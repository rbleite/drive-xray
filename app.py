#!/usr/bin/env python3
"""Streamlit UI for drive-xray.

Run with:  streamlit run app.py
"""
from __future__ import annotations

import json
import os
import shutil
import sqlite3
import subprocess
import sys
import threading
import time
from collections import defaultdict
from pathlib import Path

import streamlit as st

sys.path.insert(0, str(Path(__file__).parent))
import pandas as pd

from drive_xray import (
    open_db, fill_full_hashes, compute_dir_hashes, human,
    get_hash_version, HASH_VERSION, DX_VERSION, _duplicate_rows,
    compute_folder_sizes, generate_cleanup_script,
    CLEANUP_STRATEGIES, CLEANUP_ACTIONS,
    latest_snapshot_id, list_snapshots, diff_snapshots, resolve_root,
    registry_list, registry_remove, registry_register,
    cross_dedupe, single_copy_files, cold_data, read_drive_index_opts,
    generate_backup_script,
    verify_file, execute_file_action, QUARANTINE_DIR, AUDIT_LOG,
    read_config, write_config, get_db_dir, import_folder,
    get_exclusions, set_exclusions, SYSTEM_EXCLUDE_DIRS,
    tags_get, tags_set, tags_search, notes_get, notes_set,
    compute_auto_tags, AUTO_TAGS_YAML_PATH, write_default_auto_tag_rules,
    get_auto_tag_rules,
)


def pick_folder_dialog(initial: str | None = None) -> str | None:
    """Open the OS-native folder chooser and return the chosen path, or None if
    cancelled/unavailable. Runs on the machine hosting the app — intended for
    local desktop use (it blocks the script until the user picks).

    macOS uses AppleScript (reliable, unlike tkinter off Streamlit's worker
    thread); Windows a WinForms dialog; Linux tries zenity then kdialog.
    """
    import platform
    system = platform.system()
    use_init = bool(initial and os.path.isdir(initial))
    try:
        if system == "Darwin":
            loc = f' default location (POSIX file "{initial}")' if use_init else ""
            out = subprocess.run(
                ["osascript",
                 "-e", f'set c to choose folder with prompt "Escolher pasta"{loc}',
                 "-e", "POSIX path of c"],
                capture_output=True, text=True, timeout=180)
            return out.stdout.strip() or None
        if system == "Windows":
            ps = ("Add-Type -AssemblyName System.Windows.Forms;"
                  "$d = New-Object System.Windows.Forms.FolderBrowserDialog;"
                  + (f"$d.SelectedPath='{initial}';" if use_init else "")
                  + "if($d.ShowDialog() -eq 'OK'){$d.SelectedPath}")
            out = subprocess.run(["powershell", "-NoProfile", "-Command", ps],
                                 capture_output=True, text=True, timeout=180)
            return out.stdout.strip() or None
        for tool, argv in (("zenity", ["--file-selection", "--directory"]),
                           ("kdialog", ["--getexistingdirectory", initial or "."])):
            try:
                out = subprocess.run([tool, *argv], capture_output=True,
                                     text=True, timeout=180)
                if out.stdout.strip():
                    return out.stdout.strip()
            except FileNotFoundError:
                continue
        return None
    except Exception:
        return None

DB_DIR = get_db_dir()
SCRIPT = Path(__file__).parent / "drive_xray.py"
DB_DIR.mkdir(parents=True, exist_ok=True)


# ---------- Rust binary fallback ----------
# When the Rust `dx` binary is available it produces bit-identical .db
# files but ~10× faster. Prefer it for the long-running subprocess
# commands (index / refresh / snapshot / compact). The Python script is
# always used for the in-process helpers (drive_info, dup_file_groups,
# treemap_rows, …) because no subprocess is involved there.
def _dx_probe(cand: str) -> str | None:
    """A candidate only counts as the Rust engine if `dx --version` runs and
    identifies itself — any executable that merely happens to be called `dx`
    (fairly common name) must not be trusted with index/refresh commands.
    Returns the reported version ("1.3.0") or None when rejected."""
    try:
        out = subprocess.run([cand, "--version"], capture_output=True,
                             text=True, timeout=10)
        first = (out.stdout or "").strip().splitlines()[:1]
        if out.returncode == 0 and first and first[0].startswith("dx "):
            return first[0].split()[1]
        return None
    except Exception:
        return None


def _dx_command_prefix() -> tuple[list[str], str, str]:
    """Resolve the indexer command prefix once. Returns (cmd, reason, version)
    where reason explains a Python fallback ("" when Rust is used) and
    version is the dx binary's self-reported version ("" for Python).
    Order of preference:
      1. $DRIVE_XRAY_DX env var (explicit override)
      2. ./rust/target/{universal,release,debug}/dx adjacent to this file
      3. dx(.exe) next to app.py (the README's drop-in install on Windows)
      4. well-known install dirs — Finder/Dock launches often have a PATH
         without /opt/homebrew/bin, which made the engine flip between
         Rust and Python depending on how the app was started
      5. `dx` on PATH
      6. fall back to Python: [sys.executable, drive_xray.py]
    Every candidate must pass a `dx --version` probe before being used.
    """
    rejected: list[str] = []

    def try_cand(p) -> tuple[list[str], str] | None:
        p = str(p)
        ver = _dx_probe(p)
        if ver:
            return [p], ver
        rejected.append(p)
        return None

    env = os.environ.get("DRIVE_XRAY_DX")
    if env and Path(env).is_file() and os.access(env, os.X_OK):
        got = try_cand(env)
        if got:
            return got[0], "", got[1]
    here = Path(__file__).parent
    exe = "dx.exe" if os.name == "nt" else "dx"
    cands = [here / "rust" / "target" / "universal" / exe,
             here / "rust" / "target" / "release" / exe,
             here / "rust" / "target" / "debug" / exe,
             here / exe]
    if os.name != "nt":
        cands += [Path("/opt/homebrew/bin/dx"), Path("/usr/local/bin/dx")]
    for cand in cands:
        if cand.is_file() and os.access(cand, os.X_OK):
            got = try_cand(cand)
            if got:
                return got[0], "", got[1]
    on_path = shutil.which("dx")
    if on_path:
        got = try_cand(on_path)
        if got:
            return got[0], "", got[1]
    reason = ("rejected: " + ", ".join(rejected)) if rejected else "dx not found"
    return [sys.executable, str(SCRIPT)], reason, ""


DX_CMD, DX_FALLBACK_REASON, DX_BIN_VERSION = _dx_command_prefix()
DX_IS_RUST = len(DX_CMD) == 1  # heuristic: a single token means a binary

_CACHE_FLOOR = 1048576  # 1 MB floor for dup groups

# One lock per cache key: only one thread computes, others wait.
# session_state stores the result directly (no pickle copy overhead).
_LOCKS: dict[str, threading.Lock] = {}
_LOCKS_LOCK = threading.Lock()


def _get_lock(key: str) -> threading.Lock:
    with _LOCKS_LOCK:
        if key not in _LOCKS:
            _LOCKS[key] = threading.Lock()
        return _LOCKS[key]


def _ss_compute(key: str, fn):
    """Return session_state[key], computing via fn() if missing.
    Thread-safe: only one concurrent caller runs fn(); others wait."""
    if key in st.session_state:
        return st.session_state[key]
    lock = _get_lock(key)
    with lock:
        if key not in st.session_state:  # double-check under lock
            st.session_state[key] = fn()
    return st.session_state[key]


st.set_page_config(page_title="drive-xray", layout="wide", page_icon="💾")


# ---------- i18n ----------

TRANSLATIONS = {
    "pt": {
        "indexed_drives": "Drives indexadas",
        "no_drives": "Ainda não tens drives indexadas.",
        "delete_tooltip": "Remover esta drive do índice",
        "confirm_delete": "Confirmar remoção",
        "delete_question": "Apagar o índice **{name}**?\n\nO ficheiro `.db` e os auxiliares (`-wal`, `-shm`) vão ser removidos. **Não** afecta a drive original.",
        "yes_delete": "Sim, apagar",
        "cancel": "Cancelar",
        "index_new_drive": "Indexar nova drive",
        "path_label": "Caminho",
        "browse_btn": "📁 Procurar…",
        "browse_hint": "Abre o Finder para escolheres a pasta.",
        "label_label": "Etiqueta",
        "full_hash_label": "Hash completo (--full)",
        "full_hash_help": "Lento, mas permite comparações offline confirmadas.",
        "one_fs_label": "Apenas este filesystem (-x)",
        "one_fs_help": "Não atravessa mount points. Evita /Volumes/* e firmlinks do APFS.",
        "skip_cloud_label": "Ignorar pastas de cloud (--skip-cloud)",
        "skip_cloud_help": "Salta iCloud, OneDrive, Google Drive, Dropbox, Box, MEGA, Proton Drive, etc.",
        "skip_system_label": "Ignorar pastas de sistema",
        "skip_system_help": "Exclui {names}. Recomendado ao indexar o disco de sistema (C:\\, /); desliga se estas pastas te interessarem.",
        "index_button": "Indexar",
        "path_not_exist": "Caminho inexistente.",
        "indexing": "⏳ A indexar **{label}**…",
        "completed": "✅ {label} concluída",
        "exited_with_code": "❌ {label} saiu com código {code}",
        "clear_log": "Limpar log",
        "log": "Log",
        "no_output_yet": "(sem output ainda)",
        "main_welcome_title": "drive-xray",
        "main_welcome_body": "Indexa o teu Mac ou drives externas, encontra ficheiros e pastas duplicados, compara drives offline.\n\nComeça por **indexar uma drive** na barra lateral.",
        "no_metadata_error": "`{name}` não contém metadados de drive. Re-indexa.",
        "drive_busy": "⏳ `{name}` está ocupada — indexação/refresh em curso. Os dados estão intactos; espera que termine.",
        "drive_busy_retry": "🔄 Tentar de novo",
        "drive_indexing": "⏳ a indexar… (PID {pid})",
        "op_log_title": "📜 Log da operação",
        "op_log_refresh": "Atualizar o log",
        "engine_stale": "⚠️ O binário dx é a v{have} mas a app é a v{want} — "
                        "substitui o dx(.exe) pela release v{want}; o binário "
                        "antigo não tem as correções e funcionalidades novas.",
        "indexed_on": "indexada em",
        "tab_summary": "📊 Resumo",
        "tab_dupes": "🔁 Duplicados",
        "tab_compare": "⚖️ Comparar",
        "files": "Ficheiros",
        "folders": "Pastas",
        "total_size": "Tamanho total",
        "top_ext": "Top 20 extensões por tamanho",
        "ext_col": "extensão",
        "files_col": "ficheiros",
        "size_col": "tamanho",
        "ignore_smaller_mb": "Ignorar ficheiros menores que (MB)",
        "drive_not_mounted": "A drive `{root}` não está montada — só vão ser usados hashes já existentes na .db (índices feitos sem `--full` podem ter resultados incompletos).",
        "find_dupes": "Procurar duplicados",
        "calculating": "A calcular…",
        "load_duplicates": "Procurar duplicados",
        "confirm_expander": "🔬 Confirmar com hash completo (opcional, lento)",
        "confirm_caption": "Lê todos os ficheiros candidatos para confirmar ≈ como = exacto. Pode demorar muito em drives grandes.",
        "confirm_btn": "Confirmar duplicados (lento)",
        "confirming_candidates": "A confirmar candidatos com hash completo…",
        "files_hashed": "{n} ficheiros hashados.",
        "computing_merkle": "A calcular hashes Merkle das pastas…",
        "done": "Concluído.",
        "file_groups": "Grupos de ficheiros",
        "folder_groups": "Grupos de pastas",
        "wasted_space": "Espaço desperdiçado (ficheiros)",
        "duplicate_files": "Ficheiros duplicados",
        "sorted_top200": "Ordenado por espaço desperdiçado. Top 200.",
        "wasted": "desperdício",
        "groups_not_shown": "+ {n} grupos não mostrados",
        "duplicate_folders": "Pastas duplicadas",
        "identical_folders": "pastas idênticas",
        "need_two_drives": "Precisas de pelo menos duas drives indexadas para comparar.",
        "cross_title": "🔀 Duplicados entre todas as drives",
        "cross_caption": "Compara todos os índices em simultâneo. As drives não precisam de estar montadas.",
        "cross_btn": "Procurar duplicados entre todas as drives",
        "cross_groups": "Grupos",
        "cross_wasted": "Espaço desperdiçado",
        "cross_confirmed": "Confirmados (=)",
        "cross_approx": "Aproximados (≈)",
        "cross_no_results": "Nenhum duplicado entre drives encontrado com o tamanho mínimo seleccionado.",
        "cross_col_group": "#",
        "cross_col_drive": "drive",
        "cross_col_path": "caminho",
        "cross_col_size": "tamanho",
        "cross_col_match": "match",
        "cross_need_drives": "Precisa de pelo menos 2 drives indexadas.",
        "cross_firmlink_warn": "ℹ️ **{drives}** — têm múltiplas cópias internas do mesmo conteúdo (backups ou cópias manuais). Os grupos marcados com ⚠️ incluem essas cópias.",
        "cross_firmlink_groups": "Grupos com cópias internas duplicadas na mesma drive:",
        "cross_matrix_title": "📊 Sobreposição entre drives",
        "cross_matrix_caption": "GB partilhados entre cada par de drives. Quanto mais escuro, mais duplicados.",
        "cross_download_csv": "⬇️ Exportar CSV (todos os grupos)",
        "cross_details": "📋 Detalhes (top 500 grupos)",
        # single-copy (no backup) — inverse of cross-dedupe
        "sc_title": "🛟 Ficheiros sem cópia de segurança",
        "sc_caption": "O inverso: conteúdo que existe em apenas UMA drive. Se essa drive falhar, perde-o para sempre. Cópias internas na mesma drive não contam como backup.",
        "sc_need_drives": "Precisa de pelo menos 2 drives indexadas com dados para comparar.",
        "sc_insufficient": "⚠️ Apenas uma drive tinha dados ({drives}). Sem uma segunda drive para comparar, o resultado seria enganador (tudo pareceria sem cópia). Reindexe as drives vazias/stub.",
        "sc_scope": "Drive a analisar",
        "sc_scope_all": "Todas",
        "sc_btn": "Procurar ficheiros sem cópia",
        "sc_no_results": "✅ Nada em risco — todo o conteúdo (acima do tamanho mínimo) existe em pelo menos duas drives.",
        "sc_metric_items": "Itens sem cópia",
        "sc_metric_bytes": "Total em risco",
        "sc_metric_drives": "Drives comparadas",
        "sc_per_drive": "Por drive",
        "sc_by_folder": "📁 Pastas sem cópia (top 40)",
        "sc_col_drive": "drive",
        "sc_col_folder": "pasta",
        "sc_col_path": "caminho",
        "sc_col_bytes": "tamanho",
        "sc_col_count": "nº itens",
        "sc_col_copies": "cópias internas",
        "sc_files_title": "📋 Ficheiros sem cópia (top {n} por tamanho)",
        "sc_download_csv": "⬇️ Exportar CSV (todos os ficheiros em risco)",
        "bkp_title": "💾 Gerar script de backup",
        "bkp_caption": "Cria um script que copia estes ficheiros sem cópia para uma drive-alvo (por drive de origem). NÃO executa nada — revês e corres tu. As drives de origem têm de estar montadas.",
        "bkp_target": "Pasta de destino",
        "bkp_shell": "Formato",
        "bkp_download": "⬇️ Descarregar script de backup",
        "bkp_preview": "Ver script",
        "sc_truncated": "Tabela mostra os maiores {shown} de {total} itens; o CSV inclui todos.",
        "compare_with": "Comparar com",
        "minimum_mb": "Mínimo (MB)",
        "compare_button": "Comparar",
        "crosschecking": "A cruzar índices…",
        "matches": "Matches",
        "confirmed_eq": "Confirmados (=)",
        "only_in_a": "Apenas em A",
        "matching_size": "Tamanho coincidente",
        "matches_not_shown": "+ {n} matches não mostrados",
        "in_drive": "em {label}",
        "match_col": "match",
        "size_match_col": "tamanho",
        "hardlink_tag": "hardlink",
        "hash_version_mismatch": "⚠️ Versões de partial-hash diferentes (A=v{va}, B=v{vb}). Os matches podem não ser fiáveis — re-indexa ambas as drives com a versão actual (v{cur}).",
        "refresh_tooltip": "Re-indexar incrementalmente (reutiliza hashes de ficheiros inalterados)",
        "download_csv": "⬇️ Exportar CSV",
        "download_xlsx": "⬇️ Exportar Excel (XLSX)",
        "db_size": "Tamanho .db",
        "compact_button": "🧹 Compactar",
        "verify_title": "🔬 Verificar integridade (bit-rot)",
        "verify_caption": "Relê os ficheiros e compara o hash com o guardado — deteta corrupção silenciosa (mesmo tamanho e data, conteúdo diferente). Precisa da drive montada.",
        "verify_full": "Verificação profunda (ficheiro inteiro)",
        "verify_full_help": "Mais lento e completo; só para ficheiros com hash completo guardado. Por defeito usa o hash parcial (rápido).",
        "verify_btn": "🔬 Verificar agora",
        "verify_running": "A verificar…",
        "verify_corrupt_short": "corrompidos",
        "verify_changed": "Tamanho mudou",
        "verify_missing": "Em falta",
        "verify_unmounted": "⚠️ Drive não montada em `{root}` — não dá para verificar. Monta o disco.",
        "verify_rot_found": "🚨 {n} ficheiro(s) com BIT-ROT — conteúdo alterado com o mesmo tamanho+data. NÃO os copies por cima dos backups bons.",
        "verify_clean": "✅ Sem corrupção — todos os ficheiros batem certo com o hash guardado.",
        "compact_help": "VACUUM + checkpoint WAL para libertar espaço. Sem perda de dados.",
        # treemap
        "tab_map": "🗺️ Mapa",
        "map_caption": "Treemap de utilização de disco. Cada rectângulo é uma pasta; o tamanho é proporcional ao espaço ocupado. Clica para entrar.",
        "map_min_mb": "Tamanho mínimo (MB)",
        "map_include_files": "Incluir ficheiros individuais (não só pastas)",
        "map_empty": "Sem pastas acima do tamanho mínimo. Baixa o threshold.",
        "map_legend": "A mostrar {n} elementos.",
        # tags
        "tags_expander": "🏷️ Etiquetar pastas",
        "tags_caption": "Associa etiquetas livres a pastas para identificação rápida. Visíveis no tooltip do mapa.",
        "tags_select": "Pasta (do mapa actual)",
        "tags_input": "Etiquetas (separadas por vírgula)",
        "tags_save": "Guardar",
        "tags_remove_btn": "Remover etiquetas",
        "tags_saved": "Etiquetas guardadas.",
        "tags_removed": "Etiquetas removidas.",
        "tags_active": "Pastas etiquetadas nesta drive",
        "tags_none": "Nenhuma pasta etiquetada ainda.",
        "tags_col_path": "Pasta",
        "tags_col_tags": "Etiquetas",
        "tags_filter": "🔍 Filtrar por etiqueta ou pasta",
        "tags_filter_empty": "Nenhuma etiqueta corresponde ao filtro.",
        "tags_note_label": "Nota (texto livre)",
        "tags_note_placeholder": "Backup de 2023-03, já verificado. Contém runs STP e ZP.",
        "tags_note_save": "Guardar nota",
        "tags_col_note": "Nota",
        "tags_legend": "Legenda de cores",
        "auto_tags_detected": "Detetado automaticamente",
        "auto_tags_promote": "⬆️ Usar como tags manuais",
        "auto_tags_legend_note": "pastas com tags automáticas (sem tag manual)",
        "at_rules_title": "⚙️ Regras de auto-tag ({n} regras)",
        "at_rules_src_default": "A usar as regras predefinidas (embutidas).",
        "at_rules_src_custom": "A usar o teu ficheiro de regras personalizado.",
        "at_rules_init_btn": "Criar ficheiro editável",
        "at_rules_edit_hint": "Edita este ficheiro (formato `tag: [ext1, ext2]`); as alterações aplicam-se de imediato.",
        "at_rules_path": "Ficheiro de regras",
        "at_rules_created": "Ficheiro criado — edita-o para personalizar as regras.",
        # cold data (archive candidates) — Map tab
        "upd_title": "🔄 Atualizações (GitHub)",
        "upd_check": "Verificar atualizações",
        "upd_uptodate": "✅ Já estás na versão mais recente.",
        "upd_available": "🆕 {n} atualização(ões) disponível(eis):",
        "upd_apply": "⬇️ Atualizar agora",
        "upd_applying": "A atualizar…",
        "excl_title": "🚫 Excluir pastas da indexação",
        "excl_caption": "Pastas que NÃO queres indexar. Escreve um **nome** (`node_modules`) para excluir em qualquer nível, um **padrão** (`*cache*`, `*Extras*`), ou um **caminho** (`Series/Extras`) a partir da raiz. Aplicam-se no próximo refresh.",
        "excl_add": "Escolher pasta a excluir",
        "excl_or_type": "…ou nome/padrão",
        "excl_add_btn": "➕ Excluir",
        "excl_added": "Exclusão adicionada — faz refresh para aplicar.",
        "excl_system_btn": "🖥️ Adicionar pastas de sistema",
        "excl_system_help": "Adiciona de uma vez: {names}. Essencial ao indexar o disco de sistema (C:\\, /) — corta a maior parte dos ficheiros irrelevantes.",
        "excl_system_added": "{n} exclusões de sistema adicionadas — faz refresh para aplicar.",
        "excl_current": "Excluídas nesta drive:",
        "excl_refresh_hint": "🔄 Faz refresh da drive para aplicar as exclusões.",
        "excl_none": "Sem exclusões nesta drive.",
        "cold_title": "❄️ Dados frios (candidatos a arquivo)",
        "cold_badge": "dados frios",
        "cold_map_hint": "❄️ As pastas frias estão realçadas a azul-gélido no TreeMap acima.",
        "cold_caption": "Pastas cujo ficheiro mais recente é anterior ao limite — candidatas a arquivo/cold storage. Mostra o topo de cada subárvore fria.",
        "cold_years": "Mais antigo que (anos)",
        "cold_btn": "Procurar dados frios",
        "cold_none": "✅ Nenhuma pasta fria acima do tamanho mínimo com este limite.",
        "cold_metric_folders": "Pastas frias",
        "cold_metric_bytes": "Total arquivável",
        "cold_metric_cutoff": "Anterior a",
        "cold_col_folder": "pasta",
        "cold_col_size": "tamanho",
        "cold_col_newest": "ficheiro mais recente",
        "cold_col_files": "nº ficheiros",
        "cold_download_csv": "⬇️ Exportar CSV (todas as pastas frias)",
        "cold_truncated": "Tabela mostra as maiores {shown} de {total} pastas; o CSV inclui todas.",
        # cross-drive tag search (Compare tab)
        "tag_search_title": "🔍 Pesquisar por etiqueta",
        "tag_search_caption": "Procura em todas as drives indexadas. Corresponde à etiqueta, ao caminho ou à nota.",
        "tag_search_input": "Etiqueta, caminho ou nota",
        "tag_search_col_drive": "Drive",
        "tag_search_col_path": "Pasta",
        "tag_search_col_tags": "Etiquetas",
        "tag_search_col_note": "Nota",
        "tag_search_empty": "Nenhuma pasta etiquetada encontrada.",
        "tag_search_no_tags": "Ainda não existem etiquetas em nenhuma drive. Usa o tab Mapa para etiquetar pastas.",
        # cleanup
        "cleanup_title": "🧽 Assistente de limpeza",
        "cleanup_caption": "Gera um script shell com as remoções/movimentos sugeridos. **Não apaga nada automaticamente** — revê e corre manualmente.",
        "cleanup_strategy": "Qual cópia manter",
        "cleanup_action": "Acção para as outras",
        "strategy_shortest": "Caminho mais curto (recomendado)",
        "strategy_oldest": "Mais antiga (mtime)",
        "strategy_newest": "Mais recente (mtime)",
        "strategy_alphabetical": "Alfabética",
        "action_quarantine": "Mover para quarentena (~/.drive-xray-quarantine/)",
        "action_delete": "Apagar (rm — irreversível!)",
        "cleanup_generate": "Gerar plano",
        "cleanup_ready": "✅ Plano gerado: {n} acções propostas.",
        "cleanup_download": "⬇️ Descarregar script .sh",
        "cleanup_preview": "Pré-visualizar script",
        # interactive delete
        "del_title": "🗑️ Apagar seleccionados",
        "del_caption": "Selecciona as cópias a eliminar (por defeito: mantém o caminho mais curto). Antes de apagar, verifica-se que a cópia a manter ainda existe em disco.",
        "del_col": "apagar",
        "del_action_label": "Acção",
        "del_verify_btn": "Verificar e apagar {n} ficheiro(s) · {size}",
        "del_none_selected": "Marca pelo menos um ficheiro para apagar.",
        "del_root_unmounted": "⚠️ Drive não montada — não é possível verificar nem apagar ficheiros.",
        "del_dialog_title": "Verificação antes de apagar",
        "del_ok_summary": "✅ {n} grupo(s) verificados · {f} ficheiro(s) · {size} a libertar",
        "del_no_keeper": "🚫 {n} grupo(s) ignorados — todas as cópias marcadas (ficaria sem nenhuma)",
        "del_keeper_missing": "⚠️ {n} grupo(s) ignorados — cópia a manter não encontrada em disco",
        "del_already_gone": "ℹ️ {n} ficheiro(s) já não existem — serão ignorados",
        "del_nothing_to_do": "Nenhum ficheiro verificado para apagar.",
        "del_detail": "Ver detalhes por grupo",
        "del_keeper_label": "manter",
        "del_execute_btn": "Executar · {n} ficheiro(s) · {size}",
        "del_done": "✅ {n} ficheiro(s) processados · {size} libertados",
        "del_errors": "❌ {n} erro(s)",
        # info / rename
        "info_tooltip": "Informação e renomear",
        "drive_info_title": "Informação da drive",
        "drive_db_path": "Ficheiro .db",
        "drive_root": "Raiz",
        "drive_last_indexed": "Última indexação",
        "rename_label": "Renomear etiqueta",
        "rename_save": "Guardar",
        # snapshots / history
        "snapshot_tooltip": "Tirar novo snapshot (preserva o histórico anterior)",
        "snapshot_button": "📸 Tirar snapshot",
        "tab_history": "📅 Histórico",
        "history_no_snaps": "Esta drive ainda não tem snapshots. Indexa-a primeiro.",
        "history_title": "{n} snapshots",
        "history_taken_at": "data",
        "diff_title": "Comparar snapshots",
        "diff_from": "De",
        "diff_to": "Até",
        "diff_same": "Escolhe snapshots diferentes para o diff.",
        "diff_compute": "Calcular diff",
        "diff_added": "Adicionados",
        "diff_removed": "Removidos",
        "diff_modified": "Modificados",
        "diff_net": "Variação líquida",
        "diff_top_growth": "Top pastas por crescimento",
        "diff_top_shrink": "Top pastas por redução",
        # settings / sync
        "settings_title": "⚙️ Configurações",
        "settings_db_dir": "Pasta dos índices (.db)",
        "settings_db_dir_help": "Pasta onde são guardados os ficheiros .db. Usa uma pasta do OneDrive/Google Drive/Dropbox para partilha automática entre máquinas.",
        "settings_save": "Guardar",
        "settings_saved": "✅ Pasta actualizada. A reiniciar…",
        "settings_invalid": "Pasta inválida ou sem permissão de escrita.",
        "settings_import_btn": "Importar .db desta pasta",
        "settings_import_help": "Regista todos os ficheiros .db encontrados na pasta configurada.",
        "settings_imported": "✅ {n} drive(s) importada(s) · {s} já existiam",
        "settings_import_none": "Nenhum .db válido encontrado nessa pasta.",
        "settings_current_dir": "Pasta actual",
    },
    "en": {
        "indexed_drives": "Indexed drives",
        "no_drives": "No drives indexed yet.",
        "delete_tooltip": "Remove this drive's index",
        "confirm_delete": "Confirm removal",
        "delete_question": "Delete index **{name}**?\n\nThe `.db` file and its sidecars (`-wal`, `-shm`) will be removed. This does **not** affect the original drive.",
        "yes_delete": "Yes, delete",
        "cancel": "Cancel",
        "index_new_drive": "Index new drive",
        "path_label": "Path",
        "browse_btn": "📁 Browse…",
        "browse_hint": "Opens Finder to pick the folder.",
        "label_label": "Label",
        "full_hash_label": "Full hash (--full)",
        "full_hash_help": "Slow, but enables confirmed offline comparisons.",
        "one_fs_label": "Single filesystem only (-x)",
        "one_fs_help": "Does not cross mount points. Avoids /Volumes/* and APFS firmlinks.",
        "skip_cloud_label": "Skip cloud folders (--skip-cloud)",
        "skip_cloud_help": "Skips iCloud, OneDrive, Google Drive, Dropbox, Box, MEGA, Proton Drive, etc.",
        "skip_system_label": "Ignore system folders",
        "skip_system_help": "Excludes {names}. Recommended when indexing the system disk (C:\\, /); untick if you want those folders indexed.",
        "index_button": "Index",
        "path_not_exist": "Path does not exist.",
        "indexing": "⏳ Indexing **{label}**…",
        "completed": "✅ {label} completed",
        "exited_with_code": "❌ {label} exited with code {code}",
        "clear_log": "Clear log",
        "log": "Log",
        "no_output_yet": "(no output yet)",
        "main_welcome_title": "drive-xray",
        "main_welcome_body": "Index your Mac or external drives, find duplicate files and folders, compare drives offline.\n\nStart by **indexing a drive** in the sidebar.",
        "no_metadata_error": "`{name}` has no drive metadata. Please re-index.",
        "drive_busy": "⏳ `{name}` is busy — an index/refresh is in progress. Your data is intact; wait for it to finish.",
        "drive_busy_retry": "🔄 Try again",
        "drive_indexing": "⏳ indexing… (PID {pid})",
        "op_log_title": "📜 Operation log",
        "op_log_refresh": "Refresh the log",
        "engine_stale": "⚠️ The dx binary is v{have} but the app is v{want} — "
                        "replace dx(.exe) with the v{want} release build; the "
                        "old binary is missing new fixes and features.",
        "indexed_on": "indexed on",
        "tab_summary": "📊 Summary",
        "tab_dupes": "🔁 Duplicates",
        "tab_compare": "⚖️ Compare",
        "files": "Files",
        "folders": "Folders",
        "total_size": "Total size",
        "top_ext": "Top 20 extensions by size",
        "ext_col": "extension",
        "files_col": "files",
        "size_col": "size",
        "ignore_smaller_mb": "Ignore files smaller than (MB)",
        "drive_not_mounted": "Drive `{root}` is not mounted — only hashes already in the .db will be used (indexes built without `--full` may have incomplete results).",
        "find_dupes": "Find duplicates",
        "calculating": "Calculating…",
        "load_duplicates": "Find duplicates",
        "confirm_expander": "🔬 Confirm with full hash (optional, slow)",
        "confirm_caption": "Reads every candidate file to upgrade ≈ matches to exact =. Can take a long time on large drives.",
        "confirm_btn": "Confirm duplicates (slow)",
        "confirming_candidates": "Confirming candidates with full hashes…",
        "files_hashed": "{n} files hashed.",
        "computing_merkle": "Computing folder Merkle hashes…",
        "done": "Done.",
        "file_groups": "File groups",
        "folder_groups": "Folder groups",
        "wasted_space": "Wasted space (files)",
        "duplicate_files": "Duplicate files",
        "sorted_top200": "Sorted by wasted space. Top 200.",
        "wasted": "wasted",
        "groups_not_shown": "+ {n} groups not shown",
        "duplicate_folders": "Duplicate folders",
        "identical_folders": "identical folders",
        "need_two_drives": "You need at least two indexed drives to compare.",
        "cross_title": "🔀 Duplicates across all drives",
        "cross_caption": "Compares all indexes at once. Drives do not need to be mounted.",
        "cross_btn": "Find duplicates across all drives",
        "cross_groups": "Groups",
        "cross_wasted": "Wasted space",
        "cross_confirmed": "Confirmed (=)",
        "cross_approx": "Approximate (≈)",
        "cross_no_results": "No cross-drive duplicates found at the selected minimum size.",
        "cross_col_group": "#",
        "cross_col_drive": "drive",
        "cross_col_path": "path",
        "cross_col_size": "size",
        "cross_col_match": "match",
        "cross_need_drives": "Need at least 2 indexed drives.",
        "cross_firmlink_warn": "ℹ️ **{drives}** — have multiple internal copies of the same content (backups or manual copies). Groups marked ⚠️ include these copies.",
        "cross_firmlink_groups": "Groups with internal duplicate copies on the same drive:",
        "cross_matrix_title": "📊 Drive overlap",
        "cross_matrix_caption": "Shared GB between each pair of drives. Darker = more duplicates.",
        "cross_download_csv": "⬇️ Export CSV (all groups)",
        "cross_details": "📋 Details (top 500 groups)",
        # single-copy (no backup) — inverse of cross-dedupe
        "sc_title": "🛟 Files with no backup",
        "sc_caption": "The inverse: content that lives on only ONE drive. If that drive dies, it's gone for good. Internal copies on the same drive do not count as a backup.",
        "sc_need_drives": "Need at least 2 indexed drives with data to compare.",
        "sc_insufficient": "⚠️ Only one drive had data ({drives}). Without a second drive to compare against, the result would be misleading (everything would look un-backed-up). Re-index the empty/stub drives.",
        "sc_scope": "Drive to analyse",
        "sc_scope_all": "All",
        "sc_btn": "Find files with no backup",
        "sc_no_results": "✅ Nothing at risk — all content (above the minimum size) exists on at least two drives.",
        "sc_metric_items": "Un-backed items",
        "sc_metric_bytes": "Total at risk",
        "sc_metric_drives": "Drives compared",
        "sc_per_drive": "Per drive",
        "sc_by_folder": "📁 Folders with no backup (top 40)",
        "sc_col_drive": "drive",
        "sc_col_folder": "folder",
        "sc_col_path": "path",
        "sc_col_bytes": "size",
        "sc_col_count": "items",
        "sc_col_copies": "internal copies",
        "sc_files_title": "📋 Files with no backup (top {n} by size)",
        "sc_download_csv": "⬇️ Export CSV (all at-risk files)",
        "bkp_title": "💾 Generate backup script",
        "bkp_caption": "Builds a script that copies these un-backed-up files to a target drive (per source drive). It does NOT run anything — review and run it yourself. Source drives must be mounted.",
        "bkp_target": "Target folder",
        "bkp_shell": "Format",
        "bkp_download": "⬇️ Download backup script",
        "bkp_preview": "Preview script",
        "sc_truncated": "Table shows the largest {shown} of {total} items; the CSV has them all.",
        "compare_with": "Compare with",
        "minimum_mb": "Minimum (MB)",
        "compare_button": "Compare",
        "crosschecking": "Cross-checking indexes…",
        "matches": "Matches",
        "confirmed_eq": "Confirmed (=)",
        "only_in_a": "Only in A",
        "matching_size": "Matching size",
        "matches_not_shown": "+ {n} matches not shown",
        "in_drive": "in {label}",
        "match_col": "match",
        "size_match_col": "size",
        "hardlink_tag": "hardlink",
        "hash_version_mismatch": "⚠️ Partial-hash versions differ (A=v{va}, B=v{vb}). Matches may be unreliable — re-index both drives with the current version (v{cur}).",
        "refresh_tooltip": "Refresh incrementally (reuses hashes of unchanged files)",
        "download_csv": "⬇️ Export CSV",
        "download_xlsx": "⬇️ Export Excel (XLSX)",
        "db_size": ".db size",
        "compact_button": "🧹 Compact",
        "verify_title": "🔬 Verify integrity (bit-rot)",
        "verify_caption": "Re-reads files and compares hashes to the stored ones — catches silent corruption (same size+date, different content). Drive must be mounted.",
        "verify_full": "Deep verify (whole file)",
        "verify_full_help": "Slower and thorough; only for files with a stored full hash. Defaults to the fast partial hash.",
        "verify_btn": "🔬 Verify now",
        "verify_running": "Verifying…",
        "verify_corrupt_short": "corrupted",
        "verify_changed": "Size changed",
        "verify_missing": "Missing",
        "verify_unmounted": "⚠️ Drive not mounted at `{root}` — cannot verify. Mount it.",
        "verify_rot_found": "🚨 {n} file(s) with BIT-ROT — content changed with the same size+date. Do NOT copy these over your good backups.",
        "verify_clean": "✅ No corruption — every file matches its stored hash.",
        "compact_help": "VACUUM + WAL checkpoint to reclaim space. No data loss.",
        # treemap
        "tab_map": "🗺️ Map",
        "map_caption": "Disk usage treemap. Each rectangle is a folder; size is proportional to space used. Click to drill in.",
        "map_min_mb": "Minimum size (MB)",
        "map_include_files": "Include individual files (not just folders)",
        "map_empty": "No folders above the minimum size. Lower the threshold.",
        "map_legend": "Showing {n} items.",
        # tags
        "tags_expander": "🏷️ Tag folders",
        "tags_caption": "Attach free-form tags to folders for quick identification. Visible in map tooltips.",
        "tags_select": "Folder (from current map)",
        "tags_input": "Tags (comma-separated)",
        "tags_save": "Save",
        "tags_remove_btn": "Remove tags",
        "tags_saved": "Tags saved.",
        "tags_removed": "Tags removed.",
        "tags_active": "Tagged folders in this drive",
        "tags_none": "No folders tagged yet.",
        "tags_col_path": "Folder",
        "tags_col_tags": "Tags",
        "tags_filter": "🔍 Filter by tag or path",
        "tags_filter_empty": "No tags match the filter.",
        "tags_note_label": "Note (free text)",
        "tags_note_placeholder": "Backup from 2023-03, verified. Contains STP and ZP runs.",
        "tags_note_save": "Save note",
        "tags_col_note": "Note",
        "tags_legend": "Color legend",
        "auto_tags_detected": "Auto-detected",
        "auto_tags_promote": "⬆️ Use as manual tags",
        "auto_tags_legend_note": "folders with auto-detected tags (no manual tag)",
        "at_rules_title": "⚙️ Auto-tag rules ({n} rules)",
        "at_rules_src_default": "Using the built-in default rules.",
        "at_rules_src_custom": "Using your custom rules file.",
        "at_rules_init_btn": "Create editable file",
        "at_rules_edit_hint": "Edit this file (format `tag: [ext1, ext2]`); changes apply immediately.",
        "at_rules_path": "Rules file",
        "at_rules_created": "File created — edit it to customise the rules.",
        # cold data (archive candidates) — Map tab
        "upd_title": "🔄 Updates (GitHub)",
        "upd_check": "Check for updates",
        "upd_uptodate": "✅ You're on the latest version.",
        "upd_available": "🆕 {n} update(s) available:",
        "upd_apply": "⬇️ Update now",
        "upd_applying": "Updating…",
        "excl_title": "🚫 Exclude folders from indexing",
        "excl_caption": "Folders you do NOT want indexed. Type a **name** (`node_modules`) to exclude at any depth, a **pattern** (`*cache*`, `*Extras*`), or a **path** (`Series/Extras`) from the root. Applied on the next refresh.",
        "excl_add": "Pick a folder to exclude",
        "excl_or_type": "…or name/pattern",
        "excl_add_btn": "➕ Exclude",
        "excl_added": "Exclusion added — refresh to apply.",
        "excl_system_btn": "🖥️ Add system folders",
        "excl_system_help": "Adds in one click: {names}. Essential when indexing the system disk (C:\\, /) — cuts most of the irrelevant files.",
        "excl_system_added": "{n} system exclusions added — refresh to apply.",
        "excl_current": "Excluded on this drive:",
        "excl_refresh_hint": "🔄 Refresh the drive to apply exclusions.",
        "excl_none": "No exclusions on this drive.",
        "cold_title": "❄️ Cold data (archive candidates)",
        "cold_badge": "cold data",
        "cold_map_hint": "❄️ Cold folders are highlighted in icy blue on the TreeMap above.",
        "cold_caption": "Folders whose newest file predates the cutoff — candidates for archival/cold storage. Shows the top of each cold subtree.",
        "cold_years": "Older than (years)",
        "cold_btn": "Find cold data",
        "cold_none": "✅ No cold folders above the minimum size at this cutoff.",
        "cold_metric_folders": "Cold folders",
        "cold_metric_bytes": "Total archivable",
        "cold_metric_cutoff": "Older than",
        "cold_col_folder": "folder",
        "cold_col_size": "size",
        "cold_col_newest": "newest file",
        "cold_col_files": "files",
        "cold_download_csv": "⬇️ Export CSV (all cold folders)",
        "cold_truncated": "Table shows the largest {shown} of {total} folders; the CSV has them all.",
        # cross-drive tag search (Compare tab)
        "tag_search_title": "🔍 Search by tag",
        "tag_search_caption": "Searches all indexed drives. Matches tag name, folder path, or note.",
        "tag_search_input": "Tag, path or note",
        "tag_search_col_drive": "Drive",
        "tag_search_col_path": "Folder",
        "tag_search_col_tags": "Tags",
        "tag_search_col_note": "Note",
        "tag_search_empty": "No tagged folders found.",
        "tag_search_no_tags": "No tags exist on any drive yet. Use the Map tab to tag folders.",
        # cleanup
        "cleanup_title": "🧽 Cleanup assistant",
        "cleanup_caption": "Generates a shell script of suggested removals/moves. **It does NOT delete anything automatically** — review and run manually.",
        "cleanup_strategy": "Which copy to keep",
        "cleanup_action": "What to do with the others",
        "strategy_shortest": "Shortest path (recommended)",
        "strategy_oldest": "Oldest (mtime)",
        "strategy_newest": "Newest (mtime)",
        "strategy_alphabetical": "Alphabetical",
        "action_quarantine": "Move to quarantine (~/.drive-xray-quarantine/)",
        "action_delete": "Delete (rm — irreversible!)",
        "cleanup_generate": "Generate plan",
        "cleanup_ready": "✅ Plan generated: {n} proposed actions.",
        "cleanup_download": "⬇️ Download .sh script",
        "cleanup_preview": "Preview script",
        # interactive delete
        "del_title": "🗑️ Delete selected",
        "del_caption": "Select which copies to remove (default: keep shortest path). Before deleting, the tool verifies the copy to keep still exists on disk.",
        "del_col": "delete",
        "del_action_label": "Action",
        "del_verify_btn": "Verify and delete {n} file(s) · {size}",
        "del_none_selected": "Mark at least one file for deletion.",
        "del_root_unmounted": "⚠️ Drive not mounted — cannot verify or delete files.",
        "del_dialog_title": "Pre-delete verification",
        "del_ok_summary": "✅ {n} group(s) verified · {f} file(s) · {size} to free",
        "del_no_keeper": "🚫 {n} group(s) skipped — all copies marked (would leave zero copies)",
        "del_keeper_missing": "⚠️ {n} group(s) skipped — keeper not found on disk",
        "del_already_gone": "ℹ️ {n} file(s) already gone — will be skipped",
        "del_nothing_to_do": "No files verified for deletion.",
        "del_detail": "Show details per group",
        "del_keeper_label": "keep",
        "del_execute_btn": "Execute · {n} file(s) · {size}",
        "del_done": "✅ {n} file(s) processed · {size} freed",
        "del_errors": "❌ {n} error(s)",
        # info / rename
        "info_tooltip": "Info & rename",
        "drive_info_title": "Drive info",
        "drive_db_path": ".db file",
        "drive_root": "Root",
        "drive_last_indexed": "Last indexed",
        "rename_label": "Rename label",
        "rename_save": "Save",
        # snapshots / history
        "snapshot_tooltip": "Take new snapshot (preserves previous history)",
        "snapshot_button": "📸 Take snapshot",
        "tab_history": "📅 History",
        "history_no_snaps": "This drive has no snapshots yet. Index it first.",
        "history_title": "{n} snapshots",
        "history_taken_at": "taken at",
        "diff_title": "Compare snapshots",
        "diff_from": "From",
        "diff_to": "To",
        "diff_same": "Pick different snapshots for the diff.",
        "diff_compute": "Compute diff",
        "diff_added": "Added",
        "diff_removed": "Removed",
        "diff_modified": "Modified",
        "diff_net": "Net change",
        "diff_top_growth": "Top folders by growth",
        "diff_top_shrink": "Top folders by shrink",
        # settings / sync
        "settings_title": "⚙️ Settings",
        "settings_db_dir": "Index folder (.db files)",
        "settings_db_dir_help": "Folder where .db files are saved. Point to a OneDrive/Google Drive/Dropbox folder for automatic multi-machine sync.",
        "settings_save": "Save",
        "settings_saved": "✅ Folder updated. Restarting…",
        "settings_invalid": "Invalid folder or no write permission.",
        "settings_import_btn": "Import .db files from this folder",
        "settings_import_help": "Register all valid .db files found in the configured folder.",
        "settings_imported": "✅ {n} drive(s) imported · {s} already existed",
        "settings_import_none": "No valid .db files found in that folder.",
        "settings_current_dir": "Current folder",
    },
}


def t(key: str, **fmt) -> str:
    lang = st.session_state.get("lang", "pt")
    s = TRANSLATIONS.get(lang, TRANSLATIONS["pt"]).get(key, key)
    return s.format(**fmt) if fmt else s


# ---------- subprocess: indexer ----------

# Local staging for .db files that live inside a cloud-synced folder — see
# stage_for_write/finalize_staged in drive_xray.py. Imported there so the
# logic is unit-testable without Streamlit.
from drive_xray import STAGING_DIR as _STAGING_DIR
from drive_xray import finalize_staged as _finalize_staged
from drive_xray import stage_for_write as _stage_for_write


def _log_path(db: Path) -> Path:
    """Persistent log of the last index/refresh/snapshot of this drive —
    lives next to the .db so it survives closed tabs and app restarts."""
    return Path(str(db) + ".log")


def _log_tail(db: Path, n: int = 20) -> list[str]:
    try:
        lines = _log_path(db).read_text(
            encoding="utf-8", errors="replace").splitlines()
        return [l for l in lines if l.strip()][-n:]
    except OSError:
        return []


def _spawn(cmd: list[str], label: str, db: Path,
           staged: Path | None = None) -> None:
    """Launch a long-running CLI subprocess. Output streams into
    st.session_state.idx_log AND into <db>.log, so progress can still be
    followed after the browser tab (and its session log) is gone.
    When `staged` is set, the indexer is writing a local staging copy; on
    success it is moved back over `db` (one cloud upload instead of many)."""
    try:
        lf = open(_log_path(db), "w", encoding="utf-8", errors="replace")
        lf.write("$ " + " ".join(cmd) + "\n")
        if staged:
            lf.write(f"[staging: writing {staged}, will move to {db} when done]\n")
        lf.flush()
    except OSError:
        lf = None
    proc = subprocess.Popen(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        bufsize=1, text=True,
    )
    log: list[str] = []

    def _logline(piece: str) -> None:
        log.append(piece)
        if lf:
            try:
                lf.write(piece + "\n")
                lf.flush()
            except OSError:
                pass

    def reader():
        for line in iter(proc.stdout.readline, ""):
            for piece in line.replace("\r", "\n").splitlines():
                if piece.strip():
                    _logline(piece)
        proc.stdout.close()
        code = proc.wait()
        if staged and code == 0:
            _logline(f"[staging: {_finalize_staged(staged, db)}]")
        elif staged:
            _logline(f"[staging: kept at {staged} for resume (exit {code})]")
        if lf:
            try:
                lf.write(f"[exit {code}]\n")
                lf.close()
            except OSError:
                pass

    threading.Thread(target=reader, daemon=True).start()
    st.session_state.idx_proc = proc
    st.session_state.idx_log = log
    st.session_state.idx_label = label


def start_indexer(root: str, label: str, do_full: bool,
                  one_fs: bool, skip_cloud: bool) -> None:
    db_out = DB_DIR / f"{label}.db"
    target, staged = _stage_for_write(db_out)
    cmd = [*DX_CMD, "index", root, "--label", label, "--db", str(target)]
    if do_full:
        cmd.append("--full")
    if one_fs:
        cmd.append("--one-filesystem")
    if skip_cloud:
        cmd.append("--skip-cloud")
    _spawn(cmd, label, db_out, staged=target if staged else None)


def start_refresh(db: Path) -> None:
    target, staged = _stage_for_write(db)
    _spawn([*DX_CMD, "refresh", str(target)], db.stem, db,
           staged=target if staged else None)


def start_compact(db: Path) -> None:
    target, staged = _stage_for_write(db)
    _spawn([*DX_CMD, "compact", str(target)], db.stem, db,
           staged=target if staged else None)


def start_snapshot(db: Path) -> None:
    target, staged = _stage_for_write(db)
    _spawn([*DX_CMD, "snapshot", "take", str(target)], db.stem, db,
           staged=target if staged else None)


@st.cache_data(ttl=3, show_spinner=False)
def _active_index_procs() -> dict:
    """Map resolved-db-path -> pid for any live `dx` index/refresh/snapshot/
    compact subprocess. Reflects real OS state — survives reruns, hot-reloads,
    error pages and app restarts, unlike the session-scoped `idx_proc`. This is
    what prevents a second refresh from being launched over a running one.
    Cached for 3s so we don't run `ps` on every micro-rerun."""
    try:
        if os.name == "nt":
            # no `ps` on Windows — same "pid commandline" shape via CIM
            out = subprocess.run(
                ["powershell", "-NoProfile", "-Command",
                 "Get-CimInstance Win32_Process | ForEach-Object"
                 " { '{0} {1}' -f $_.ProcessId, $_.CommandLine }"],
                capture_output=True, text=True, timeout=15,
            ).stdout
        else:
            out = subprocess.run(
                ["ps", "-Ao", "pid=,command="],
                capture_output=True, text=True, timeout=5,
            ).stdout
    except Exception:
        return {}
    procs: dict[str, int] = {}
    subs = ("refresh", "index", "snapshot", "compact")
    for line in out.splitlines():
        line = line.strip()
        if not line or " <defunct>" in line:
            continue  # skip zombies — they hold no lock
        if not any(f" {s}" in line for s in subs):
            continue
        parts = line.split()
        try:
            pid = int(parts[0])
        except (ValueError, IndexError):
            continue
        for tok in parts[1:]:
            if tok.endswith(".db"):
                try:
                    procs[str(Path(tok).resolve())] = pid
                except Exception:
                    procs[tok] = pid
                break
    return procs


def _busy_pid_for(db: Path, procs: dict) -> int | None:
    """PID of a live index/refresh writer for this drive — whether it is
    writing the db in place or its local staging copy (cloud-synced dbs)."""
    return (procs.get(str(db.resolve()))
            or procs.get(str((_STAGING_DIR / db.name).resolve())))


# ---------- db helpers ----------

def list_dbs() -> list[Path]:
    """Return all known .db files: local DB_DIR + anything in the registry."""
    known: dict[Path, None] = {}
    for p in sorted(DB_DIR.glob("*.db")):
        known[p.resolve()] = None
    for entry in registry_list():
        if entry["exists"]:
            known[entry["db"].resolve()] = None
    return sorted(known.keys())


DRIVE_LOCKED = "__locked__"  # sentinel: db busy (e.g. an index/refresh is writing)


def drive_info(db: Path) -> dict | str | None:
    try:
        conn = open_db(db)
        row = conn.execute(
            "SELECT label, root_path, indexed_at, total_files, total_dirs, total_size"
            " FROM drive LIMIT 1"
        ).fetchone()
        # latest snapshot stats (preferred when available)
        sid = latest_snapshot_id(conn)
        n_snapshots = conn.execute(
            "SELECT COUNT(*) FROM snapshots"
        ).fetchone()[0]
        # a drive indexed on another OS/machine mounts elsewhere (e.g.
        # /Volumes/X on macOS → E:\ on Windows) — resolve to the live mount
        resolved_root = str(resolve_root(conn, row[1])) if row and row[1] else None
        conn.close()
    except sqlite3.OperationalError as e:
        # "database is locked" during an in-progress index/refresh is transient,
        # NOT missing metadata — surface it distinctly so we don't tell the user
        # to re-index a perfectly good drive.
        if "locked" in str(e).lower():
            return DRIVE_LOCKED
        return None
    except sqlite3.DatabaseError:
        return None
    if not row:
        return None
    d = dict(zip(
        ["label", "root", "indexed_at", "files", "dirs", "size"], row
    ))
    if resolved_root:
        d["root"] = resolved_root
    d["snapshot_id"] = sid
    d["n_snapshots"] = n_snapshots
    return d


def _treemap_precompute(db: Path):
    """Fetch dirs + direct-file-size sums; propagate sizes bottom-up.
    Returns (by_id, sizes) using only ~133k + 101k rows instead of 4M."""
    conn = open_db(db)
    sid = latest_snapshot_id(conn)
    if sid is None:
        conn.close()
        return {}, {}
    dirs = conn.execute(
        "SELECT id, rel_path, parent_id, size FROM entries"
        " WHERE snapshot_id=? AND is_dir=1", (sid,),
    ).fetchall()
    file_sums_rows = conn.execute(
        "SELECT parent_id, SUM(size) FROM entries"
        " WHERE snapshot_id=? AND is_dir=0 GROUP BY parent_id", (sid,),
    ).fetchall()
    conn.close()

    file_sums = {row[0]: row[1] or 0 for row in file_sums_rows}
    by_id = {eid: (eid, rp, pid, True, sz) for eid, rp, pid, sz in dirs}

    children: dict[int, list[int]] = defaultdict(list)
    roots: list[int] = []
    for eid, _, pid, _ in dirs:
        if pid is None:
            roots.append(eid)
        else:
            children[pid].append(eid)

    sizes: dict[int, int] = {eid: file_sums.get(eid, 0) for eid, _, _, _ in dirs}
    for root in roots:
        stack: list[tuple[int, bool]] = [(root, False)]
        while stack:
            nid, visited = stack.pop()
            if visited:
                for c in children[nid]:
                    sizes[nid] = sizes.get(nid, 0) + sizes.get(c, 0)
            else:
                stack.append((nid, True))
                for c in children[nid]:
                    stack.append((c, False))
    return by_id, sizes


def treemap_rows(db: Path, min_size: int, include_files: bool = False,
                 _precomputed=None) -> list[dict]:
    """Build rows for a plotly treemap. Only folders ≥ min_size are kept;
    their ancestors are added to keep the tree connected. Each row has
    id / parent / name / size / kind.

    Pass _precomputed=(by_id, sizes) to skip the expensive DB query."""
    if _precomputed is not None:
        by_id, sizes = _precomputed
        raw = list(by_id.values())
    else:
        conn = open_db(db)
        sid = latest_snapshot_id(conn)
        if sid is None:
            conn.close()
            return []
        raw = list(conn.execute(
            "SELECT id, rel_path, parent_id, is_dir, size FROM entries"
            " WHERE snapshot_id=?",
            (sid,),
        ))
        conn.close()
        if not raw:
            return []
        sizes = compute_folder_sizes([(e[0], e[2], e[3], e[4]) for e in raw])
        by_id = {e[0]: e for e in raw}

    kept_ids: set[int] = set()
    keep: list[tuple] = []
    for eid, rp, pid, isdir, sz in raw:
        s = sizes.get(eid, sz or 0)
        if isdir and s >= min_size:
            kept_ids.add(eid); keep.append((eid, rp, pid, isdir, s))
        elif not isdir and include_files and s >= min_size:
            kept_ids.add(eid); keep.append((eid, rp, pid, isdir, s))

    # walk up ancestors so plotly's tree is connected
    for eid, rp, pid, isdir, s in list(keep):
        cur = pid
        while cur is not None and cur not in kept_ids:
            anc = by_id.get(cur)
            if anc is None:
                break
            a_id, a_rp, a_pid, a_isdir, a_sz = anc
            a_size = sizes.get(a_id, a_sz or 0)
            kept_ids.add(a_id)
            keep.append((a_id, a_rp, a_pid, a_isdir, a_size))
            cur = a_pid

    rows = []
    for eid, rp, pid, isdir, sz in keep:
        name = os.path.basename(rp) if rp != "." else "/"
        rows.append({
            "id": str(eid),
            "parent": str(pid) if (pid is not None and pid in kept_ids) else "",
            "name": name + ("/" if isdir else ""),
            "size": sz,
            "size_human": human(sz),
            "kind": "folder" if isdir else "file",
            "rel_path": rp,
        })
    return rows


def _hex(b) -> str:
    """Render a BLOB (bytes) or legacy hex string as a hex string."""
    if isinstance(b, (bytes, bytearray)):
        return b.hex()
    return str(b) if b is not None else ""


def dup_file_groups(db: Path, min_size: int,
                    snapshot_id: int | None = None) -> list[dict]:
    """Return groups of duplicate files within one snapshot. Hardlink-aware.

    Uses a single JOIN query (no N+1) to fetch all candidate entries at once,
    then groups in Python. Drive need not be mounted.

    confirmed=True  (=)  all copies share the same full_hash
    confirmed=False (≈)  full_hash not yet computed for some/all copies
    Partial collisions (same partial, different full) are silently skipped.
    """
    conn = open_db(db)
    sid = snapshot_id if snapshot_id is not None else latest_snapshot_id(conn)
    if sid is None:
        conn.close()
        return []

    # Single query: fetch every file that belongs to a candidate group.
    # The inner SELECT identifies (size, partial_hash) pairs with >1 entry;
    # the outer JOIN retrieves all their fields in one round-trip.
    rows = conn.execute(
        "SELECT e.size, e.partial_hash, e.rel_path, e.inode, e.device, e.full_hash"
        " FROM entries e"
        " JOIN ("
        "   SELECT size, partial_hash FROM entries"
        "   WHERE snapshot_id=? AND is_dir=0"
        "     AND partial_hash IS NOT NULL AND size>=?"
        "   GROUP BY size, partial_hash HAVING COUNT(*)>1"
        " ) c ON e.size=c.size AND e.partial_hash=c.partial_hash"
        " WHERE e.snapshot_id=? AND e.is_dir=0",
        (sid, min_size, sid),
    ).fetchall()
    conn.close()

    # Group rows by (size, partial_hash) in Python
    by_key: dict[tuple, list] = defaultdict(list)
    for size, partial, rel, ino, dev, fh in rows:
        by_key[(size, partial)].append((rel, ino, dev, fh))

    out: list[dict] = []
    for (size, partial), members in by_key.items():
        # inode-dedup: hardlinks share storage, count only once
        seen_inodes: set[tuple] = set()
        deduped: list[tuple] = []
        hardlink_count = 0
        for rel, ino, dev, fh in members:
            key = (ino, dev) if ino is not None and dev is not None else None
            if key and key in seen_inodes:
                hardlink_count += 1
            else:
                if key:
                    seen_inodes.add(key)
                deduped.append((rel, fh))

        if len(deduped) < 2:
            continue

        fhs = [fh for _, fh in deduped if fh is not None]
        all_have_fh = len(fhs) == len(deduped)

        if all_have_fh:
            # sub-group by full_hash: each matching sub-group is a confirmed dup
            by_fh: dict = defaultdict(list)
            for rel, fh in deduped:
                by_fh[fh].append(rel)
            for fh, grp_paths in by_fh.items():
                if len(grp_paths) < 2:
                    continue  # partial collision — different content, skip
                wasted = size * (len(grp_paths) - 1)
                out.append({
                    "hash": _hex(fh),
                    "count": len(grp_paths),
                    "size": size,
                    "wasted": wasted,
                    "distinct_inodes": len(grp_paths),
                    "hardlinks": hardlink_count,
                    "paths": [{"path": p, "hardlink": False} for p in grp_paths],
                    "confirmed": True,
                })
        else:
            # approximate: full_hash not yet available for all copies
            wasted = size * (len(deduped) - 1)
            out.append({
                "hash": _hex(partial),
                "count": len(deduped),
                "size": size,
                "wasted": wasted,
                "distinct_inodes": len(deduped),
                "hardlinks": hardlink_count,
                "paths": [{"path": p, "hardlink": False} for p, _ in deduped],
                "confirmed": False,
            })

    out.sort(key=lambda g: -g["wasted"])
    return out


def dup_folder_groups(db: Path,
                      snapshot_id: int | None = None) -> list[dict]:
    conn = open_db(db)
    sid = snapshot_id if snapshot_id is not None else latest_snapshot_id(conn)
    if sid is None:
        conn.close()
        return []
    # Single JOIN query — replaces N+1 (one query per group).
    rows = conn.execute(
        "SELECT e.full_hash, c.cnt, e.rel_path"
        " FROM entries e"
        " JOIN ("
        "   SELECT full_hash, COUNT(*) cnt FROM entries"
        "   WHERE snapshot_id=? AND is_dir=1 AND full_hash IS NOT NULL"
        "   GROUP BY full_hash HAVING cnt > 1"
        " ) c ON e.full_hash = c.full_hash"
        " WHERE e.snapshot_id=? AND is_dir=1"
        " ORDER BY c.cnt DESC, e.full_hash",
        (sid, sid),
    ).fetchall()
    conn.close()
    by_hash: dict = {}
    for fh, cnt, rel in rows:
        key = bytes(fh)
        if key not in by_hash:
            by_hash[key] = {"hash": _hex(fh), "count": cnt, "paths": []}
        by_hash[key]["paths"].append(rel)
    return list(by_hash.values())


def extension_breakdown(db: Path, limit: int = 20) -> list[tuple]:
    """Group files by extension. v3 schema has no `name` column — derive via
    a registered SQLite function on `rel_path`."""
    def _ext(rel_path):
        n = os.path.basename(rel_path or "")
        if "." in n:
            return n[n.index(".") + 1:].lower()
        return "(no ext)"
    conn = open_db(db)
    conn.create_function("ext_of", 1, _ext, deterministic=True)
    sid = latest_snapshot_id(conn)
    if sid is None:
        conn.close()
        return []
    rows = conn.execute(
        "SELECT ext_of(rel_path) AS ext, COUNT(*) c, SUM(size) s"
        " FROM entries WHERE snapshot_id=? AND is_dir=0"
        " GROUP BY ext ORDER BY s DESC LIMIT ?",
        (sid, limit),
    ).fetchall()
    conn.close()
    return rows


def build_csv(rows: list[dict]) -> bytes:
    import csv, io
    if not rows:
        return b""
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    w.writeheader()
    w.writerows(rows)
    return buf.getvalue().encode("utf-8")


def build_xlsx(rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicates"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="EEEEEE")
        for cell in ws[1]:
            cell.font = bold
            cell.fill = fill
        for r in rows:
            ws.append([r[h] for h in headers])
        from openpyxl.utils import get_column_letter
        widths = {"path": 60, "hash": 28, "size_human": 12, "wasted_human": 14}
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 14)
        ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def delete_db_files(target: Path) -> None:
    """Remove the .db plus any -wal/-shm/-journal/.log sidecars, and deregister."""
    registry_remove(target)
    target.unlink(missing_ok=True)
    for ext in ("-wal", "-shm", "-journal", ".log"):
        sib = Path(str(target) + ext)
        sib.unlink(missing_ok=True)


# ---------- sidebar ----------

# registry lookup used in both the sidebar loop and the info dialog
_reg_entries: dict = {e["db"].resolve(): e for e in registry_list()}

with st.sidebar:
    # language toggle
    cur_lang = st.session_state.get("lang", "pt")
    lc1, lc2 = st.columns(2)
    if lc1.button(
        "🇵🇹 PT", width="stretch",
        type="primary" if cur_lang == "pt" else "secondary",
        key="lang_pt",
    ):
        st.session_state.lang = "pt"
        st.rerun()
    if lc2.button(
        "🇬🇧 EN", width="stretch",
        type="primary" if cur_lang == "en" else "secondary",
        key="lang_en",
    ):
        st.session_state.lang = "en"
        st.rerun()

    st.title("💾 drive-xray")
    st.caption(f"engine: {'🦀 Rust' if DX_IS_RUST else '🐍 Python'}  ·  `{DX_CMD[0]}`",
               help=(None if DX_IS_RUST else
                     f"Rust dx unavailable ({DX_FALLBACK_REASON}); "
                     "using the Python engine."))
    # a stale dx binary silently misses features (exclusions, cross-OS mount
    # resolution, checkpointing) — warn when it doesn't match the app version
    if DX_IS_RUST and DX_BIN_VERSION and DX_BIN_VERSION != DX_VERSION:
        st.warning(t("engine_stale", have=DX_BIN_VERSION, want=DX_VERSION))

    # ── self-update from GitHub ────────────────────────────────────────────
    with st.expander(t("upd_title"), expanded=False):
        import update as _upd
        if st.button(t("upd_check"), key="upd_check", width="stretch"):
            st.session_state["upd_status"] = _upd.check_updates()
        _st = st.session_state.get("upd_status")
        if _st:
            if not _st.get("ok"):
                st.warning(_st.get("error"))
            elif _st["behind"] == 0:
                st.success(t("upd_uptodate"))
            else:
                st.info(t("upd_available", n=_st["behind"]))
                for _c in _st["commits"][:8]:
                    st.caption(f"• {_c}")
                if st.button(t("upd_apply"), type="primary", key="upd_apply"):
                    with st.spinner(t("upd_applying")):
                        _res = _upd.apply_update()
                    (st.success if _res.get("ok") else st.error)(_res.get("message"))
                    st.session_state.pop("upd_status", None)

    # whether an indexer/refresher is currently running (this session)
    proc_running = (
        st.session_state.get("idx_proc") is not None
        and st.session_state.idx_proc.poll() is None
    )
    # ground-truth map of drives with a live dx write-process (any session)
    _busy_procs = _active_index_procs()

    # drives list with refresh + delete buttons
    dbs = list_dbs()
    selected_db: Path | None = None

    if dbs:
        st.subheader(t("indexed_drives"))
        current_path = st.session_state.get("db_choice_path")
        # clear stale selection
        if current_path and not any(str(d) == current_path for d in dbs):
            current_path = None
            st.session_state.pop("db_choice_path", None)
        # auto-select the first if nothing chosen yet
        if current_path is None and dbs:
            current_path = str(dbs[0])
            st.session_state.db_choice_path = current_path

        for db in dbs:
            c1, c2, c3, c4, c5 = st.columns([4, 1, 1, 1, 1])
            _reg = _reg_entries.get(db.resolve(), {})
            _display_label = _reg.get("label", db.stem)
            is_current = (current_path == str(db))
            # is THIS drive being indexed/refreshed right now?
            _db_pid = _busy_pid_for(db, _busy_procs)
            _db_busy = _db_pid is not None
            # disable write actions if a process is touching this db (real OS
            # state) or this session spawned one that's still alive.
            _write_disabled = proc_running or _db_busy
            if c1.button(
                (("⏳ " if _db_busy else ("▶ " if is_current else "   "))
                 + _display_label),
                key=f"sel_{db}",
                width="stretch",
                type="primary" if is_current else "secondary",
            ):
                st.session_state.db_choice_path = str(db)
                st.rerun()
            if c2.button(
                "ℹ️", key=f"info_{db}", help=t("info_tooltip"),
            ):
                st.session_state.pending_info = str(db)
                st.rerun()
            if c3.button(
                "📸", key=f"snap_{db}", help=t("snapshot_tooltip"),
                disabled=_write_disabled,
            ):
                start_snapshot(db)
                st.rerun()
            if c4.button(
                "🔄", key=f"ref_{db}", help=t("refresh_tooltip"),
                disabled=_write_disabled,
            ):
                # belt-and-suspenders: bypass the 3s cache and re-check the OS
                # right before spawning, so a stale render can't launch a
                # second concurrent writer over a running one.
                _active_index_procs.clear()
                if _busy_pid_for(db, _active_index_procs()):
                    st.rerun()  # already indexing — refuse silently
                else:
                    start_refresh(db)
                    st.rerun()
            if c5.button(
                "🗑️", key=f"del_{db}", help=t("delete_tooltip"),
            ):
                st.session_state.pending_delete = str(db)
                st.rerun()
            if _db_busy:
                _last = _log_tail(db, 1)
                c1.caption(t("drive_indexing", pid=_db_pid)
                           + (f"\n\n`{_last[-1][:60]}`" if _last else ""))
        if current_path:
            selected_db = Path(current_path)
    else:
        st.info(t("no_drives"))

    st.divider()
    st.subheader(t("index_new_drive"))

    # native folder picker — must live outside the form (forms allow only a
    # submit button), so it writes the path into session_state and reruns.
    _browse_default = ("/Volumes/" if sys.platform == "darwin"
                       else "" if os.name == "nt" else "/media/")
    st.session_state.setdefault("idx_root_v", _browse_default)
    _pk1, _pk2 = st.columns([1, 2])
    if _pk1.button(t("browse_btn"), key="pick_idx_btn", width="stretch"):
        _picked = pick_folder_dialog(st.session_state.get("idx_root_v") or _browse_default)
        if _picked:
            st.session_state["idx_root_v"] = _picked
            st.rerun()
    _pk2.caption(t("browse_hint"))

    with st.form("idx_form"):
        new_root = st.text_input(t("path_label"), key="idx_root_v")
        new_label = st.text_input(t("label_label"), "")
        do_full = st.checkbox(
            t("full_hash_label"), help=t("full_hash_help"),
        )
        one_fs = st.checkbox(
            t("one_fs_label"), value=True, help=t("one_fs_help"),
        )
        skip_cloud = st.checkbox(
            t("skip_cloud_label"), value=True, help=t("skip_cloud_help"),
        )
        skip_system = st.checkbox(
            t("skip_system_label"), value=True,
            help=t("skip_system_help", names=", ".join(SYSTEM_EXCLUDE_DIRS)),
        )
        submitted = st.form_submit_button(
            t("index_button"), type="primary",
            disabled=proc_running or bool(_busy_procs),
        )

    if submitted:
        expanded = str(Path(new_root).expanduser()) if new_root else ""
        if not expanded or not Path(expanded).exists():
            st.error(t("path_not_exist"))
        else:
            label = (new_label.strip()
                     or Path(expanded.rstrip("/")).name
                     or "drive")
            if skip_system:
                # seed the exclusions BEFORE the indexer starts: both engines
                # read the db's exclusion list at walk time, and a fresh index
                # wipes entries/snapshots/drive but preserves exclusions.
                _db_out = DB_DIR / f"{label}.db"
                _existing = get_exclusions(_db_out) if _db_out.exists() else []
                set_exclusions(_db_out, _existing + SYSTEM_EXCLUDE_DIRS)
            start_indexer(expanded, label, do_full, one_fs, skip_cloud)
            st.rerun()

    # indexer status
    proc = st.session_state.get("idx_proc")
    if proc is not None:
        log = st.session_state.idx_log
        label = st.session_state.get("idx_label", "")
        if proc.poll() is None:
            st.info(t("indexing", label=label))
            if st.button(t("cancel"), key="cancel_idx"):
                proc.terminate()
                time.sleep(0.3)
                st.rerun()
        else:
            if proc.returncode == 0:
                st.success(t("completed", label=label))
            else:
                st.error(t("exited_with_code", label=label,
                            code=proc.returncode))
            if st.button(t("clear_log"), key="clear_idx"):
                st.session_state.idx_proc = None
                st.session_state.idx_log = []
                st.rerun()
        with st.expander(t("log"), expanded=proc.poll() is None):
            st.code("\n".join(log[-15:]) or t("no_output_yet"))
        if proc.poll() is None:
            time.sleep(1)
            st.rerun()

    # ---------- settings (db folder / cloud sync) ----------
    st.divider()
    with st.expander(t("settings_title")):
        _cfg = read_config()
        _cur_dir = str(get_db_dir())
        st.caption(f"{t('settings_current_dir')}: `{_cur_dir}`")

        st.session_state.setdefault("cfg_db_dir_input", _cur_dir)
        _dcol1, _dcol2 = st.columns([4, 1])
        _new_dir = _dcol1.text_input(
            t("settings_db_dir"),
            help=t("settings_db_dir_help"),
            key="cfg_db_dir_input",
        )
        _dcol2.markdown("<div style='height:1.75em'></div>",
                        unsafe_allow_html=True)
        if _dcol2.button(t("browse_btn"), key="pick_cfg_btn", width="stretch"):
            _p = pick_folder_dialog(
                st.session_state.get("cfg_db_dir_input") or str(Path.home()))
            if _p:
                st.session_state["cfg_db_dir_input"] = _p
                st.rerun()
        _scol1, _scol2 = st.columns(2)
        if _scol1.button(t("settings_save"), width="stretch",
                         key="cfg_save_btn"):
            _p = Path(_new_dir).expanduser()
            try:
                _p.mkdir(parents=True, exist_ok=True)
                _ = (_p / ".dx_write_test").write_text("ok")
                (_p / ".dx_write_test").unlink(missing_ok=True)
                _cfg["db_dir"] = str(_p.resolve())
                write_config(_cfg)
                st.success(t("settings_saved"))
                time.sleep(0.8)
                st.rerun()
            except Exception:
                st.error(t("settings_invalid"))

        if _scol2.button(t("settings_import_btn"), width="stretch",
                         key="cfg_import_btn",
                         help=t("settings_import_help")):
            _imp = import_folder(Path(_new_dir))
            _new_count = sum(1 for r in _imp if not r["already_registered"])
            _old_count = sum(1 for r in _imp if r["already_registered"])
            if _imp:
                st.success(t("settings_imported", n=_new_count, s=_old_count))
                time.sleep(0.8)
                st.rerun()
            else:
                st.warning(t("settings_import_none"))


# ---------- pre-delete verification dialog ----------

if "del_plan" in st.session_state:
    _dplan = st.session_state["del_plan"]
    _daction = st.session_state.get("del_action", "quarantine")
    _droot = Path(st.session_state.get("del_root_path",
                  st.session_state.get("db_choice_path", "/")))

    @st.dialog(t("del_dialog_title"), width="large")
    def _del_dialog():
        _ok_grps   = [g for g in _dplan if g["status"] == "ok"]
        _no_keeper = [g for g in _dplan if g["status"] == "no_keeper"]
        _miss      = [g for g in _dplan if g["status"] == "keeper_missing"]
        _exec_files = [d for g in _ok_grps for d in g["to_delete"] if d["exists"]]
        _gone_files = [d for g in _ok_grps for d in g["to_delete"] if not d["exists"]]
        _bytes_free = sum(d["size"] for d in _exec_files)

        if _exec_files:
            st.success(t("del_ok_summary", n=len(_ok_grps),
                         f=len(_exec_files), size=human(_bytes_free)))
        if _no_keeper:
            st.error(t("del_no_keeper", n=len(_no_keeper)))
        if _miss:
            st.warning(t("del_keeper_missing", n=len(_miss)))
        if _gone_files:
            st.info(t("del_already_gone", n=len(_gone_files)))

        with st.expander(t("del_detail"), expanded=len(_ok_grps) <= 10):
            for _g in _ok_grps:
                _k = _g["keeper"]
                _ico = "✅" if _k["ok"] else "⚠️"
                st.markdown(
                    f"**#{_g['group']}** — {_ico} `{_k['path']}` "
                    f"*({t('del_keeper_label')})*"
                )
                for _d in _g["to_delete"]:
                    _di = "🗑️" if _d["exists"] else "👻"
                    st.markdown(f"  {_di} `{_d['path']}`"
                                + ("" if _d["exists"] else " *(já eliminado)*"))

        if not _exec_files:
            st.warning(t("del_nothing_to_do"))
            if st.button(t("cancel"), width="stretch", key="dd_cancel"):
                st.session_state.pop("del_plan", None)
                st.rerun()
        else:
            c1, c2 = st.columns(2)
            if c1.button(t("cancel"), width="stretch", key="dd_cancel"):
                st.session_state.pop("del_plan", None)
                st.rerun()
            if c2.button(
                t("del_execute_btn", n=len(_exec_files), size=human(_bytes_free)),
                type="primary", width="stretch", key="dd_exec",
            ):
                _results = [
                    execute_file_action(
                        d["full_path"], _daction,
                        root_path=_droot,
                        db_path=str(st.session_state.get("db_choice_path", "")),
                    )
                    for d in _exec_files
                ]
                _ok_count = sum(1 for r in _results if r["ok"])
                _freed = sum(d["size"] for d, r in zip(_exec_files, _results)
                             if r["ok"])
                st.session_state["del_result"] = {
                    "ok": _ok_count,
                    "freed_bytes": _freed,
                    "errors": [r for r in _results if not r["ok"]],
                }
                st.session_state.pop("del_plan", None)
                st.session_state.pop("dupes_ready", None)  # force re-scan
                st.rerun()

    _del_dialog()


# ---------- info / rename dialog ----------

if "pending_info" in st.session_state:
    _pending_info = st.session_state["pending_info"]

    @st.dialog(t("drive_info_title"))
    def _info_dialog():
        _db = Path(_pending_info)
        _reg = _reg_entries.get(_db.resolve(), {})
        st.markdown(f"**{t('drive_db_path')}**")
        st.code(str(_db), language=None)
        st.markdown(f"**{t('drive_root')}:** `{_reg.get('root', '?')}`")
        st.markdown(f"**{t('drive_last_indexed')}:** {_reg.get('last_indexed', '?')}")
        st.divider()
        _current = _reg.get("label", _db.stem)
        _new = st.text_input(t("rename_label"), value=_current, key="rename_inp")
        rc1, rc2 = st.columns(2)
        if rc1.button(t("rename_save"), type="primary",
                      width="stretch", key="rename_ok",
                      disabled=not _new.strip() or _new.strip() == _current):
            _nl = _new.strip()
            _conn = open_db(_db)
            _conn.execute("UPDATE drive SET label = ?", (_nl,))
            _conn.commit()
            _conn.close()
            registry_register(_db, _nl, Path(_reg.get("root", "/")))
            st.session_state.pop("pending_info", None)
            st.rerun()
        if rc2.button(t("cancel"), width="stretch", key="rename_cancel"):
            st.session_state.pop("pending_info", None)
            st.rerun()

    _info_dialog()


# ---------- delete confirmation dialog ----------

if "pending_delete" in st.session_state:
    _pending = st.session_state["pending_delete"]

    @st.dialog(t("confirm_delete"))
    def _confirm_delete_dialog():
        target = Path(_pending)
        st.markdown(t("delete_question", name=target.stem))
        c1, c2 = st.columns(2)
        if c1.button(
            t("yes_delete"), type="primary",
            width="stretch", key="pd_yes",
        ):
            delete_db_files(target)
            st.session_state.pop("pending_delete", None)
            if st.session_state.get("db_choice_path") == _pending:
                st.session_state.pop("db_choice_path", None)
            st.session_state.pop("dupes_ready", None)
            st.rerun()
        if c2.button(
            t("cancel"), width="stretch", key="pd_no",
        ):
            st.session_state.pop("pending_delete", None)
            st.rerun()

    _confirm_delete_dialog()


# ---------- main ----------

if selected_db is None:
    st.title(t("main_welcome_title"))
    st.markdown(t("main_welcome_body"))
    st.stop()

def _render_op_status(db: Path, running: bool) -> None:
    """Live status + reviewable log of the drive's index/refresh, read from
    the persistent <db>.log — works even after the tab that launched the
    operation was closed (the in-session log dies with the session)."""
    tail = _log_tail(db, 25)
    if not tail:
        return
    if running:
        st.caption(f"⏳ `{tail[-1][:120]}`")
    with st.expander(t("op_log_title"), expanded=running):
        st.code("\n".join(tail))
    if running and st.button("🔄", key=f"oplog_refresh_{db}",
                             help=t("op_log_refresh")):
        st.rerun()


info = drive_info(selected_db)
_selected_busy = _busy_pid_for(selected_db, _busy_procs) is not None
if info == DRIVE_LOCKED:
    st.info(t("drive_busy", name=selected_db.name))
    _render_op_status(selected_db, running=True)
    if st.button(t("drive_busy_retry"), key="drive_busy_retry_btn"):
        st.rerun()
    st.stop()
if info is None:
    st.error(t("no_metadata_error", name=selected_db.name))
    st.stop()

st.title(f"💾 {info['label']}")
st.caption(f"`{info['root']}`  ·  {t('indexed_on')} {info['indexed_at']}")
if _selected_busy:
    _render_op_status(selected_db, running=True)

tab_summary, tab_dupes, tab_map, tab_history, tab_compare = st.tabs(
    [t("tab_summary"), t("tab_dupes"), t("tab_map"),
     t("tab_history"), t("tab_compare")]
)

# --- Summary ---
with tab_summary:
    c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
    c1.metric(t("files"), f"{info['files']:,}")
    c2.metric(t("folders"), f"{info['dirs']:,}")
    c3.metric(t("total_size"), human(info["size"] or 0))
    # db file size + compact button
    db_size = selected_db.stat().st_size if selected_db.exists() else 0
    wal = Path(str(selected_db) + "-wal")
    if wal.exists():
        db_size += wal.stat().st_size
    c4.metric(t("db_size"), human(db_size))
    if st.button(t("compact_button"), help=t("compact_help"),
                 disabled=proc_running
                 or _busy_pid_for(selected_db, _busy_procs) is not None):
        start_compact(selected_db)
        st.rerun()

    # ── structural DB health check ────────────────────────────────────────────
    with st.expander("🩺 Doctor", expanded=False):
        st.caption("Valida a estrutura do índice sem modificar a base de dados.")
        if st.button("🩺 Correr doctor", key="doctor_btn"):
            from drive_xray import doctor_db
            st.session_state["doctor_result"] = doctor_db(selected_db)
            st.session_state["doctor_result_db"] = str(selected_db)
        _doc = st.session_state.get("doctor_result")
        if _doc and st.session_state.get("doctor_result_db") == str(selected_db):
            if _doc.get("ok"):
                st.success("Todos os checks passaram.")
            else:
                st.error("Foram encontrados problemas neste índice.")
            st.dataframe(
                [{"": "✓" if c["ok"] else "✗",
                  "check": c["name"],
                  "detalhe": c["detail"]}
                 for c in _doc.get("checks", [])],
                width="stretch", hide_index=True,
            )

    # ── integrity / bit-rot verification ──────────────────────────────────────
    with st.expander(t("verify_title"), expanded=False):
        st.caption(t("verify_caption"))
        _vfull = st.checkbox(t("verify_full"), value=False,
                             help=t("verify_full_help"))
        if st.button(t("verify_btn"), key="verify_btn"):
            from drive_xray import verify_integrity
            _vbar = st.progress(0.0, text=t("verify_running"))
            def _vcb(i, n, c):
                _vbar.progress(i / max(n, 1),
                               text=f"{i}/{n} · {c} {t('verify_corrupt_short')}")
            _vres = verify_integrity(selected_db, full=_vfull, progress=_vcb)
            st.session_state["verify_result"] = _vres
            st.session_state["verify_result_db"] = str(selected_db)
        _vr = st.session_state.get("verify_result")
        if _vr and st.session_state.get("verify_result_db") == str(selected_db):
            if not _vr["root_mounted"]:
                st.warning(t("verify_unmounted", root=_vr["root"]))
            else:
                vc1, vc2, vc3, vc4 = st.columns(4)
                vc1.metric("✅ OK", f"{_vr['ok']:,}")
                vc2.metric("⚠️ " + t("verify_corrupt_short"),
                           f"{len(_vr['corrupted']):,}")
                vc3.metric(t("verify_changed"), f"{_vr['size_changed']:,}")
                vc4.metric(t("verify_missing"), f"{_vr['missing']:,}")
                if _vr["corrupted"]:
                    st.error(t("verify_rot_found", n=len(_vr["corrupted"])))
                    st.dataframe(
                        [{"": "⚠️", t("tags_col_path"): c["rel_path"],
                          t("db_size"): human(c["size"] or 0)}
                         for c in _vr["corrupted"][:500]],
                        width="stretch", hide_index=True)
                else:
                    st.success(t("verify_clean"))

    def _compute_ext():
        db = selected_db
        if DX_IS_RUST:
            p = subprocess.run(
                [*DX_CMD, "ext-breakdown", str(db), "--limit", "20", "--json"],
                capture_output=True, text=True,
            )
            if p.returncode == 0 and p.stdout.strip():
                return [(r["ext"], r["files"], r["size_bytes"]) for r in json.loads(p.stdout)]
        return extension_breakdown(db)

    with st.spinner(t("calculating")):
        ext_rows = _ss_compute(f"ext_{selected_db}", _compute_ext)
    if ext_rows:
        st.subheader(t("top_ext"))
        st.dataframe(
            [{t("ext_col"): e, t("files_col"): c, t("size_col"): human(s)}
             for e, c, s in ext_rows],
            width="stretch", hide_index=True,
        )

# --- Duplicates ---
with tab_dupes:
    root_path = Path(info["root"])
    root_mounted = root_path.exists()
    if not root_mounted:
        st.info(t("drive_not_mounted", root=str(root_path)))

    _dupes_key = f"dupes_{selected_db}"
    if _dupes_key not in st.session_state:
        if st.button(t("load_duplicates"), type="primary", key="btn_load_dupes"):
            def _compute_dupes():
                db = selected_db
                if DX_IS_RUST:
                    rf: list | None = None
                    rd: list | None = None
                    def _rf():
                        nonlocal rf
                        p = subprocess.run(
                            [*DX_CMD, "dup-groups", str(db),
                             "--min-size", str(_CACHE_FLOOR), "--json"],
                            capture_output=True, text=True,
                        )
                        rf = json.loads(p.stdout) if p.returncode == 0 and p.stdout.strip() \
                             else dup_file_groups(db, _CACHE_FLOOR)
                    def _rd():
                        nonlocal rd
                        p = subprocess.run(
                            [*DX_CMD, "dup-folders", str(db), "--json"],
                            capture_output=True, text=True,
                        )
                        rd = json.loads(p.stdout) if p.returncode == 0 and p.stdout.strip() \
                             else dup_folder_groups(db)
                    t1 = threading.Thread(target=_rf)
                    t2 = threading.Thread(target=_rd)
                    t1.start(); t2.start()
                    t1.join(); t2.join()
                    return rf, rd
                return dup_file_groups(db, _CACHE_FLOOR), dup_folder_groups(db)

            with st.spinner(t("calculating")):
                st.session_state[_dupes_key] = _ss_compute(_dupes_key, _compute_dupes)
            st.rerun()
    else:
        _all_files, _all_folders = st.session_state[_dupes_key]

        min_size_mb = st.slider(t("ignore_smaller_mb"), 1, 500, 1)
        min_size = min_size_mb * 1024 * 1024
        files = [g for g in _all_files if g["size"] >= min_size]
        folders = _all_folders

        c1, c2, c3 = st.columns(3)
        c1.metric(t("file_groups"), f"{len(files):,}")
        c2.metric(t("folder_groups"), f"{len(folders):,}")
        c3.metric(
            t("wasted_space"),
            human(sum(g["wasted"] for g in files)),
        )

        # Optional: confirm approximate matches with full hash (slow)
        if root_mounted:
            with st.expander(t("confirm_expander"), expanded=False):
                st.caption(t("confirm_caption"))
                if st.button(t("confirm_btn"), key="confirm_full_hash_btn"):
                    st.session_state.pop(_dupes_key, None)
                    with st.status(t("calculating"), expanded=True) as _status:
                        if DX_IS_RUST:
                            st.write(t("confirming_candidates"))
                            _proc = subprocess.run(
                                [*DX_CMD, "dedupe", str(selected_db),
                                 "--min-size", str(min_size)],
                                capture_output=True, text=True,
                            )
                            for line in _proc.stderr.splitlines():
                                if line.strip():
                                    st.write(line)
                        else:
                            conn = open_db(selected_db)
                            n = fill_full_hashes(conn, root_path, min_size)
                            st.write(t("files_hashed", n=n))
                            st.write(t("computing_merkle"))
                            compute_dir_hashes(conn)
                            conn.close()
                        _status.update(label=t("done"), state="complete")
                    st.rerun()

        # export buttons
        if files:
            conn = open_db(selected_db)
            export_rows = _duplicate_rows(conn, min_size)
            conn.close()
            ec1, ec2 = st.columns(2)
            ec1.download_button(
                t("download_csv"),
                data=build_csv(export_rows),
                file_name=f"{selected_db.stem}-duplicates.csv",
                mime="text/csv",
                width="stretch",
            )
            ec2.download_button(
                t("download_xlsx"),
                data=build_xlsx(export_rows),
                file_name=f"{selected_db.stem}-duplicates.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

        st.subheader(t("duplicate_files"))
        st.caption(t("sorted_top200"))
        for g in files[:200]:
            hl_tag = (f"  ·  {g['hardlinks']} {t('hardlink_tag')}"
                      if g["hardlinks"] else "")
            match_tag = "=" if g.get("confirmed") else "≈"
            with st.expander(
                f"{match_tag} {g['count']}× {human(g['size'])}  ·  "
                f"{t('wasted')} {human(g['wasted'])}{hl_tag}  ·  {g['hash'][:12]}"
            ):
                for p in g["paths"]:
                    suffix = f"   ← {t('hardlink_tag')}" if p["hardlink"] else ""
                    st.code(p["path"] + suffix, language=None)
        if len(files) > 200:
            st.caption(t("groups_not_shown", n=len(files) - 200))

        st.subheader(t("duplicate_folders"))
        for g in folders[:100]:
            with st.expander(
                f"{g['count']}× {t('identical_folders')}  ·  {g['hash'][:12]}"
            ):
                for p in g["paths"]:
                    st.code(p + "/", language=None)
        if len(folders) > 100:
            st.caption(t("groups_not_shown", n=len(folders) - 100))

        # ----- Interactive delete -----
        st.divider()
        st.subheader(t("del_title"))
        st.caption(t("del_caption"))

        if not root_mounted:
            st.warning(t("del_root_unmounted"))
        elif files:
            # Build editor dataframe — cap at 200 groups (same as display)
            _del_rows = []
            for _gi, _g in enumerate(files[:200], 1):
                _sorted_paths = sorted(_g["paths"], key=lambda p: len(p["path"]))
                for _pi, _p in enumerate(_sorted_paths):
                    _del_rows.append({
                        t("del_col"): _pi > 0,      # keep shortest, mark rest
                        "#": _gi,
                        t("cross_col_size"): human(_g["size"]),
                        t("cross_col_path"): _p["path"],
                        "_size": _g["size"],
                    })
            _del_df = pd.DataFrame(_del_rows)
            _edited = st.data_editor(
                _del_df,
                column_config={
                    t("del_col"): st.column_config.CheckboxColumn(
                        t("del_col"), default=False),
                    "#": st.column_config.NumberColumn(
                        "#", disabled=True, width="small"),
                    t("cross_col_size"): st.column_config.TextColumn(
                        t("cross_col_size"), disabled=True, width="small"),
                    t("cross_col_path"): st.column_config.TextColumn(
                        t("cross_col_path"), disabled=True, width="large"),
                    "_size": None,
                },
                disabled=["#", t("cross_col_size"), t("cross_col_path")],
                hide_index=True,
                width="stretch",
                key="del_editor",
            )

            _del_action = st.selectbox(
                t("del_action_label"),
                options=list(CLEANUP_ACTIONS),
                format_func=lambda a: t(f"action_{a}"),
                key="del_action_sel",
            )

            _marked = _edited[_edited[t("del_col")].astype(bool)]
            _n_marked = len(_marked)
            _bytes_marked = int(_marked["_size"].sum()) if _n_marked else 0

            if _n_marked == 0:
                st.caption(t("del_none_selected"))
            else:
                if st.button(
                    t("del_verify_btn", n=_n_marked, size=human(_bytes_marked)),
                    type="primary", key="del_verify_btn",
                ):
                    _plan = []
                    for _gid in _marked["#"].unique():
                        _gdf = _edited[_edited["#"] == _gid]
                        _keepers = _gdf[~_gdf[t("del_col")].astype(bool)]
                        _to_del = _gdf[_gdf[t("del_col")].astype(bool)]

                        if _keepers.empty:
                            _plan.append({"group": int(_gid),
                                          "status": "no_keeper",
                                          "keeper": None, "to_delete": []})
                            continue

                        _kr = _keepers.iloc[0]
                        _kcheck = verify_file(
                            root_path, _kr[t("cross_col_path")], int(_kr["_size"]))
                        _del_checks = []
                        for _, _dr in _to_del.iterrows():
                            _dc = verify_file(
                                root_path, _dr[t("cross_col_path")], int(_dr["_size"]))
                            _del_checks.append({
                                "path": _dr[t("cross_col_path")],
                                "full_path": _dc["full_path"],
                                "exists": _dc["ok"],
                                "size": int(_dr["_size"]),
                            })
                        _plan.append({
                            "group": int(_gid),
                            "status": "ok" if _kcheck["ok"] else "keeper_missing",
                            "keeper": {"path": _kr[t("cross_col_path")],
                                       "full_path": _kcheck["full_path"],
                                       "ok": _kcheck["ok"],
                                       "reason": _kcheck["reason"]},
                            "to_delete": _del_checks,
                        })
                    st.session_state["del_plan"] = _plan
                    st.session_state["del_action"] = _del_action
                    st.session_state["del_root_path"] = str(root_path)
                    st.rerun()

        # show execution result banner
        if "del_result" in st.session_state:
            _res = st.session_state.pop("del_result")
            st.success(t("del_done", n=_res["ok"],
                         size=human(_res["freed_bytes"])))
            if _res["errors"]:
                st.error(t("del_errors", n=len(_res["errors"])))

        # ----- Assisted cleanup -----
        st.divider()
        st.subheader(t("cleanup_title"))
        st.caption(t("cleanup_caption"))
        cc1, cc2 = st.columns(2)
        cleanup_strategy = cc1.selectbox(
            t("cleanup_strategy"),
            options=list(CLEANUP_STRATEGIES),
            format_func=lambda s: t(f"strategy_{s}"),
        )
        cleanup_action = cc2.selectbox(
            t("cleanup_action"),
            options=list(CLEANUP_ACTIONS),
            format_func=lambda a: t(f"action_{a}"),
        )
        if st.button(t("cleanup_generate"), key="gen_cleanup"):
            script = generate_cleanup_script(
                selected_db, min_size,
                strategy=cleanup_strategy, action=cleanup_action,
            )
            st.session_state["cleanup_script"] = script
            st.session_state["cleanup_db"] = str(selected_db)

        if (st.session_state.get("cleanup_script")
                and st.session_state.get("cleanup_db") == str(selected_db)):
            script = st.session_state["cleanup_script"]
            n_actions = sum(1 for l in script.splitlines()
                            if l.startswith(("rm ", "mv ")))
            st.success(t("cleanup_ready", n=n_actions))
            st.download_button(
                t("cleanup_download"),
                data=script.encode("utf-8"),
                file_name=f"{selected_db.stem}-cleanup-{cleanup_action}.sh",
                mime="text/x-shellscript",
                type="primary",
            )
            with st.expander(t("cleanup_preview"), expanded=False):
                st.code(script, language="bash")

# --- TreeMap ---
with tab_map:
    st.caption(t("map_caption"))
    map_min_mb = st.slider(t("map_min_mb"), 1, 5000, 100, key="map_min_mb")
    map_min = map_min_mb * 1024 * 1024
    map_include_files = st.checkbox(t("map_include_files"), value=False,
                                    key="map_include_files")
    _map_pre_key = f"map_pre_{selected_db}"
    with st.spinner(t("calculating")):
        _pre = _ss_compute(_map_pre_key, lambda: _treemap_precompute(selected_db))
    rows = treemap_rows(selected_db, map_min, map_include_files, _precomputed=_pre)

    _folder_tags = tags_get(selected_db)

    # auto-tags — computed from file extensions in latest snapshot, cached by db mtime
    _at_mtime_key = f"_at_mtime_{selected_db}"
    _at_data_key  = f"_auto_tags_{selected_db}"
    # cache key = (db mtime, rules-file mtime) so editing the YAML rules busts
    # the cache immediately, no restart needed.
    _db_mtime = selected_db.stat().st_mtime if selected_db.exists() else 0
    _rules_mtime = (AUTO_TAGS_YAML_PATH.stat().st_mtime
                    if AUTO_TAGS_YAML_PATH.exists() else 0)
    _at_key = (_db_mtime, _rules_mtime)
    if st.session_state.get(_at_mtime_key) != _at_key:
        st.session_state[_at_data_key]  = compute_auto_tags(selected_db)
        st.session_state[_at_mtime_key] = _at_key
    _auto_tags: dict = st.session_state.get(_at_data_key, {})

    # colour palette — one colour per unique tag NAME, shared by manual AND
    # auto tags, so an auto-classified "NGS" gets the same colour as a manual
    # "NGS" (rather than every auto-tagged folder sharing one grey).
    _TAG_PALETTE = [
        "#e74c3c", "#3498db", "#2ecc71", "#f39c12",
        "#9b59b6", "#1abc9c", "#e67e22", "#e91e63",
        "#00bcd4", "#8bc34a", "#ff7043", "#7986cb",
    ]
    _all_unique_tags = sorted(
        {tg for tgs in _folder_tags.values() for tg in tgs}
        | {tg for tgs in _auto_tags.values() for tg in tgs}
    )
    _tag_color = {tg: _TAG_PALETTE[i % len(_TAG_PALETTE)]
                  for i, tg in enumerate(_all_unique_tags)}
    _DEFAULT_FOLDER_COLOR = "#1f77b4"
    _DEFAULT_FILE_COLOR   = "#888888"
    _COLD_COLOR           = "#8ecae6"  # icy blue — cold / archive candidates

    # cold folders — populated once the cold-data panel has been computed for
    # THIS drive (defaults to >2 years untouched). A folder is cold if it IS,
    # or sits under, a maximal cold candidate.
    _cold_candidates: set = set()
    if (st.session_state.get("cold_result_db") == str(selected_db)
            and st.session_state.get("cold_result")):
        _cold_candidates = {c["folder"]
                            for c in st.session_state["cold_result"]["candidates"]}

    def _is_cold(rp: str) -> bool:
        if not _cold_candidates:
            return False
        if rp in _cold_candidates:
            return True
        parts = rp.split("/")
        return any("/".join(parts[:k]) in _cold_candidates
                   for k in range(1, len(parts)))

    if not rows:
        st.info(t("map_empty"))
    else:
        import plotly.graph_objects as go
        ids = [r["id"] for r in rows]
        labels = [r["name"] for r in rows]
        parents = [r["parent"] for r in rows]
        values = [r["size"] for r in rows]

        # id → rel_path lookup for click-to-select
        _id_to_relpath = {r["id"]: r["rel_path"] for r in rows}
        # folder list needed both for click handling and the selectbox below
        _map_folders = sorted({
            r["rel_path"] for r in rows if r["kind"] == "folder" and r["rel_path"] != "."
        })

        def _cell_color(rp, kind):
            if kind != "folder":
                return _DEFAULT_FILE_COLOR
            tgs = _folder_tags.get(rp, [])
            if tgs:                                      # manual tag wins
                return _tag_color.get(tgs[0], _DEFAULT_FOLDER_COLOR)
            if _is_cold(rp):                             # then cold (archive)
                return _COLD_COLOR
            atgs = _auto_tags.get(rp, [])
            if atgs:                                     # then auto-tag colour
                return _tag_color.get(atgs[0], _DEFAULT_FOLDER_COLOR)
            return _DEFAULT_FOLDER_COLOR

        def _hover_extra(rp):
            tgs = _folder_tags.get(rp, [])
            atgs = _auto_tags.get(rp, [])
            nt = notes_get(selected_db, rp)
            parts = []
            if tgs:
                parts.append("🏷️ " + " · ".join(tgs))
            if atgs:
                parts.append("🤖 " + " · ".join(atgs))
            if _is_cold(rp):
                parts.append("❄️ " + t("cold_badge"))
            if nt:
                parts.append("📝 " + (nt[:80] + "…" if len(nt) > 80 else nt))
            return ("<br>" + "<br>".join(parts)) if parts else ""

        cell_colors = [_cell_color(r["rel_path"], r["kind"]) for r in rows]
        customdata  = [r["size_human"] + _hover_extra(r["rel_path"]) for r in rows]

        fig = go.Figure(go.Treemap(
            ids=ids, labels=labels, parents=parents, values=values,
            branchvalues="total",
            customdata=customdata,
            hovertemplate=("<b>%{label}</b><br>%{customdata}"
                           "<br>%{percentParent:.1%} of parent<extra></extra>"),
            marker=dict(colors=cell_colors, showscale=False),
            textinfo="label+value",
            texttemplate="<b>%{label}</b><br>%{customdata}",
        ))
        fig.update_layout(margin=dict(t=10, l=0, r=0, b=0), height=700,
                          clickmode="event+select")
        _map_event = st.plotly_chart(
            fig, width="stretch",
            on_select="rerun", key="treemap_plot",
        )
        st.caption(t("map_legend", n=len(rows)))

        # colour legend — per-tag colours (manual + auto share them) plus a
        # cold chip once cold data has been computed.
        _leg_items = list(_tag_color.items())
        if _cold_candidates:
            _leg_items.append(("❄️ " + t("cold_badge"), _COLD_COLOR))
        if _leg_items:
            with st.expander(t("tags_legend"), expanded=False):
                _leg_cols = st.columns(min(len(_leg_items), 4))
                for i, (tg, col) in enumerate(_leg_items):
                    _leg_cols[i % len(_leg_cols)].markdown(
                        f'<span style="background:{col};color:#fff;'
                        f'padding:2px 8px;border-radius:4px;font-size:0.85em">'
                        f'{tg}</span>', unsafe_allow_html=True)

        # pre-select clicked folder in the tag editor
        if _map_event and _map_event.selection and _map_event.selection.get("points"):
            _clicked_id = str(_map_event.selection["points"][0].get("id", ""))
            _clicked_rp = _id_to_relpath.get(_clicked_id, "")
            if _clicked_rp and _clicked_rp != "." and _clicked_rp in _map_folders:
                if st.session_state.get("tags_folder_sel") != _clicked_rp:
                    st.session_state["tags_folder_sel"] = _clicked_rp
                    st.session_state["_tags_last_folder"] = None

    # ── Tag editor ────────────────────────────────────────────────────────────
    with st.expander(t("tags_expander"), expanded=bool(_folder_tags)):
        st.caption(t("tags_caption"))

        _map_folders = sorted({
            r["rel_path"] for r in rows if r["kind"] == "folder" and r["rel_path"] != "."
        }) if rows else []

        _tc1, _tc2 = st.columns([3, 2])
        _sel_folder = _tc1.selectbox(
            t("tags_select"), _map_folders,
            index=None,
            key="tags_folder_sel",
        )
        _existing_tags = ", ".join(_folder_tags.get(_sel_folder, [])) if _sel_folder else ""
        _input_key = f"tags_input_{_sel_folder}"
        if st.session_state.get("_tags_last_folder") != _sel_folder:
            st.session_state["_tags_last_folder"] = _sel_folder
            st.session_state[_input_key] = _existing_tags

        # auto-tag suggestions for the selected folder
        _sel_auto = _auto_tags.get(_sel_folder, []) if _sel_folder else []
        if _sel_auto:
            _chips_html = " ".join(
                f'<span style="background:{_tag_color.get(tg, "#78909c")};'
                f'color:#fff;padding:2px 8px;border-radius:12px;'
                f'font-size:0.8em;margin-right:2px">{tg}</span>'
                for tg in _sel_auto
            )
            st.markdown(
                f'🤖 {t("auto_tags_detected")}: {_chips_html}',
                unsafe_allow_html=True,
            )
            if st.button(t("auto_tags_promote"), key="auto_promote_btn"):
                _merged = list(dict.fromkeys(
                    [x.strip() for x in st.session_state.get(_input_key, "").split(",") if x.strip()]
                    + _sel_auto
                ))
                st.session_state[_input_key] = ", ".join(_merged)
                st.rerun()

        _tag_input = _tc2.text_input(
            t("tags_input"),
            placeholder="backup, importante, NGS",
            key=_input_key,
        )

        # note field
        _note_key = f"note_input_{_sel_folder}"
        if st.session_state.get("_tags_last_folder") != _sel_folder or _note_key not in st.session_state:
            _existing_note = notes_get(selected_db, _sel_folder) if _sel_folder else ""
            st.session_state[_note_key] = _existing_note
        _note_input = st.text_area(
            t("tags_note_label"),
            placeholder=t("tags_note_placeholder"),
            key=_note_key,
            height=80,
            disabled=not _sel_folder,
        )

        _tb1, _tb2 = st.columns(2)
        if _tb1.button(t("tags_save"), type="primary",
                       disabled=not _sel_folder, key="tags_save_btn"):
            _new_tags = [x.strip() for x in _tag_input.split(",") if x.strip()]
            tags_set(selected_db, _sel_folder, _new_tags, note=_note_input)
            _folder_tags = tags_get(selected_db)
            st.toast(t("tags_saved"), icon="✅")
        if _tb2.button(t("tags_remove_btn"),
                       disabled=not _sel_folder or (
                           _sel_folder not in _folder_tags
                           and not notes_get(selected_db, _sel_folder)
                       ),
                       key="tags_remove_btn"):
            tags_set(selected_db, _sel_folder, [], note="")
            _folder_tags = tags_get(selected_db)
            st.toast(t("tags_removed"), icon="🗑️")

        st.divider()
        st.caption(t("tags_active"))
        if _folder_tags:
            _tag_filter = st.text_input(
                t("tags_filter"), placeholder="backup", key="tags_filter_input",
                label_visibility="collapsed",
            )
            _q = _tag_filter.strip().lower()
            _filtered = {
                p: tgs for p, tgs in sorted(_folder_tags.items())
                if not _q or _q in p.lower() or any(_q in tg.lower() for tg in tgs)
            }
            if _filtered:
                st.dataframe(
                    [{t("tags_col_path"): p,
                      t("tags_col_tags"): " · ".join(tgs),
                      t("tags_col_note"): notes_get(selected_db, p)}
                     for p, tgs in _filtered.items()],
                    width="stretch", hide_index=True,
                )
            else:
                st.info(t("tags_filter_empty"))
        else:
            st.info(t("tags_none"))

    # ── auto-tag rules (editable YAML) ────────────────────────────────────────
    _at_rules = get_auto_tag_rules()
    with st.expander(t("at_rules_title", n=len(_at_rules)), expanded=False):
        _rules_custom = AUTO_TAGS_YAML_PATH.exists()
        st.caption(t("at_rules_src_custom") if _rules_custom
                   else t("at_rules_src_default"))
        if _rules_custom:
            st.text_input(t("at_rules_path"), value=str(AUTO_TAGS_YAML_PATH),
                          disabled=True, key="at_rules_path_display")
            st.caption(t("at_rules_edit_hint"))
        else:
            if st.button(t("at_rules_init_btn"), key="at_rules_init"):
                _p = write_default_auto_tag_rules()
                st.toast(t("at_rules_created"), icon="⚙️")
                st.rerun()
        st.dataframe(
            [{t("tags_col_tags"): _tag, "ext": ", ".join(sorted(_exts))}
             for _exts, _tag in _at_rules],
            width="stretch", hide_index=True,
        )

    # ── folder exclusions ─────────────────────────────────────────────────────
    with st.expander(t("excl_title"), expanded=False):
        st.caption(t("excl_caption"))
        _excl = get_exclusions(selected_db)
        _excl_folders = sorted({
            r["rel_path"] for r in rows
            if r["kind"] == "folder" and r["rel_path"] != "."
        }) if rows else []
        _ec1, _ec2 = st.columns([3, 1])
        _pick = _ec1.selectbox(t("excl_add"), [""] + _excl_folders,
                               key="excl_pick")
        _typed = _ec2.text_input(t("excl_or_type"), key="excl_typed",
                                 label_visibility="visible")
        if st.button(t("excl_add_btn"), key="excl_add_btn"):
            _new = (_typed.strip() or _pick).strip("/").replace("\\", "/")
            if _new and _new not in _excl:
                set_exclusions(selected_db, _excl + [_new])
                st.toast(t("excl_added"), icon="🚫")
                st.rerun()
        _sys_missing = [d for d in SYSTEM_EXCLUDE_DIRS if d not in _excl]
        if st.button(t("excl_system_btn"), key="excl_system_btn",
                     disabled=not _sys_missing,
                     help=t("excl_system_help",
                            names=", ".join(SYSTEM_EXCLUDE_DIRS))):
            set_exclusions(selected_db, _excl + _sys_missing)
            st.toast(t("excl_system_added", n=len(_sys_missing)), icon="🖥️")
            st.rerun()
        if _excl:
            st.caption(t("excl_current"))
            for _e in _excl:
                _c1, _c2 = st.columns([5, 1])
                _c1.code(_e, language=None)
                if _c2.button("✕", key=f"excl_rm_{_e}"):
                    set_exclusions(selected_db, [x for x in _excl if x != _e])
                    st.rerun()
            st.info(t("excl_refresh_hint"))
        else:
            st.caption(t("excl_none"))

    # ── cold data (archive candidates) ────────────────────────────────────────
    with st.expander(t("cold_title"), expanded=False):
        st.caption(t("cold_caption"))
        _cd_c1, _cd_c2 = st.columns(2)
        _cd_years = _cd_c1.slider(t("cold_years"), 0.5, 10.0, 2.0, 0.5,
                                  key="cold_years")
        _cd_min_mb = _cd_c2.slider(t("minimum_mb"), 0, 2000, 100, key="cold_min")
        if st.button(t("cold_btn"), type="primary", key="cold_run"):
            with st.spinner(t("calculating")):
                st.session_state["cold_result"] = cold_data(
                    selected_db,
                    older_than_days=int(_cd_years * 365),
                    min_size=_cd_min_mb * 1024 * 1024,
                    max_rows=5000,
                )
            st.session_state["cold_result_db"] = str(selected_db)

        if ("cold_result" in st.session_state
                and st.session_state.get("cold_result_db") == str(selected_db)):
            _cd = st.session_state["cold_result"]
            if _cd["total_folders"] == 0:
                st.success(t("cold_none"))
            else:
                _cm1, _cm2, _cm3 = st.columns(3)
                _cm1.metric(t("cold_metric_folders"), f"{_cd['total_folders']:,}")
                _cm2.metric(t("cold_metric_bytes"), human(_cd["total_bytes"]))
                _cm3.metric(t("cold_metric_cutoff"), _cd["cutoff_iso"][:10])
                st.caption(t("cold_map_hint"))

                import csv as _cd_csv, io as _cd_io
                _cd_buf = _cd_io.StringIO()
                _cd_w = _cd_csv.writer(_cd_buf)
                _cd_w.writerow(["folder", "size_bytes", "size_human",
                                "newest_file", "file_count"])
                for _c in _cd["candidates"]:
                    _cd_w.writerow([_c["folder"], _c["size"], human(_c["size"]),
                                    _c["newest_iso"], _c["file_count"]])
                st.download_button(
                    t("cold_download_csv"), data=_cd_buf.getvalue().encode(),
                    file_name="cold-data.csv", mime="text/csv", key="cold_csv",
                )
                st.dataframe(
                    [{t("cold_col_folder"): _c["folder"] + "/",
                      t("cold_col_size"): human(_c["size"]),
                      t("cold_col_newest"): _c["newest_iso"][:10],
                      t("cold_col_files"): _c["file_count"]}
                     for _c in _cd["candidates"][:1000]],
                    width="stretch", hide_index=True,
                )
                if len(_cd["candidates"]) > 1000:
                    st.caption(t("cold_truncated", shown=1000,
                                 total=_cd["total_folders"]))

# --- History ---
with tab_history:
    conn = open_db(selected_db)
    snaps = list_snapshots(conn)
    conn.close()

    if not snaps:
        st.info(t("history_no_snaps"))
    else:
        c1, c2 = st.columns([3, 1])
        c1.subheader(t("history_title", n=len(snaps)))
        if c2.button(t("snapshot_button"), type="primary",
                     disabled=proc_running
                     or _busy_pid_for(selected_db, _busy_procs) is not None,
                     key="snap_btn_history"):
            start_snapshot(selected_db)
            st.rerun()

        st.dataframe(
            [
                {
                    "id": s["id"],
                    t("history_taken_at"): s["taken_at"],
                    t("files"): f"{(s['total_files'] or 0):,}",
                    t("total_size"): human(s["total_size"] or 0),
                    "label": s.get("label") or "",
                }
                for s in snaps
            ],
            width="stretch", hide_index=True,
        )

        if len(snaps) >= 2:
            st.divider()
            st.subheader(t("diff_title"))
            ids = [s["id"] for s in snaps]
            labels = {s["id"]: f"#{s['id']} — {s['taken_at']}" for s in snaps}
            dc1, dc2 = st.columns(2)
            from_id = dc1.selectbox(
                t("diff_from"), ids[1:],
                format_func=lambda i: labels[i],
                index=0, key="diff_from",
            )
            to_id = dc2.selectbox(
                t("diff_to"), ids,
                format_func=lambda i: labels[i],
                index=0, key="diff_to",
            )
            if from_id == to_id:
                st.warning(t("diff_same"))
            elif st.button(t("diff_compute"), type="primary",
                           key="diff_btn"):
                with st.spinner(t("calculating")):
                    d = diff_snapshots(selected_db, from_id=from_id,
                                       to_id=to_id, top_n=10)
                net = (d["added_bytes"] - d["removed_bytes"]
                       + d["modified_delta_bytes"])
                m1, m2, m3, m4 = st.columns(4)
                m1.metric(t("diff_added"), f"{d['added_count']:,}",
                          f"+{human(d['added_bytes'])}")
                m2.metric(t("diff_removed"), f"{d['removed_count']:,}",
                          f"−{human(d['removed_bytes'])}",
                          delta_color="inverse")
                m3.metric(t("diff_modified"), f"{d['modified_count']:,}",
                          f"{'+' if d['modified_delta_bytes']>=0 else '−'}{human(abs(d['modified_delta_bytes']))}")
                m4.metric(t("diff_net"),
                          f"{'+' if net>=0 else '−'}{human(abs(net))}")

                gc1, gc2 = st.columns(2)
                gc1.subheader(t("diff_top_growth"))
                gc1.dataframe(
                    [{"folder": k + "/", "Δ": human(v)}
                     for k, v in d["top_growth"] if v > 0] or
                    [{"folder": "—", "Δ": "—"}],
                    width="stretch", hide_index=True,
                )
                if any(v < 0 for _, v in d["top_shrink"]):
                    gc2.subheader(t("diff_top_shrink"))
                    gc2.dataframe(
                        [{"folder": k + "/", "Δ": f"−{human(-v)}"}
                         for k, v in d["top_shrink"] if v < 0],
                        width="stretch", hide_index=True,
                    )


# --- Compare ---
with tab_compare:
    # ── cross-drive tag search ───────────────────────────────────────────────
    st.subheader(t("tag_search_title"))
    st.caption(t("tag_search_caption"))

    _ts_all = tags_search("")   # all tagged folders across all drives
    if not _ts_all:
        st.info(t("tag_search_no_tags"))
    else:
        _ts_query = st.text_input(
            t("tag_search_input"),
            placeholder="backup",
            key="tag_search_q",
        )
        _ts_results = tags_search(_ts_query) if _ts_query.strip() else _ts_all
        if _ts_results:
            st.dataframe(
                [{
                    t("tag_search_col_drive"): r["label"],
                    t("tag_search_col_path"): r["rel_path"],
                    t("tag_search_col_tags"): " · ".join(r["tags"]),
                    t("tag_search_col_note"): r.get("note", ""),
                } for r in _ts_results],
                width="stretch", hide_index=True,
            )
        else:
            st.info(t("tag_search_empty"))

    st.divider()

    # ── cross-drive dedupe (all drives at once) ──────────────────────────────
    st.subheader(t("cross_title"))
    st.caption(t("cross_caption"))

    if len(dbs) < 2:
        st.info(t("cross_need_drives"))
    else:
        min_size_mb_x = st.slider(
            t("minimum_mb"), 0, 500, 1, key="xdp_min"
        )
        # built unconditionally: the results block below runs on every rerun from
        # session_state, so it must not depend on the button-click branch.
        db_labels = []
        for _db in dbs:
            _reg = _reg_entries.get(_db.resolve(), {})
            db_labels.append((_db, _reg.get("label", _db.stem)))
        if st.button(t("cross_btn"), type="primary", key="xdp_btn"):
            _min_size_bytes = min_size_mb_x * 1024 * 1024
            with st.spinner(t("calculating")):
                if DX_IS_RUST:
                    _proc = subprocess.run(
                        [*DX_CMD, "cross-dedupe",
                         "--min-size", str(_min_size_bytes),
                         "--json"]
                        + [str(db) for db, _ in db_labels],
                        capture_output=True, text=True,
                    )
                    if _proc.returncode == 0 and _proc.stdout.strip():
                        _xgroups = json.loads(_proc.stdout)
                    else:
                        if _proc.stderr:
                            st.warning(_proc.stderr[:500])
                        _xgroups = cross_dedupe(db_labels, min_size=_min_size_bytes)
                else:
                    _xgroups = cross_dedupe(db_labels, min_size=_min_size_bytes)
            st.session_state["xdp_groups"] = _xgroups

        if "xdp_groups" in st.session_state:
            _xgroups = st.session_state["xdp_groups"]
            if not _xgroups:
                st.info(t("cross_no_results"))
            else:
                # ── metrics ─────────────────────────────────────────────────
                _total_wasted = sum(g["wasted_bytes"] for g in _xgroups)
                _confirmed = sum(1 for g in _xgroups if g["confirmed"])
                _approx = len(_xgroups) - _confirmed
                xc1, xc2, xc3, xc4 = st.columns(4)
                xc1.metric(t("cross_groups"), f"{len(_xgroups):,}")
                xc2.metric(t("cross_wasted"), human(_total_wasted))
                xc3.metric(t("cross_confirmed"), f"{_confirmed:,}")
                xc4.metric(t("cross_approx"), f"{_approx:,}")

                # ── firmlink / no-one-fs warning ─────────────────────────────
                _intra_drives: set[str] = set()
                _intra_groups: list[int] = []
                for _gi2, _g2 in enumerate(_xgroups, 1):
                    if _g2.get("intra_drives"):
                        _intra_drives.update(_g2["intra_drives"])
                        _intra_groups.append(_gi2)

                _no_x_drives: set[str] = {
                    lbl for lbl, opts in read_drive_index_opts(db_labels).items()
                    if not opts.get("one_fs")
                }
                _suspect = _intra_drives | _no_x_drives
                if _suspect:
                    st.info(t("cross_firmlink_warn",
                               drives=", ".join(f"**{d}**" for d in sorted(_suspect))))
                    if _intra_groups:
                        with st.expander(t("cross_firmlink_groups"), expanded=False):
                            st.write(f"Grupos: {_intra_groups[:50]}"
                                     + (f" … +{len(_intra_groups)-50}" if len(_intra_groups) > 50 else ""))

                # ── N×N heatmap ──────────────────────────────────────────────
                st.subheader(t("cross_matrix_title"))
                st.caption(t("cross_matrix_caption"))

                # collect unique drive labels that appear in results
                _drive_set: set[str] = set()
                for _g in _xgroups:
                    for _c in _g["copies"]:
                        _drive_set.add(_c["drive"])
                _drive_order = sorted(_drive_set)
                _didx = {d: i for i, d in enumerate(_drive_order)}
                _n = len(_drive_order)

                # build symmetric matrix (bytes)
                _mat = [[0.0] * _n for _ in range(_n)]
                for _g in _xgroups:
                    _ud = list({_c["drive"] for _c in _g["copies"]})
                    for _i in range(len(_ud)):
                        for _j in range(_i + 1, len(_ud)):
                            _di, _dj = _didx[_ud[_i]], _didx[_ud[_j]]
                            _mat[_di][_dj] += _g["size"]
                            _mat[_dj][_di] += _g["size"]

                # convert to GB, mask diagonal as None (shows as grey)
                _mat_gb = [
                    [None if _i == _j else _mat[_i][_j] / 1e9
                     for _j in range(_n)]
                    for _i in range(_n)
                ]
                _text = [
                    ["" if _i == _j
                     else (f"{_mat[_i][_j]/1e9:.1f} GB"
                           if _mat[_i][_j] > 0 else "0")
                     for _j in range(_n)]
                    for _i in range(_n)
                ]

                import plotly.graph_objects as go
                _fig = go.Figure(go.Heatmap(
                    z=_mat_gb,
                    x=_drive_order,
                    y=_drive_order,
                    colorscale="Blues",
                    showscale=True,
                    colorbar=dict(title="GB"),
                    text=_text,
                    texttemplate="%{text}",
                    hovertemplate=(
                        "%{y} ↔ %{x}<br>%{text}<extra></extra>"
                    ),
                ))
                _fig.update_layout(
                    height=max(300, 80 * _n),
                    margin=dict(l=10, r=10, t=10, b=10),
                    xaxis=dict(side="bottom"),
                )
                st.plotly_chart(_fig, width="stretch")

                # ── flat table + CSV download ─────────────────────────────────
                st.subheader(t("cross_details"))

                import csv, io as _io
                _csv_buf = _io.StringIO()
                _csv_w = csv.writer(_csv_buf)
                _csv_w.writerow(["group", "match", "size_bytes", "size_human",
                                 "drive", "path"])
                _rows = []
                for _i, _g in enumerate(_xgroups, 1):
                    _tag = "=" if _g["confirmed"] else "≈"
                    _has_intra = bool(_g.get("intra_drives"))
                    _intra_set = set(_g.get("intra_drives", []))
                    for _c in _g["copies"]:
                        _drive_label = (_c["drive"] + " ⚠️"
                                        if _c["drive"] in _intra_set else _c["drive"])
                        _match_tag = _tag + (" ⚠️" if _has_intra else "")
                        _row = {
                            t("cross_col_group"): _i,
                            t("cross_col_match"): _match_tag,
                            t("cross_col_size"): human(_g["size"]),
                            t("cross_col_drive"): _drive_label,
                            t("cross_col_path"): _c["path"],
                        }
                        _rows.append(_row)
                        _csv_w.writerow([_i, _match_tag, _g["size"],
                                         human(_g["size"]),
                                         _c["drive"], _c["path"]])

                st.download_button(
                    t("cross_download_csv"),
                    data=_csv_buf.getvalue().encode(),
                    file_name="cross-dedupe.csv",
                    mime="text/csv",
                    key="xdp_csv",
                )
                st.dataframe(
                    _rows[:500 * 10],  # rows are per-copy, groups×copies
                    width="stretch",
                    hide_index=True,
                )
                if len(_xgroups) > 500:
                    st.caption(t("groups_not_shown", n=len(_xgroups) - 500))

    st.divider()

    # ── single-copy: files with no backup (inverse of cross-dedupe) ───────────
    st.subheader(t("sc_title"))
    st.caption(t("sc_caption"))

    if len(dbs) < 2:
        st.info(t("sc_need_drives"))
    else:
        _sc_labels_map = {
            _db: _reg_entries.get(_db.resolve(), {}).get("label", _db.stem)
            for _db in dbs
        }
        _sc_c1, _sc_c2 = st.columns([2, 3])
        _sc_min_mb = _sc_c1.slider(t("minimum_mb"), 0, 500, 50, key="sc_min")
        _sc_scope_opts = [t("sc_scope_all")] + sorted(_sc_labels_map.values())
        _sc_scope = _sc_c2.selectbox(
            t("sc_scope"), _sc_scope_opts, key="sc_scope_sel"
        )

        if st.button(t("sc_btn"), type="primary", key="sc_run_btn"):
            _sc_db_labels = [(_db, lbl) for _db, lbl in _sc_labels_map.items()]
            _sc_target = None if _sc_scope == t("sc_scope_all") else _sc_scope
            with st.spinner(t("calculating")):
                st.session_state["sc_result"] = single_copy_files(
                    _sc_db_labels,
                    min_size=_sc_min_mb * 1024 * 1024,
                    target_label=_sc_target,
                    max_files=200000,
                )

        if "sc_result" in st.session_state:
            _sc = st.session_state["sc_result"]
            if _sc.get("insufficient"):
                st.warning(t("sc_insufficient",
                             drives=", ".join(_sc["drives"]) or "—"))
            elif _sc["at_risk_count"] == 0:
                st.success(t("sc_no_results"))
            else:
                _scm1, _scm2, _scm3 = st.columns(3)
                _scm1.metric(t("sc_metric_items"), f"{_sc['at_risk_count']:,}")
                _scm2.metric(t("sc_metric_bytes"), human(_sc["at_risk_bytes"]))
                _scm3.metric(t("sc_metric_drives"), f"{len(_sc['drives'])}")

                if _sc["per_drive"]:
                    st.caption(t("sc_per_drive"))
                    st.dataframe(
                        [{t("sc_col_drive"): _d,
                          t("sc_col_bytes"): human(_v["bytes"]),
                          t("sc_col_count"): _v["count"]}
                         for _d, _v in sorted(_sc["per_drive"].items(),
                                              key=lambda kv: -kv[1]["bytes"])],
                        width="stretch", hide_index=True,
                    )

                if _sc["by_folder"]:
                    st.subheader(t("sc_by_folder"))
                    st.dataframe(
                        [{t("sc_col_drive"): _f["drive"],
                          t("sc_col_folder"): _f["folder"] + "/",
                          t("sc_col_bytes"): human(_f["bytes"]),
                          t("sc_col_count"): _f["count"]}
                         for _f in _sc["by_folder"]],
                        width="stretch", hide_index=True,
                    )

                # per-file list + full CSV export
                st.subheader(t("sc_files_title", n=min(len(_sc["at_risk"]), 2000)))
                import csv as _sc_csv, io as _sc_io
                _sc_buf = _sc_io.StringIO()
                _sc_w = _sc_csv.writer(_sc_buf)
                _sc_w.writerow(["drive", "path", "size_bytes",
                                "size_human", "internal_copies"])
                for _r in _sc["at_risk"]:
                    _sc_w.writerow([_r["drive"], _r["path"], _r["size"],
                                    human(_r["size"]), _r["internal_copies"]])
                st.download_button(
                    t("sc_download_csv"), data=_sc_buf.getvalue().encode(),
                    file_name="single-copy.csv", mime="text/csv", key="sc_csv",
                )
                st.dataframe(
                    [{t("sc_col_drive"): _r["drive"],
                      t("sc_col_path"): _r["path"],
                      t("sc_col_bytes"): human(_r["size"]),
                      t("sc_col_copies"): _r["internal_copies"]}
                     for _r in _sc["at_risk"][:2000]],
                    width="stretch", hide_index=True,
                )
                if len(_sc["at_risk"]) > 2000:
                    st.caption(t("sc_truncated", shown=2000,
                                 total=_sc["at_risk_count"]))

                # ── generate a backup script for the at-risk files ────────────
                st.divider()
                st.markdown(f"**{t('bkp_title')}**")
                st.caption(t("bkp_caption"))
                _bt1, _bt2 = st.columns([3, 1])
                _bkp_default = ("/Volumes/" if sys.platform == "darwin"
                                else "" if os.name == "nt" else "/media/")
                _bkp_target = _bt1.text_input(t("bkp_target"),
                                              value=_bkp_default, key="bkp_target")
                _bkp_shell = _bt2.selectbox(t("bkp_shell"),
                                            ["rsync (.sh)", "robocopy (.bat)"],
                                            key="bkp_shell")
                if _bkp_target.strip():
                    _shell = "bat" if "bat" in _bkp_shell else "sh"
                    _db_labels_b = [(_db, _reg_entries.get(_db.resolve(), {})
                                     .get("label", _db.stem)) for _db in dbs]
                    _script = generate_backup_script(_sc, _db_labels_b,
                                                     _bkp_target.strip(), _shell)
                    st.download_button(
                        t("bkp_download"), data=_script.encode(),
                        file_name=f"backup-at-risk.{'bat' if _shell=='bat' else 'sh'}",
                        mime="text/plain", key="bkp_dl")
                    with st.expander(t("bkp_preview"), expanded=False):
                        st.code(_script[:4000]
                                + ("\n… (truncado)" if len(_script) > 4000 else ""),
                                language="bash")

    st.divider()

    # ── pair comparison ──────────────────────────────────────────────────────
    other_dbs = [d for d in dbs if d != selected_db]
    if not other_dbs:
        st.info(t("need_two_drives"))
    else:
        other = st.selectbox(
            t("compare_with"), other_dbs,
            format_func=lambda p: p.stem,
        )
        min_size_mb_c = st.slider(
            t("minimum_mb"), 0, 500, 1, key="cmp_min"
        )
        min_size_c = min_size_mb_c * 1024 * 1024

        if st.button(t("compare_button"), type="primary", key="cmp_btn"):
            # warn if the two indexes used different partial-hash algorithms
            ca_v = open_db(selected_db)
            cb_v = open_db(other)
            va = get_hash_version(ca_v); vb = get_hash_version(cb_v)
            ca_v.close(); cb_v.close()
            if va != vb:
                st.warning(t("hash_version_mismatch",
                             va=va, vb=vb, cur=HASH_VERSION))
            with st.spinner(t("crosschecking")):
                b_index: dict[tuple[int, str], list[tuple[str, str | None]]] = defaultdict(list)
                cb = open_db(other)
                for size, partial, rel, fh in cb.execute(
                    "SELECT size, partial_hash, rel_path, full_hash FROM entries"
                    " WHERE is_dir=0 AND size >= ? AND partial_hash IS NOT NULL",
                    (min_size_c,),
                ):
                    b_index[(size, partial)].append((rel, fh))
                cb.close()

                matches = []
                only_a = 0
                ca = open_db(selected_db)
                for size, partial, rel_a, fh_a in ca.execute(
                    "SELECT size, partial_hash, rel_path, full_hash FROM entries"
                    " WHERE is_dir=0 AND size >= ? AND partial_hash IS NOT NULL",
                    (min_size_c,),
                ):
                    hits = b_index.get((size, partial))
                    if not hits:
                        only_a += 1
                        continue
                    for rel_b, fh_b in hits:
                        if fh_a and fh_b:
                            if fh_a != fh_b:
                                continue
                            tag = "="
                        else:
                            tag = "≈"
                        matches.append({
                            t("match_col"): tag,
                            t("size_match_col"): human(size),
                            t("in_drive", label=info["label"]): rel_a,
                            t("in_drive", label=Path(other).stem): rel_b,
                            "_bytes": size,
                        })
                ca.close()

            matches.sort(key=lambda m: -m["_bytes"])
            total_bytes = sum(m["_bytes"] for m in matches)
            confirmed = sum(1 for m in matches if m[t("match_col")] == "=")

            c1, c2, c3, c4 = st.columns(4)
            c1.metric(t("matches"), f"{len(matches):,}")
            c2.metric(t("confirmed_eq"), f"{confirmed:,}")
            c3.metric(t("only_in_a"), f"{only_a:,}")
            c4.metric(t("matching_size"), human(total_bytes))

            display = [
                {k: v for k, v in m.items() if k != "_bytes"}
                for m in matches[:1000]
            ]
            st.dataframe(
                display, width="stretch", hide_index=True,
            )
            if len(matches) > 1000:
                st.caption(t("matches_not_shown", n=len(matches) - 1000))
