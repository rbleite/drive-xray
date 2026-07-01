#!/usr/bin/env python3
"""drive-xray — index drives and find duplicate files/folders.

Subcommands:
  index <root> [--db PATH] [--label NAME] [--full]
      Walk <root>, record size/mtime/partial-hash of every file. Stores a
      SQLite "x-ray" of the drive. With --full also computes the full BLAKE2b
      hash of every file (slow, but needed for offline cross-drive compare).

  dedupe <db> [--min-size BYTES] [--dirs-only|--files-only]
      Find duplicates inside one indexed drive. Computes full hashes lazily
      on candidates, then folds duplicate files into duplicate folders via a
      Merkle hash. Requires the drive to still be mounted at its root.

  compare <db_a> <db_b> [--min-size BYTES]
      Cross-drive comparison from the stored x-rays. Matches files by
      (size, partial-hash); if both sides have full_hash, confirms exactly.

Hashing strategy: BLAKE2b (stdlib). Partial-hash = first 64 KiB + last 64 KiB
of the file plus its size — extremely unlikely to collide for unrelated files
of the same size, and ~constant time regardless of file size.
"""
from __future__ import annotations

import argparse
import concurrent.futures
import datetime
import hashlib
import json
import os
import shutil
import sqlite3
import stat
import sys
import time
from collections import defaultdict
from pathlib import Path

PARTIAL_CHUNK = 64 * 1024
READ_CHUNK = 1024 * 1024
# Bump when partial_hash() algorithm changes. Stored per drive in `drive.hash_version`.
# v1 = head + tail; v2 = head + middle + tail (defends against bio formats like
# BAM/VCF where header/footer are stable but body varies).
HASH_VERSION = 2
DX_VERSION = "1.0.0"
SCHEMA_VERSION = 5  # see _migrate_to_v5 / SCHEMA constant below
SKIP_DIR_NAMES = {
    ".Spotlight-V100", ".Trashes", ".fseventsd", ".TemporaryItems",
    ".DocumentRevisions-V100", ".PKInstallSandboxManager",
    ".PKInstallSandboxManager-SystemSoftware", ".HFS+ Private Directory Data",
    ".vol", "System Volume Information", "$RECYCLE.BIN",
}

# Cloud sync folder patterns. Matched case-insensitively against directory
# names; `startswith` so variants like "OneDrive - Acme Corp" also match.
CLOUD_DIR_PREFIXES = (
    "cloudstorage",        # ~/Library/CloudStorage/* (macOS Monterey+ hub)
    "mobile documents",    # ~/Library/Mobile Documents (iCloud Drive)
    "icloud drive",
    "onedrive",            # also OneDrive - <Tenant>
    "google drive",
    "googledrive",
    "dropbox",
    "pclouddrive",
    "box sync", "box-box",
    "creative cloud files",
    "mega", "megasync",
    "tresorit",
    "proton drive", "protondrive",
)


def is_cloud_dir(name: str) -> bool:
    n = name.lower()
    return any(n.startswith(p) for p in CLOUD_DIR_PREFIXES)


_INT64_MAX = 0x7FFFFFFFFFFFFFFF
_UINT64_MOD = 0x10000000000000000


def _i64(n: int | None) -> int | None:
    """SQLite INTEGER is signed 64-bit. macOS reports st_ino / st_dev as
    Python ints that may exceed 2^63-1 on exFAT/NTFS/some APFS volumes
    (raises OverflowError on insert). Wrap to signed range while preserving
    uniqueness (a→a-2^64 is a bijection)."""
    if n is None:
        return None
    if n > _INT64_MAX:
        return n - _UINT64_MOD
    return n


# ---------- hashing ----------

def partial_hash(path: Path, size: int) -> bytes | None:
    """v2: BLAKE2b of size + first 64KB + middle 64KB + last 64KB.

    Returns 16 raw bytes, or None on OSError. Empty files get a deterministic
    digest of just the size header so they all compare equal.
    """
    h = hashlib.blake2b(digest_size=16)
    h.update(size.to_bytes(8, "little"))
    if size == 0:
        return h.digest()
    try:
        with path.open("rb") as f:
            if size <= 3 * PARTIAL_CHUNK:
                h.update(f.read())
            else:
                h.update(f.read(PARTIAL_CHUNK))                   # head
                f.seek(size // 2 - PARTIAL_CHUNK // 2)            # middle
                h.update(f.read(PARTIAL_CHUNK))
                f.seek(-PARTIAL_CHUNK, os.SEEK_END)               # tail
                h.update(f.read(PARTIAL_CHUNK))
    except OSError:
        return None
    return h.digest()


def full_hash(path: Path) -> bytes | None:
    """BLAKE2b of the full file. Returns 32 raw bytes, or None on OSError."""
    h = hashlib.blake2b(digest_size=32)
    try:
        with path.open("rb") as f:
            while chunk := f.read(READ_CHUNK):
                h.update(chunk)
    except OSError:
        return None
    return h.digest()


# ---------- db schema ----------

# Schema v5 — path interning ("Tier 3"):
#   * new `paths(id, parent_id, segment)` table — each unique directory or
#     file name lives once. Common prefixes ("Users/rleite/...") shared.
#   * `entries.path_id` references `paths(id)`. `rel_path` stays as a
#     denormalized cache column so existing queries keep working unchanged,
#     but the heavy `UNIQUE INDEX (snapshot_id, rel_path)` is replaced by
#     `UNIQUE INDEX (snapshot_id, path_id)` — int+int rather than int+text.
#     On a 770k-file db this typically cuts ~20–25% off the total .db size.
#   * `paths` is **global** across snapshots: same string ⇒ same path_id,
#     which also makes `diff` cleaner (path_id equality is a fast int join).
SCHEMA = """
CREATE TABLE IF NOT EXISTS drive (
    id INTEGER PRIMARY KEY,
    label TEXT,
    root_path TEXT NOT NULL,
    indexed_at TEXT NOT NULL,
    total_files INTEGER,
    total_dirs INTEGER,
    total_size INTEGER,
    hash_version INTEGER,
    opt_one_fs INTEGER,
    opt_skip_cloud INTEGER
);
CREATE TABLE IF NOT EXISTS snapshots (
    id INTEGER PRIMARY KEY,
    taken_at TEXT NOT NULL,
    label TEXT,
    total_files INTEGER,
    total_dirs INTEGER,
    total_size INTEGER,
    hash_version INTEGER,
    opt_one_fs INTEGER,
    opt_skip_cloud INTEGER
);
CREATE TABLE IF NOT EXISTS paths (
    id INTEGER PRIMARY KEY,
    parent_id INTEGER REFERENCES paths(id),
    segment TEXT NOT NULL
);
CREATE UNIQUE INDEX IF NOT EXISTS idx_paths_parent_seg
    ON paths(parent_id, segment);
CREATE TABLE IF NOT EXISTS entries (
    id INTEGER PRIMARY KEY,
    snapshot_id INTEGER NOT NULL REFERENCES snapshots(id),
    rel_path TEXT NOT NULL,
    path_id INTEGER REFERENCES paths(id),
    parent_id INTEGER REFERENCES entries(id),
    is_dir INTEGER NOT NULL,
    size INTEGER,
    mtime REAL,
    partial_hash BLOB,
    full_hash BLOB,
    is_symlink INTEGER DEFAULT 0,
    error TEXT,
    inode INTEGER,
    device INTEGER
);
CREATE UNIQUE INDEX IF NOT EXISTS idx_snap_path_id
    ON entries(snapshot_id, path_id);
CREATE INDEX IF NOT EXISTS idx_snap_parent ON entries(snapshot_id, parent_id);
CREATE INDEX IF NOT EXISTS idx_snap_size_partial
    ON entries(snapshot_id, size, partial_hash) WHERE is_dir=0;
CREATE INDEX IF NOT EXISTS idx_full ON entries(full_hash);
CREATE INDEX IF NOT EXISTS idx_snap_inode
    ON entries(snapshot_id, inode, device);
"""


def _migrate_to_v3(conn: sqlite3.Connection) -> bool:
    """If `entries` is still v2 (has `name`/`parent_path` columns, hex hashes),
    rebuild it to v3. Returns True if a migration ran."""
    cols = {r[1] for r in conn.execute("PRAGMA table_info(entries)")}
    needs_migration = ("name" in cols) or ("parent_path" in cols)
    if not needs_migration:
        return False

    # empty-file partial hash: blake2b(size=0). Used to convert "EMPTY" strings.
    _eh = hashlib.blake2b(digest_size=16)
    _eh.update((0).to_bytes(8, "little"))
    empty_blob = _eh.digest()

    def hex_to_blob(s):
        if s is None:
            return None
        if s == "EMPTY":
            return empty_blob
        if s.startswith("ERR:"):
            return None  # store as NULL; original error string is lost
        try:
            return bytes.fromhex(s)
        except (ValueError, AttributeError):
            return None

    print("  migrating .db schema to v3 (compact paths + BLOB hashes)...",
          file=sys.stderr)
    t0 = time.time()

    # ensure v2-era columns are present so the SELECT below doesn't fail
    if "inode" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN inode INTEGER")
    if "device" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN device INTEGER")

    conn.executescript("""
        DROP TABLE IF EXISTS entries_v3_new;
        CREATE TABLE entries_v3_new (
            id INTEGER PRIMARY KEY,
            rel_path TEXT NOT NULL UNIQUE,
            parent_id INTEGER REFERENCES entries_v3_new(id),
            is_dir INTEGER NOT NULL,
            size INTEGER,
            mtime REAL,
            partial_hash BLOB,
            full_hash BLOB,
            is_symlink INTEGER DEFAULT 0,
            error TEXT,
            inode INTEGER,
            device INTEGER
        );
    """)

    # Pass 1: copy rows with parent_id=NULL (preserving ids), in batches
    BATCH = 10000
    batch = []
    n = 0
    for row in conn.execute(
        "SELECT id, rel_path, is_dir, size, mtime, partial_hash, full_hash,"
        " is_symlink, error, inode, device FROM entries"
    ):
        (oid, rp, isdir, sz, mt, ph, fh, sl, er, ino, dev) = row
        batch.append((oid, rp, isdir, sz, mt,
                      hex_to_blob(ph), hex_to_blob(fh),
                      sl or 0, er, ino, dev))
        if len(batch) >= BATCH:
            conn.executemany(
                "INSERT INTO entries_v3_new (id, rel_path, parent_id, is_dir,"
                " size, mtime, partial_hash, full_hash, is_symlink, error,"
                " inode, device) VALUES (?,?,NULL,?,?,?,?,?,?,?,?,?)",
                batch,
            )
            n += len(batch); batch.clear()
            print(f"\r    copied {n} rows", end="", file=sys.stderr)
    if batch:
        conn.executemany(
            "INSERT INTO entries_v3_new (id, rel_path, parent_id, is_dir,"
            " size, mtime, partial_hash, full_hash, is_symlink, error,"
            " inode, device) VALUES (?,?,NULL,?,?,?,?,?,?,?,?,?)",
            batch,
        )
        n += len(batch)

    # Pass 2: link parents. parent_path was the rel_path of the parent dir
    # (literal "." for root's children); look up by rel_path → id.
    rel_to_id = {rp: rid for rp, rid in
                 conn.execute("SELECT rel_path, id FROM entries_v3_new")}
    updates = []
    for oid, pp in conn.execute(
        "SELECT id, parent_path FROM entries WHERE parent_path IS NOT NULL"
    ):
        pid = rel_to_id.get(pp)
        if pid is not None:
            updates.append((pid, oid))
    if updates:
        conn.executemany(
            "UPDATE entries_v3_new SET parent_id=? WHERE id=?", updates
        )

    conn.executescript("""
        DROP TABLE entries;
        ALTER TABLE entries_v3_new RENAME TO entries;
        CREATE INDEX IF NOT EXISTS idx_size_partial ON entries(size, partial_hash) WHERE is_dir=0;
        CREATE INDEX IF NOT EXISTS idx_full ON entries(full_hash);
        CREATE INDEX IF NOT EXISTS idx_parent ON entries(parent_id);
        CREATE INDEX IF NOT EXISTS idx_inode ON entries(inode, device);
    """)
    conn.commit()
    print(f"\r    migrated {n} rows in {time.time()-t0:.1f}s",
          file=sys.stderr)
    return True


def _migrate_to_v4(conn: sqlite3.Connection) -> bool:
    """If `entries` is still v3 (no `snapshot_id`), add the `snapshots`
    table, seed one initial snapshot from the `drive` row, and rebuild
    `entries` with `snapshot_id` + composite UNIQUE(snapshot_id, rel_path)."""
    ent_cols = {r[1] for r in conn.execute("PRAGMA table_info(entries)")}
    if not ent_cols:
        return False  # fresh db — SCHEMA will create v4 directly
    if "snapshot_id" in ent_cols:
        return False  # already v4
    # ensure v4 snapshots table exists
    conn.execute("""
        CREATE TABLE IF NOT EXISTS snapshots (
            id INTEGER PRIMARY KEY,
            taken_at TEXT NOT NULL,
            label TEXT,
            total_files INTEGER,
            total_dirs INTEGER,
            total_size INTEGER,
            hash_version INTEGER,
            opt_one_fs INTEGER,
            opt_skip_cloud INTEGER
        )
    """)
    # seed initial snapshot from drive row (if there is one)
    drv = conn.execute(
        "SELECT label, indexed_at, total_files, total_dirs, total_size,"
        " hash_version, opt_one_fs, opt_skip_cloud FROM drive LIMIT 1"
    ).fetchone()
    if drv is None:
        # weird state — entries exist without a drive row. Make a synthetic
        # snapshot so the FK doesn't dangle.
        conn.execute(
            "INSERT INTO snapshots (taken_at, label, hash_version)"
            " VALUES (?, ?, ?)",
            (time.strftime("%Y-%m-%dT%H:%M:%S"), "(unknown)", 1),
        )
    else:
        conn.execute(
            "INSERT INTO snapshots (taken_at, label, total_files, total_dirs,"
            " total_size, hash_version, opt_one_fs, opt_skip_cloud)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (drv[1], drv[0], drv[2], drv[3], drv[4],
             drv[5], drv[6], drv[7]),
        )
    initial_sid = conn.execute(
        "SELECT id FROM snapshots ORDER BY id DESC LIMIT 1"
    ).fetchone()[0]

    print(f"  migrating .db schema to v4 (snapshot support, initial sid={initial_sid})...",
          file=sys.stderr)
    t0 = time.time()
    conn.executescript("""
        DROP TABLE IF EXISTS entries_v4_new;
        CREATE TABLE entries_v4_new (
            id INTEGER PRIMARY KEY,
            snapshot_id INTEGER NOT NULL REFERENCES snapshots(id),
            rel_path TEXT NOT NULL,
            parent_id INTEGER REFERENCES entries_v4_new(id),
            is_dir INTEGER NOT NULL,
            size INTEGER,
            mtime REAL,
            partial_hash BLOB,
            full_hash BLOB,
            is_symlink INTEGER DEFAULT 0,
            error TEXT,
            inode INTEGER,
            device INTEGER
        );
    """)
    # copy with snapshot_id = initial_sid, preserving ids and parent_id
    n = conn.execute(
        "INSERT INTO entries_v4_new (id, snapshot_id, rel_path, parent_id,"
        " is_dir, size, mtime, partial_hash, full_hash, is_symlink, error,"
        " inode, device)"
        " SELECT id, ?, rel_path, parent_id, is_dir, size, mtime,"
        " partial_hash, full_hash, is_symlink, error, inode, device"
        " FROM entries",
        (initial_sid,),
    ).rowcount
    conn.executescript("""
        DROP TABLE entries;
        ALTER TABLE entries_v4_new RENAME TO entries;
        CREATE UNIQUE INDEX IF NOT EXISTS idx_snap_path ON entries(snapshot_id, rel_path);
        CREATE INDEX IF NOT EXISTS idx_snap_parent ON entries(snapshot_id, parent_id);
        CREATE INDEX IF NOT EXISTS idx_snap_size_partial ON entries(snapshot_id, size, partial_hash) WHERE is_dir=0;
        CREATE INDEX IF NOT EXISTS idx_full ON entries(full_hash);
        CREATE INDEX IF NOT EXISTS idx_snap_inode ON entries(snapshot_id, inode, device);
    """)
    conn.commit()
    print(f"    migrated {n} rows in {time.time()-t0:.1f}s", file=sys.stderr)
    return True


def _migrate_to_v5(conn: sqlite3.Connection) -> bool:
    """If `entries` is at v4 (no `path_id` column or the heavy
    `idx_snap_path` UNIQUE index still around), upgrade to v5 path-interning:
        * build the `paths` table from existing `rel_path` strings,
        * populate `entries.path_id`,
        * replace `idx_snap_path` (UNIQUE on text) with `idx_snap_path_id`
          (UNIQUE on int) — the actual space win.

    Idempotent: if `path_id` already populated and `idx_snap_path` is gone,
    returns False without doing work.
    """
    ent_cols = {r[1] for r in conn.execute("PRAGMA table_info(entries)")}
    if not ent_cols:
        return False  # fresh db — SCHEMA will create v5 directly
    if "snapshot_id" not in ent_cols:
        return False  # not v4 yet; v4 migration must run first
    has_path_id = "path_id" in ent_cols
    has_old_idx = any(r[1] == "idx_snap_path" for r in
                       conn.execute("PRAGMA index_list(entries)"))
    if has_path_id and not has_old_idx:
        return False  # already v5

    print("  migrating .db schema to v5 (path interning)...", file=sys.stderr)
    t0 = time.time()

    conn.execute(
        "CREATE TABLE IF NOT EXISTS paths ("
        "id INTEGER PRIMARY KEY, parent_id INTEGER REFERENCES paths(id),"
        " segment TEXT NOT NULL)"
    )
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_paths_parent_seg"
        " ON paths(parent_id, segment)"
    )
    if not has_path_id:
        conn.execute("ALTER TABLE entries ADD COLUMN path_id INTEGER")

    # Build cache rel_path → path_id, inserting as needed.
    # cache[""] is the root path_id (segment=".", parent=NULL).
    cache: dict[str, int] = {}
    cur = conn.cursor()

    def intern(rel_path: str) -> int:
        if rel_path in cache:
            return cache[rel_path]
        if rel_path in (".", ""):
            cur.execute(
                "INSERT INTO paths (parent_id, segment) VALUES (NULL, '.')"
            )
            pid = cur.lastrowid
            cache["."] = pid
            cache[""] = pid
            return pid
        parts = rel_path.split("/")
        parent_id = intern(".")
        for i, seg in enumerate(parts):
            key = "/".join(parts[: i + 1])
            if key in cache:
                parent_id = cache[key]
                continue
            cur.execute(
                "INSERT INTO paths (parent_id, segment) VALUES (?, ?)",
                (parent_id, seg),
            )
            parent_id = cur.lastrowid
            cache[key] = parent_id
        return parent_id

    # Walk every entry exactly once across all snapshots; intern its rel_path.
    n = 0
    for row in conn.execute(
        "SELECT id, rel_path FROM entries WHERE path_id IS NULL"
    ):
        eid, rp = row
        pid = intern(rp)
        cur.execute("UPDATE entries SET path_id=? WHERE id=?", (pid, eid))
        n += 1

    # Now we can safely swap the indexes.
    conn.execute("DROP INDEX IF EXISTS idx_snap_path")
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_snap_path_id"
        " ON entries(snapshot_id, path_id)"
    )
    conn.commit()
    print(f"    interned {n} entries in {time.time()-t0:.1f}s",
          file=sys.stderr)
    return True


QUARANTINE_DIR = Path.home() / ".drive-xray-quarantine"
AUDIT_LOG = Path.home() / ".config" / "drive-xray" / "audit.jsonl"


def _append_audit(record: dict) -> None:
    """Append one JSON line to the persistent audit log; silently ignored on error."""
    try:
        AUDIT_LOG.parent.mkdir(parents=True, exist_ok=True)
        with AUDIT_LOG.open("a", encoding="utf-8") as _f:
            _f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except Exception:
        pass


def verify_file(root_path: Path, rel_path: str, expected_size: int) -> dict:
    """Check a file still exists on disk with the expected size.

    Returns {"ok": bool, "full_path": str, "reason": str | None}.
    `reason` is None on success; otherwise "not_found" or "size_mismatch …".
    """
    full = root_path / rel_path
    try:
        st_result = full.stat()
    except FileNotFoundError:
        return {"ok": False, "full_path": str(full), "reason": "not_found"}
    except Exception as exc:
        return {"ok": False, "full_path": str(full), "reason": str(exc)}
    if st_result.st_size != expected_size:
        return {
            "ok": False,
            "full_path": str(full),
            "reason": (
                f"size_mismatch (db={expected_size}, disk={st_result.st_size})"
            ),
        }
    return {"ok": True, "full_path": str(full), "reason": None}


def execute_file_action(
    full_path: str,
    action: str,
    root_path: Path | None = None,
    db_path: str = "",
) -> dict:
    """Move a file to quarantine or delete it permanently.

    action: "quarantine" | "delete"
    root_path: when supplied, the target must be inside this directory (path
               containment check prevents acting on files outside the drive).
    Returns {"ok": bool, "full_path": str, "dest": str | None, "error": str | None}.
    Every completed action (ok or error) is appended to AUDIT_LOG.
    """
    # Use the original path for all file operations (do NOT resolve — symlinks in
    # /Volumes/... on macOS resolve to internal paths that break the containment check
    # and would operate on the symlink *target* instead of the entry itself).
    p = Path(full_path)

    # Containment check — os.path.abspath normalises without following symlinks,
    # so /Volumes/DriveName/file stays under /Volumes/DriveName even on macOS.
    if root_path is not None:
        p_abs   = Path(os.path.abspath(full_path))
        root_abs = Path(os.path.abspath(str(root_path)))
        try:
            p_abs.relative_to(root_abs)
        except ValueError:
            result = {
                "ok": False, "full_path": full_path, "dest": None,
                "error": f"path_outside_root: {full_path!r} not under {root_path!r}",
            }
            _append_audit({
                "ts": datetime.datetime.now().isoformat(),
                "action": action, "src": full_path, "dest": None,
                "db": db_path, **{k: result[k] for k in ("ok", "error")},
            })
            return result

    if not p.exists() and not p.is_symlink():
        result = {"ok": False, "full_path": full_path, "dest": None,
                  "error": "not_found"}
        _append_audit({
            "ts": datetime.datetime.now().isoformat(),
            "action": action, "src": full_path, "dest": None,
            "db": db_path, "ok": False, "error": "not_found",
        })
        return result

    try:
        if action == "quarantine":
            QUARANTINE_DIR.mkdir(parents=True, exist_ok=True)
            dest = QUARANTINE_DIR / p.name
            if dest.exists():
                dest = QUARANTINE_DIR / f"{p.stem}_{int(time.time())}{p.suffix}"
            shutil.move(str(p), str(dest))
            result = {"ok": True, "full_path": full_path, "dest": str(dest), "error": None}
        else:
            p.unlink()
            result = {"ok": True, "full_path": full_path, "dest": None, "error": None}
    except Exception as exc:
        result = {"ok": False, "full_path": full_path, "dest": None, "error": str(exc)}

    _append_audit({
        "ts": datetime.datetime.now().isoformat(),
        "action": action,
        "src": full_path,
        "dest": result.get("dest"),
        "db": db_path,
        "ok": result["ok"],
        "error": result["error"],
    })
    return result


# ---------- central registry ----------

REGISTRY_PATH = Path.home() / ".config" / "drive-xray" / "registry.json"
CONFIG_PATH = Path.home() / ".config" / "drive-xray" / "config.json"
_DEFAULT_DB_DIR = Path.home() / "tools" / "drive-xray"


def _registry_load() -> dict:
    if REGISTRY_PATH.exists():
        try:
            return json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"version": 1, "drives": {}}


def _registry_save(data: dict) -> None:
    REGISTRY_PATH.parent.mkdir(parents=True, exist_ok=True)
    REGISTRY_PATH.write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def registry_register(db_path: Path, label: str, root: Path) -> None:
    """Record a drive db in the central registry."""
    data = _registry_load()
    data["drives"][str(db_path.resolve())] = {
        "label": label,
        "root": str(root),
        "last_indexed": datetime.datetime.now().isoformat(timespec="seconds"),
    }
    _registry_save(data)


def registry_list() -> list[dict]:
    """Return all registered drives (sorted by label). Each dict has:
    db (Path), label, root, last_indexed, exists (bool)."""
    data = _registry_load()
    result = []
    for db_str, meta in data.get("drives", {}).items():
        db = Path(db_str)
        result.append({
            "db": db,
            "label": meta.get("label", db.stem),
            "root": meta.get("root", ""),
            "last_indexed": meta.get("last_indexed", ""),
            "exists": db.exists(),
        })
    result.sort(key=lambda x: x["label"].lower())
    return result


def registry_remove(db_path: Path) -> None:
    """Remove a drive from the registry (called on delete)."""
    data = _registry_load()
    key = str(db_path.resolve())
    if key in data.get("drives", {}):
        del data["drives"][key]
        _registry_save(data)


# ---------- folder tags ----------

def tags_get(db_path: Path) -> dict[str, list[str]]:
    """Return {rel_path: [tag, ...]} for all tagged folders in this db."""
    data = _registry_load()
    key = str(db_path.resolve())
    return dict(data.get("drives", {}).get(key, {}).get("folder_tags", {}))


def tags_set(db_path: Path, rel_path: str, tags: list[str]) -> None:
    """Set tags for a folder (replaces existing). Empty list removes the entry."""
    data = _registry_load()
    key = str(db_path.resolve())
    drive = data.setdefault("drives", {}).setdefault(key, {})
    ft = drive.setdefault("folder_tags", {})
    if tags:
        ft[rel_path] = [t.strip() for t in tags if t.strip()]
    else:
        ft.pop(rel_path, None)
    if not ft:
        del drive["folder_tags"]
    _registry_save(data)


def tags_search(query: str) -> list[dict]:
    """Search tagged folders across all registered drives.

    Matches query (case-insensitive) against tag names AND folder paths.
    Returns list of {label, db, rel_path, tags} sorted by drive label.
    """
    q = query.strip().lower()
    data = _registry_load()
    results = []
    for db_str, meta in data.get("drives", {}).items():
        label = meta.get("label", Path(db_str).stem)
        for rel_path, tags in meta.get("folder_tags", {}).items():
            if q in rel_path.lower() or any(q in tag.lower() for tag in tags):
                results.append({
                    "label": label,
                    "db": Path(db_str),
                    "rel_path": rel_path,
                    "tags": tags,
                })
    results.sort(key=lambda r: (r["label"].lower(), r["rel_path"]))
    return results


# ---------- config ----------

def read_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def write_config(data: dict) -> None:
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def get_db_dir() -> Path:
    """Return the configured .db directory, falling back to the default."""
    raw = read_config().get("db_dir")
    if raw:
        p = Path(raw).expanduser()
        if p.is_dir() or not p.exists():
            return p
    return _DEFAULT_DB_DIR


def import_folder(folder: Path) -> list[dict]:
    """Scan `folder` for .db files and register each in the central registry.

    Returns list of {db, label, root, already_registered} for every valid
    .db found. Files that can't be opened (corrupt, wrong format) are skipped.
    """
    folder = folder.expanduser().resolve()
    if not folder.is_dir():
        return []
    existing = {e["db"].resolve() for e in registry_list()}
    results = []
    for db_file in sorted(folder.glob("*.db")):
        db_file = db_file.resolve()
        already = db_file in existing
        try:
            conn = open_db(db_file)
            row = conn.execute(
                "SELECT label, root_path FROM drive LIMIT 1"
            ).fetchone()
            conn.close()
        except Exception:
            continue
        if not row:
            continue
        label, root = row[0], Path(row[1])
        if not already:
            registry_register(db_file, label, root)
        results.append({
            "db": db_file,
            "label": label,
            "root": root,
            "already_registered": already,
        })
    return results


def open_db(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    # Existing tables may be v1/v2/v3/v4; migrate forward in sequence.
    _migrate_to_v3(conn)
    _migrate_to_v4(conn)
    _migrate_to_v5(conn)
    conn.executescript(SCHEMA)
    # in case `drive` was created without these (very old .db)
    drv_cols = {r[1] for r in conn.execute("PRAGMA table_info(drive)")}
    for col in ("hash_version", "opt_one_fs", "opt_skip_cloud"):
        if col not in drv_cols:
            conn.execute(f"ALTER TABLE drive ADD COLUMN {col} INTEGER")
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA foreign_keys=ON")
    # Streamlit may issue concurrent reads while a `dx index` subprocess
    # holds the writer. With WAL, readers don't block writers, but a
    # second writer (e.g. cancelling and re-launching `index`) needs to
    # wait. 5 s covers normal contention; longer locks indicate a bug.
    conn.execute("PRAGMA busy_timeout=5000")
    return conn


def intern_path(conn: sqlite3.Connection, rel_path: str,
                cache: dict[str, int]) -> int:
    """Return the path_id for `rel_path`, inserting any missing ancestors.
    `cache` is a session-local dict (rel_path → path_id) — share it across
    a single walk to avoid repeated lookups."""
    if rel_path in cache:
        return cache[rel_path]
    cur = conn.cursor()
    if rel_path in (".", ""):
        row = conn.execute(
            "SELECT id FROM paths WHERE parent_id IS NULL AND segment='.'"
        ).fetchone()
        if row:
            pid = row[0]
        else:
            cur.execute(
                "INSERT INTO paths (parent_id, segment) VALUES (NULL, '.')"
            )
            pid = cur.lastrowid
        cache["."] = pid
        cache[""] = pid
        return pid
    parts = rel_path.split("/")
    parent_id = intern_path(conn, ".", cache)
    for i, seg in enumerate(parts):
        key = "/".join(parts[: i + 1])
        if key in cache:
            parent_id = cache[key]
            continue
        # Try INSERT; on conflict the UNIQUE index returns the existing row.
        try:
            cur.execute(
                "INSERT INTO paths (parent_id, segment) VALUES (?, ?)",
                (parent_id, seg),
            )
            parent_id = cur.lastrowid
        except sqlite3.IntegrityError:
            row = conn.execute(
                "SELECT id FROM paths WHERE parent_id=? AND segment=?",
                (parent_id, seg),
            ).fetchone()
            parent_id = row[0]
        cache[key] = parent_id
    return parent_id


# ---------- snapshot helpers ----------

def latest_snapshot_id(conn: sqlite3.Connection) -> int | None:
    row = conn.execute("SELECT MAX(id) FROM snapshots").fetchone()
    return row[0] if row else None


def list_snapshots(conn: sqlite3.Connection) -> list[dict]:
    cols = ("id", "taken_at", "label", "total_files",
            "total_dirs", "total_size")
    return [dict(zip(cols, r)) for r in conn.execute(
        f"SELECT {','.join(cols)} FROM snapshots ORDER BY id DESC"
    )]


def insert_snapshot(conn: sqlite3.Connection, label: str | None,
                    hash_version: int, one_fs: bool, skip_cloud: bool) -> int:
    cur = conn.execute(
        "INSERT INTO snapshots (taken_at, label, hash_version, opt_one_fs,"
        " opt_skip_cloud) VALUES (?,?,?,?,?)",
        (time.strftime("%Y-%m-%dT%H:%M:%S"), label,
         hash_version, int(one_fs), int(skip_cloud)),
    )
    return cur.lastrowid


def update_snapshot_totals(conn: sqlite3.Connection, sid: int,
                           total_files: int, total_dirs: int,
                           total_size: int) -> None:
    conn.execute(
        "UPDATE snapshots SET total_files=?, total_dirs=?, total_size=?"
        " WHERE id=?",
        (total_files, total_dirs, total_size, sid),
    )


def get_hash_version(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT hash_version FROM drive LIMIT 1").fetchone()
    return row[0] if row and row[0] else 1


# ---------- indexing ----------

def index_drive(root: Path, db_path: Path, label: str | None, do_full: bool,
                one_fs: bool = False, skip_cloud: bool = False,
                reuse_old: dict | None = None,
                mode: str = "fresh", target_snapshot_id: int | None = None) -> None:
    """Index `root` into `db_path`.

    mode:
      "fresh"      — wipe everything (entries, snapshots, drive) and start
                     over. Creates snapshot id=1.
      "snapshot"   — preserve existing snapshots; create a new one and write
                     entries into it. Used by snapshot_drive().
      "refresh"    — overwrite the latest snapshot in place. Uses
                     `target_snapshot_id` if given, otherwise the most recent.

    `reuse_old` carries forward partial_hash/full_hash for files matched by
    (size, mtime). Used by both snapshot and refresh paths.
    """
    root = root.resolve()
    if not root.is_dir():
        sys.exit(f"not a directory: {root}")
    conn = open_db(db_path)

    if mode == "fresh":
        conn.execute("DELETE FROM entries")
        conn.execute("DELETE FROM snapshots")
        conn.execute("DELETE FROM drive")
        snap_id = insert_snapshot(conn, label, HASH_VERSION,
                                  one_fs, skip_cloud)
    elif mode == "snapshot":
        snap_id = insert_snapshot(conn, label, HASH_VERSION,
                                  one_fs, skip_cloud)
    elif mode == "refresh":
        snap_id = target_snapshot_id or latest_snapshot_id(conn)
        if snap_id is None:
            # nothing to refresh → fall back to fresh
            conn.execute("DELETE FROM entries")
            conn.execute("DELETE FROM drive")
            snap_id = insert_snapshot(conn, label, HASH_VERSION,
                                      one_fs, skip_cloud)
        else:
            conn.execute("DELETE FROM entries WHERE snapshot_id=?", (snap_id,))
            conn.execute(
                "UPDATE snapshots SET taken_at=?, label=?, hash_version=?,"
                " opt_one_fs=?, opt_skip_cloud=? WHERE id=?",
                (time.strftime("%Y-%m-%dT%H:%M:%S"), label, HASH_VERSION,
                 int(one_fs), int(skip_cloud), snap_id),
            )
    else:
        sys.exit(f"unknown index mode: {mode}")

    cur = conn.cursor()

    root_dev = root.stat().st_dev if one_fs else None
    # Track visited directory (inode, device) pairs to detect APFS firmlinks.
    # macOS firmlinks expose the same directory tree under two paths
    # (e.g. /Users/… and /System/Volumes/Data/Users/…) with identical
    # st_dev, so --one-filesystem cannot stop the double-count. Tracking
    # inodes catches this regardless of whether -x is active.
    _root_stat = root.stat()
    _visited_dir_inodes: set[tuple[int, int]] = {
        (_root_stat.st_ino, _root_stat.st_dev)
    }
    firmlinks_skipped = 0
    crossed = 0   # count of pruned subtrees on other filesystems
    cloud_skipped = 0  # count of pruned cloud-sync subtrees
    reused = 0    # files whose hashes were carried over from reuse_old

    total_files = total_dirs = total_size = 0
    t0 = time.time()
    last_print = t0

    # parent_id_by_rel maps rel_path-of-a-dir → its row id, so children can
    # cheaply set parent_id without a SELECT.
    parent_id_by_rel: dict[str, int] = {}
    # path interning cache: rel_path → path_id (in the global `paths` table).
    path_id_cache: dict[str, int] = {}

    def insert(rel_path: str, parent_id: int | None, is_dir: bool,
               size: int | None, mtime: float | None,
               partial: bytes | None, full: bytes | None,
               is_symlink: bool, error: str | None,
               inode: int | None = None, device: int | None = None) -> int:
        path_id = intern_path(conn, rel_path, path_id_cache)
        cur.execute(
            "INSERT OR REPLACE INTO entries"
            " (snapshot_id, rel_path, path_id, parent_id, is_dir, size, mtime,"
            "  partial_hash, full_hash, is_symlink, error, inode, device)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (snap_id, rel_path, path_id, parent_id, int(is_dir), size, mtime,
             partial, full, int(is_symlink), error,
             _i64(inode), _i64(device)),
        )
        return cur.lastrowid

    # root entry
    root_id = insert(".", None, True, None, None, None, None, False, None)
    parent_id_by_rel["."] = root_id

    for dirpath, dirnames, filenames in os.walk(root, followlinks=False, onerror=lambda e: None):
        # prune skip dirs in-place
        dirnames[:] = [d for d in dirnames if d not in SKIP_DIR_NAMES]
        # prune cloud-sync folders
        if skip_cloud:
            before = len(dirnames)
            dirnames[:] = [d for d in dirnames if not is_cloud_dir(d)]
            cloud_skipped += before - len(dirnames)
        dp = Path(dirpath)
        # prune entries on other filesystems (mount points)
        if one_fs:
            kept = []
            for d in dirnames:
                try:
                    if (dp / d).lstat().st_dev == root_dev:
                        kept.append(d)
                    else:
                        crossed += 1
                except OSError:
                    pass
            dirnames[:] = kept
        # prune APFS firmlinks: skip any subdir whose (inode, device) was
        # already visited — these are the same physical directory reachable
        # via two paths (e.g. /Users and /System/Volumes/Data/Users).
        # Skipped on Windows: st_dev is 0 for all NTFS entries, making the
        # (inode, device) key unreliable and causing false positives.
        if os.name != 'nt':
            kept2 = []
            for d in dirnames:
                try:
                    _dstat = (dp / d).lstat()
                    _dkey = (_dstat.st_ino, _dstat.st_dev)
                    if _dkey in _visited_dir_inodes:
                        firmlinks_skipped += 1
                    else:
                        _visited_dir_inodes.add(_dkey)
                        kept2.append(d)
                except OSError:
                    kept2.append(d)
            dirnames[:] = kept2
        rel_dir = str(dp.relative_to(root)) if dp != root else "."
        parent_id = parent_id_by_rel.get(rel_dir)

        for d in dirnames:
            full_p = dp / d
            rel = str(full_p.relative_to(root))
            try:
                st = full_p.lstat()
                # Path.is_symlink() detects NTFS junctions on Windows;
                # stat.S_ISLNK() misses them on Python < 3.12.
                is_link = full_p.is_symlink()
                sub_id = insert(rel, parent_id, True, None, st.st_mtime,
                                None, None, is_link, None,
                                st.st_ino, st.st_dev)
                parent_id_by_rel[rel] = sub_id
                total_dirs += 1
            except OSError as e:
                insert(rel, parent_id, True, None, None, None, None,
                       False, str(e))

        for fn in filenames:
            full_p = dp / fn
            rel = str(full_p.relative_to(root))
            try:
                st = full_p.lstat()
            except OSError as e:
                insert(rel, parent_id, False, None, None, None, None,
                       False, str(e))
                continue
            is_link = full_p.is_symlink()  # handles NTFS junctions on Windows
            if is_link:
                insert(rel, parent_id, False, 0, st.st_mtime, None, None,
                       True, None, st.st_ino, st.st_dev)
                continue
            if not stat.S_ISREG(st.st_mode):
                continue
            size = st.st_size
            partial = full = None
            if reuse_old is not None:
                cached = reuse_old.get(rel)
                if cached:
                    old_size, old_mtime, old_partial, old_full = cached
                    # mtime tolerance: 1s handles HFS+ (1s) vs APFS (1ns)
                    if old_size == size and old_mtime is not None and abs(old_mtime - st.st_mtime) < 1.0:
                        partial = old_partial
                        full = old_full
                        reused += 1
            if partial is None:
                partial = partial_hash(full_p, size)
            if do_full and not full:
                full = full_hash(full_p)
            insert(rel, parent_id, False, size, st.st_mtime, partial, full,
                   False, None, st.st_ino, st.st_dev)
            total_files += 1
            total_size += size

            now = time.time()
            if now - last_print > 1.0:
                rate = total_files / (now - t0) if now > t0 else 0
                print(f"\r  {total_files:>8} files  {total_size/1e9:7.2f} GB  ({rate:.0f}/s)", end="", file=sys.stderr)
                last_print = now
                conn.commit()

    update_snapshot_totals(conn, snap_id, total_files, total_dirs, total_size)
    if mode == "fresh":
        conn.execute(
            "INSERT INTO drive (label, root_path, indexed_at, total_files,"
            " total_dirs, total_size, hash_version, opt_one_fs, opt_skip_cloud)"
            " VALUES (?,?,?,?,?,?,?,?,?)",
            (label or root.name, str(root), time.strftime("%Y-%m-%dT%H:%M:%S"),
             total_files, total_dirs, total_size, HASH_VERSION,
             int(one_fs), int(skip_cloud)),
        )
    else:
        # snapshot / refresh: update drive's "current" totals + indexed_at
        conn.execute(
            "UPDATE drive SET indexed_at=?, total_files=?, total_dirs=?,"
            " total_size=?, hash_version=?, opt_one_fs=?, opt_skip_cloud=?",
            (time.strftime("%Y-%m-%dT%H:%M:%S"), total_files, total_dirs,
             total_size, HASH_VERSION, int(one_fs), int(skip_cloud)),
        )
    conn.commit()
    conn.execute("PRAGMA optimize")  # refresh query-planner statistics after bulk inserts
    conn.close()
    elapsed = time.time() - t0
    extras = []
    if one_fs:
        extras.append(f"pruned {crossed} cross-fs subtrees")
    if firmlinks_skipped:
        extras.append(f"skipped {firmlinks_skipped} firmlink subtrees")
    if skip_cloud:
        extras.append(f"skipped {cloud_skipped} cloud subtrees")
    if reuse_old is not None:
        extras.append(f"reused {reused}/{total_files} cached hashes")
    extra = f" [{'; '.join(extras)}]" if extras else ""
    print(f"\r  indexed {total_files} files / {total_dirs} dirs / {total_size/1e9:.2f} GB in {elapsed:.1f}s{extra}", file=sys.stderr)


# ---------- dedupe ----------

def fill_full_hashes(conn: sqlite3.Connection, root: Path, min_size: int,
                     snapshot_id: int | None = None,
                     workers: int | None = None) -> int:
    """For files sharing (size, partial_hash) within one snapshot, compute
    full_hash using parallel I/O threads.

    workers: thread count (default: min(4, cpu_count)).  HDD users can pass
             workers=1 to avoid seek contention; NVMe users can go higher.
    """
    sid = snapshot_id if snapshot_id is not None else latest_snapshot_id(conn)
    if sid is None:
        return 0
    rows = conn.execute(
        "SELECT size, partial_hash FROM entries"
        " WHERE snapshot_id=? AND is_dir=0 AND error IS NULL"
        "   AND size >= ? AND partial_hash IS NOT NULL"
        " GROUP BY size, partial_hash HAVING COUNT(*) > 1",
        (sid, min_size),
    ).fetchall()
    if not rows:
        return 0
    cur = conn.cursor()
    todo: list[tuple[int, str]] = []
    for size, partial in rows:
        for (rid, rel) in conn.execute(
            "SELECT id, rel_path FROM entries"
            " WHERE snapshot_id=? AND is_dir=0 AND size=? AND partial_hash=?"
            "   AND full_hash IS NULL",
            (sid, size, partial),
        ):
            todo.append((rid, rel))
    if not todo:
        return 0

    n = workers if workers is not None else min(4, os.cpu_count() or 1)
    print(
        f"  computing full hashes for {len(todo)} candidate files "
        f"({n} thread{'s' if n != 1 else ''})...",
        file=sys.stderr,
    )
    t0 = time.time()
    done = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=n) as pool:
        # GIL is released during file read() — threads overlap I/O wait
        future_to_rid = {pool.submit(full_hash, root / rel): rid
                         for rid, rel in todo}
        for fut in concurrent.futures.as_completed(future_to_rid):
            rid = future_to_rid[fut]
            try:
                h = fut.result()
            except Exception:
                h = None
            cur.execute("UPDATE entries SET full_hash=? WHERE id=?", (h, rid))
            done += 1
            if done % 50 == 0:
                conn.commit()
                print(f"\r    {done}/{len(todo)}  ({time.time()-t0:.0f}s)",
                      end="", file=sys.stderr)

    conn.commit()
    conn.execute("PRAGMA optimize")  # full_hash updates shift index statistics significantly
    print(f"\r    done {len(todo)} hashes in {time.time()-t0:.1f}s", file=sys.stderr)
    return len(todo)


def compute_dir_hashes(conn: sqlite3.Connection,
                       snapshot_id: int | None = None) -> None:
    """Bottom-up Merkle hash for directories within one snapshot.
    Defaults to the latest snapshot."""
    sid = snapshot_id if snapshot_id is not None else latest_snapshot_id(conn)
    if sid is None:
        return
    children: dict[int, list[tuple[str, int, bytes | None, int]]] = defaultdict(list)
    for rid, rel, parent_id, is_dir, fh in conn.execute(
        "SELECT id, rel_path, parent_id, is_dir, full_hash FROM entries"
        " WHERE snapshot_id=?",
        (sid,),
    ):
        if parent_id is None:
            continue
        name = os.path.basename(rel) or rel
        children[parent_id].append((name, is_dir, fh, rid))

    dir_ids = [r[0] for r in conn.execute(
        "SELECT id FROM entries WHERE snapshot_id=? AND is_dir=1"
        " ORDER BY length(rel_path) DESC",
        (sid,),
    )]
    cur = conn.cursor()
    resolved: dict[int, bytes | None] = {}

    def hash_for(dir_id: int) -> bytes | None:
        if dir_id in resolved:
            return resolved[dir_id]
        kids = sorted(children.get(dir_id, []), key=lambda x: x[0])
        if not kids:
            resolved[dir_id] = None
            return None
        h = hashlib.blake2b(digest_size=32)
        for name, is_dir, fh, kid_id in kids:
            if is_dir:
                sub = hash_for(kid_id)
                if sub is None:
                    resolved[dir_id] = None
                    return None
                h.update(b"D")
                h.update(name.encode("utf-8", "replace"))
                h.update(sub)
            else:
                if fh is None:
                    resolved[dir_id] = None
                    return None
                h.update(b"F")
                h.update(name.encode("utf-8", "replace"))
                h.update(fh)
        digest = h.digest()
        resolved[dir_id] = digest
        return digest

    for did in dir_ids:
        h = hash_for(did)
        cur.execute("UPDATE entries SET full_hash=? WHERE id=?", (h, did))
    conn.commit()


def human(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024:
            return f"{n:.1f}{unit}"
        n /= 1024
    return f"{n:.1f}PB"


def dedupe(db_path: Path, min_size: int, mode: str,
           workers: int | None = None) -> None:
    conn = open_db(db_path)
    drive = conn.execute("SELECT root_path, label FROM drive LIMIT 1").fetchone()
    if not drive:
        sys.exit("db has no drive record — re-run `index` first")
    root = Path(drive[0])
    sid = latest_snapshot_id(conn)
    if sid is None:
        sys.exit("no snapshots — run `index` first")
    if mode != "dirs-only":
        if not root.exists():
            print(f"  warning: root {root} not mounted — using existing full_hash only", file=sys.stderr)
        else:
            fill_full_hashes(conn, root, min_size, snapshot_id=sid, workers=workers)

    if mode in ("both", "dirs-only"):
        compute_dir_hashes(conn, snapshot_id=sid)

    print(f"\n=== Duplicate files (drive: {drive[1]}, snapshot {sid}) ===")
    if mode != "dirs-only":
        groups = conn.execute(
            "SELECT full_hash, COUNT(*) c, SUM(size) s FROM entries"
            " WHERE snapshot_id=? AND is_dir=0 AND full_hash IS NOT NULL"
            "   AND size >= ?"
            " GROUP BY full_hash HAVING c > 1 ORDER BY s DESC",
            (sid, min_size),
        ).fetchall()
        total_wasted = 0
        total_hardlinks = 0
        for fh, count, _total in groups:
            rows = conn.execute(
                "SELECT rel_path, size, inode, device FROM entries"
                " WHERE snapshot_id=? AND full_hash=? AND is_dir=0",
                (sid, fh),
            ).fetchall()
            size = rows[0][1]
            inodes = {(ino, dev) for _, _, ino, dev in rows
                      if ino is not None and dev is not None}
            # if any inode info is missing, fall back to counting all rows
            distinct = len(inodes) if inodes else count
            wasted = size * (distinct - 1)
            hardlinks_here = count - distinct if inodes else 0
            total_wasted += wasted
            total_hardlinks += hardlinks_here
            tag = f"  ({hardlinks_here} hardlink{'s' if hardlinks_here != 1 else ''})" if hardlinks_here else ""
            print(f"\n[{count}× {human(size)}] wasted={human(wasted)}{tag}  hash={fh.hex()[:12]}")
            seen_inodes: dict[tuple, str] = {}
            for rel, _, ino, dev in rows:
                key = (ino, dev) if ino is not None else None
                if key and key in seen_inodes:
                    print(f"    {rel}  [↳ hardlink to {seen_inodes[key]}]")
                else:
                    if key:
                        seen_inodes[key] = rel
                    print(f"    {rel}")
        print(f"\ntotal wasted by duplicate files: {human(total_wasted)}"
              + (f"  (excluding {total_hardlinks} hardlinks)" if total_hardlinks else ""))

    if mode in ("both", "dirs-only"):
        print(f"\n=== Duplicate folders (drive: {drive[1]}, snapshot {sid}) ===")
        dir_groups = conn.execute(
            "SELECT full_hash, COUNT(*) c FROM entries"
            " WHERE snapshot_id=? AND is_dir=1 AND full_hash IS NOT NULL"
            " GROUP BY full_hash HAVING c > 1 ORDER BY c DESC",
            (sid,),
        ).fetchall()
        for fh, count in dir_groups:
            paths = [r for (r,) in conn.execute(
                "SELECT rel_path FROM entries"
                " WHERE snapshot_id=? AND full_hash=? AND is_dir=1",
                (sid, fh),
            )]
            print(f"\n[{count}× identical folder]  hash={fh.hex()[:12]}")
            for p in paths:
                print(f"    {p}/")
    conn.close()


# ---------- refresh (incremental) ----------

def _read_drive_and_cache(db_path: Path) -> tuple[Path, str, bool, bool, dict, int | None]:
    """Common setup for refresh and snapshot: read drive metadata + build
    the (size, mtime) → hash cache from the latest snapshot."""
    conn = open_db(db_path)
    drv = conn.execute(
        "SELECT root_path, label, opt_one_fs, opt_skip_cloud FROM drive LIMIT 1"
    ).fetchone()
    if not drv:
        sys.exit("db has no drive record — run `index` first")
    root_path, label, opt_one_fs, opt_skip_cloud = drv
    root = Path(root_path)
    if not root.is_dir():
        sys.exit(f"root {root} not mounted or not a directory")
    # Default one_fs=True on macOS/Linux (meaningful); False on Windows where
    # st_dev is 0 for all NTFS entries and the check would cross or block all mounts.
    one_fs = bool(opt_one_fs) if opt_one_fs is not None else (os.name != 'nt')
    skip_cloud = bool(opt_skip_cloud) if opt_skip_cloud is not None else True
    latest_sid = latest_snapshot_id(conn)
    old: dict[str, tuple] = {}
    if latest_sid is not None:
        for r in conn.execute(
            "SELECT rel_path, size, mtime, partial_hash, full_hash FROM entries"
            " WHERE snapshot_id=? AND is_dir=0 AND error IS NULL"
            "       AND is_symlink=0 AND partial_hash IS NOT NULL",
            (latest_sid,),
        ):
            old[r[0]] = (r[1], r[2], r[3], r[4])
    conn.close()
    return root, label, one_fs, skip_cloud, old, latest_sid


def refresh_drive(db_path: Path, do_full: bool = False) -> None:
    """Re-walk the original root and overwrite the latest snapshot in place.
    Reuses partial/full hashes for files unchanged by (size, mtime).
    Historical snapshots are preserved untouched."""
    root, label, one_fs, skip_cloud, old, sid = _read_drive_and_cache(db_path)
    print(f"  refresh: reusing {len(old)} cached entries from snapshot {sid}",
          file=sys.stderr)
    index_drive(root, db_path, label, do_full,
                one_fs=one_fs, skip_cloud=skip_cloud, reuse_old=old,
                mode="refresh", target_snapshot_id=sid)


def snapshot_drive(db_path: Path, do_full: bool = False,
                   auto_prune: bool = True,
                   keep_last: int = 10, keep_monthly: int = 12) -> int:
    """Re-walk the original root and create a *new* snapshot (preserving
    previous ones). Reuses hashes from the latest snapshot for files
    unchanged by (size, mtime). Returns the new snapshot id.

    If auto_prune is True, runs prune_snapshots() after with the given
    retention policy."""
    root, label, one_fs, skip_cloud, old, prev_sid = _read_drive_and_cache(db_path)
    print(f"  snapshot: reusing {len(old)} cached entries from snapshot {prev_sid}",
          file=sys.stderr)
    index_drive(root, db_path, label, do_full,
                one_fs=one_fs, skip_cloud=skip_cloud, reuse_old=old,
                mode="snapshot")
    # the snapshot just created is the new latest
    conn = open_db(db_path)
    new_sid = latest_snapshot_id(conn)
    if auto_prune:
        pruned = prune_snapshots(conn, keep_last=keep_last,
                                 keep_monthly=keep_monthly)
        if pruned:
            print(f"  prune: dropped {len(pruned)} old snapshot(s) "
                  f"({pruned})", file=sys.stderr)
    conn.commit()
    conn.close()
    return new_sid


# ---------- snapshot retention + diff ----------

def prune_snapshots(conn: sqlite3.Connection,
                    keep_last: int = 10,
                    keep_monthly: int = 12) -> list[int]:
    """Apply the retention policy: keep the `keep_last` most recent snapshots
    plus one per calendar month for the last `keep_monthly` months. Returns
    the list of pruned snapshot ids."""
    from datetime import datetime, timedelta, timezone
    snaps = list(conn.execute(
        "SELECT id, taken_at FROM snapshots ORDER BY taken_at DESC, id DESC"
    ))
    if not snaps:
        return []
    keep: set[int] = set()
    for sid, _ in snaps[:keep_last]:
        keep.add(sid)
    seen_months: set[tuple[int, int]] = set()
    cutoff = datetime.now() - timedelta(days=31 * keep_monthly)
    for sid, ts in snaps[keep_last:]:
        try:
            dt = datetime.fromisoformat(ts)
        except ValueError:
            keep.add(sid)  # unparseable → keep (defensive)
            continue
        if dt.tzinfo is not None:
            dt = dt.astimezone().replace(tzinfo=None)
        if dt < cutoff:
            continue
        month = (dt.year, dt.month)
        if month not in seen_months:
            seen_months.add(month)
            keep.add(sid)
    to_delete = [sid for sid, _ in snaps if sid not in keep]
    if to_delete:
        placeholders = ",".join("?" * len(to_delete))
        conn.execute(
            f"DELETE FROM entries WHERE snapshot_id IN ({placeholders})",
            to_delete,
        )
        conn.execute(
            f"DELETE FROM snapshots WHERE id IN ({placeholders})",
            to_delete,
        )
    return to_delete


def diff_snapshots(db_path: Path, from_id: int | None = None,
                   to_id: int | None = None, top_n: int = 10) -> dict:
    """Diff two snapshots by file. If ids are None, defaults to
    (previous, latest). Returns counts and per-folder rollups."""
    conn = open_db(db_path)
    snaps = list_snapshots(conn)
    if not snaps:
        sys.exit("no snapshots in db")
    if to_id is None:
        to_id = snaps[0]["id"]
    if from_id is None:
        if len(snaps) < 2:
            sys.exit("need at least 2 snapshots to diff (only 1 found)")
        from_id = snaps[1]["id"]
    if from_id == to_id:
        sys.exit("from and to are the same snapshot")
    a, b = from_id, to_id

    # v5 schema: path_id is a global int per unique path — same path across
    # snapshots has the same path_id. Use it for faster int joins instead of
    # text comparisons on rel_path.
    added = conn.execute(
        "SELECT b.rel_path, b.size FROM entries b"
        " WHERE b.snapshot_id=? AND b.is_dir=0"
        "   AND NOT EXISTS ("
        "     SELECT 1 FROM entries a WHERE a.snapshot_id=? AND a.path_id=b.path_id AND a.is_dir=0"
        "   )",
        (b, a),
    ).fetchall()
    removed = conn.execute(
        "SELECT a.rel_path, a.size FROM entries a"
        " WHERE a.snapshot_id=? AND a.is_dir=0"
        "   AND NOT EXISTS ("
        "     SELECT 1 FROM entries b WHERE b.snapshot_id=? AND b.path_id=a.path_id AND b.is_dir=0"
        "   )",
        (a, b),
    ).fetchall()
    modified = conn.execute(
        "SELECT b.rel_path, a.size, b.size FROM entries b"
        " JOIN entries a ON a.path_id=b.path_id"
        " WHERE b.snapshot_id=? AND a.snapshot_id=? AND b.is_dir=0 AND a.is_dir=0"
        "   AND (b.size != a.size OR b.partial_hash IS NOT a.partial_hash)",
        (b, a),
    ).fetchall()

    def top_folder_key(rel: str, depth: int = 2) -> str:
        parts = rel.split("/")
        return "/".join(parts[:depth]) if len(parts) > depth else parts[0]

    growth: dict[str, int] = defaultdict(int)
    addcount: dict[str, int] = defaultdict(int)
    for rp, sz in added:
        k = top_folder_key(rp)
        growth[k] += sz or 0
        addcount[k] += 1
    for rp, sz in removed:
        k = top_folder_key(rp)
        growth[k] -= sz or 0
        addcount[k] -= 1
    for rp, old_sz, new_sz in modified:
        k = top_folder_key(rp)
        growth[k] += (new_sz or 0) - (old_sz or 0)

    snap_a = next(s for s in snaps if s["id"] == a)
    snap_b = next(s for s in snaps if s["id"] == b)
    conn.close()

    return {
        "from": snap_a,
        "to": snap_b,
        "added_count": len(added),
        "added_bytes": sum(sz or 0 for _, sz in added),
        "removed_count": len(removed),
        "removed_bytes": sum(sz or 0 for _, sz in removed),
        "modified_count": len(modified),
        "modified_delta_bytes": sum((nsz or 0) - (osz or 0)
                                    for _, osz, nsz in modified),
        "top_growth": sorted(growth.items(), key=lambda kv: -kv[1])[:top_n],
        "top_shrink": sorted(growth.items(), key=lambda kv: kv[1])[:top_n],
        "top_count": sorted(addcount.items(), key=lambda kv: -kv[1])[:top_n],
    }


def print_diff(d: dict) -> None:
    fa = d["from"]; tb = d["to"]
    print(f"\n=== diff: snapshot #{fa['id']} ({fa['taken_at']}) "
          f"→ #{tb['id']} ({tb['taken_at']}) ===\n")
    print(f"  + {d['added_count']:>8,} files  +{human(d['added_bytes'])}")
    print(f"  − {d['removed_count']:>8,} files  −{human(d['removed_bytes'])}")
    delta = d['modified_delta_bytes']
    print(f"  ~ {d['modified_count']:>8,} modified  (size Δ "
          f"{'+' if delta >= 0 else ''}{human(abs(delta))})")
    net = d['added_bytes'] - d['removed_bytes'] + d['modified_delta_bytes']
    print(f"  ─────────────────────────────────")
    print(f"  net size change: {'+' if net >= 0 else '−'}{human(abs(net))}")
    print(f"\nTop folders by growth:")
    for k, v in d["top_growth"]:
        if v <= 0: break
        print(f"  +{human(v):>8}   {k}/")
    if any(v < 0 for _, v in d["top_shrink"]):
        print(f"\nTop folders by shrink:")
        for k, v in d["top_shrink"]:
            if v >= 0: break
            print(f"  −{human(-v):>8}   {k}/")


# ---------- export ----------

def _duplicate_rows(conn: sqlite3.Connection, min_size: int,
                    snapshot_id: int | None = None) -> list[dict]:
    """Flat row representation of duplicate file groups for export.
    Single JOIN query instead of N+1 (one query per group)."""
    sid = snapshot_id if snapshot_id is not None else latest_snapshot_id(conn)
    if sid is None:
        return []
    # Single JOIN: fetch all members of all duplicate groups at once.
    all_rows = conn.execute(
        "SELECT e.full_hash, g.cnt, e.rel_path, e.size, e.inode, e.device"
        " FROM entries e"
        " JOIN ("
        "   SELECT full_hash, COUNT(*) cnt FROM entries"
        "   WHERE snapshot_id=? AND is_dir=0 AND full_hash IS NOT NULL AND size>=?"
        "   GROUP BY full_hash HAVING cnt > 1"
        " ) g ON e.full_hash = g.full_hash"
        " WHERE e.snapshot_id=? AND e.is_dir=0"
        " ORDER BY e.full_hash, e.rel_path",
        (sid, min_size, sid),
    ).fetchall()

    # Group in Python (already ordered by full_hash)
    by_hash: dict[bytes, list] = defaultdict(list)
    for fh, cnt, rel, size, ino, dev in all_rows:
        by_hash[bytes(fh)].append((cnt, rel, size, ino, dev))

    rows: list[dict] = []
    for group_id, (fh, members) in enumerate(by_hash.items(), 1):
        count = members[0][0]
        size = members[0][2]
        hash_hex = fh.hex()
        distinct = {(ino, dev) for _, _, _, ino, dev in members if ino is not None}
        d = len(distinct) if distinct else count
        wasted = size * (d - 1)
        seen: set[tuple] = set()
        for _, rel, _, ino, dev in members:
            key = (ino, dev) if ino is not None else None
            is_hl = key is not None and key in seen
            if key is not None:
                seen.add(key)
            rows.append({
                "group_id": group_id,
                "hash": hash_hex,
                "size_bytes": size,
                "size_human": human(size),
                "group_count": count,
                "distinct_inodes": d,
                "wasted_bytes": wasted,
                "wasted_human": human(wasted),
                "path": rel,
                "is_hardlink": int(is_hl),
            })
    # sort: largest waste first, then alphabetical within group
    rows.sort(key=lambda r: (-r["wasted_bytes"], r["group_id"], r["path"]))
    return rows


def export_duplicates(db_path: Path, out_path: Path, fmt: str,
                      min_size: int, workers: int | None = None) -> None:
    conn = open_db(db_path)
    drv = conn.execute("SELECT root_path FROM drive LIMIT 1").fetchone()
    if drv:
        root = Path(drv[0])
        if root.is_dir():
            fill_full_hashes(conn, root, min_size, workers=workers)
        else:
            print(f"  warning: root {root} not mounted — using existing"
                  " full_hash only", file=sys.stderr)
    rows = _duplicate_rows(conn, min_size)
    conn.close()
    if not rows:
        print("  no duplicate groups found above min-size — nothing to export",
              file=sys.stderr)
        return

    if fmt == "csv":
        import csv
        with out_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
    elif fmt == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            sys.exit("openpyxl not installed. Install with: pip install openpyxl")
        wb = Workbook()
        ws = wb.active
        ws.title = "Duplicates"
        headers = list(rows[0].keys())
        ws.append(headers)
        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="EEEEEE")
        for cell in ws[1]:
            cell.font = bold
            cell.fill = fill
        for r in rows:
            ws.append([r[h] for h in headers])
        # column widths
        widths = {"path": 60, "hash": 28, "size_human": 12, "wasted_human": 14}
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = widths.get(h, 14)
        ws.freeze_panes = "A2"
        wb.save(out_path)
    else:
        sys.exit(f"unknown export format: {fmt}")
    print(f"  wrote {len(rows)} rows to {out_path}", file=sys.stderr)


# ---------- folder size aggregation (treemap) ----------

def compute_folder_sizes(entries: list[tuple]) -> dict[int, int]:
    """Given iterable of (id, parent_id, is_dir, size), return id → cumulative
    size dict. Folder size = sum of all descendant file sizes. Iterative DFS
    so it handles deep trees without recursion-limit issues."""
    children: dict[int, list[int]] = defaultdict(list)
    is_dir_of: dict[int, bool] = {}
    file_size: dict[int, int] = {}
    roots: list[int] = []
    for eid, pid, isdir, sz in entries:
        is_dir_of[eid] = bool(isdir)
        if pid is None:
            roots.append(eid)
        else:
            children[pid].append(eid)
        if not isdir:
            file_size[eid] = sz or 0

    sizes: dict[int, int] = dict(file_size)
    for root in roots:
        # post-order iterative DFS
        stack: list[tuple[int, bool]] = [(root, False)]
        while stack:
            nid, visited = stack.pop()
            if visited:
                if nid not in sizes:
                    sizes[nid] = sum(sizes.get(c, 0) for c in children[nid])
            else:
                stack.append((nid, True))
                for c in children[nid]:
                    if c not in sizes:
                        stack.append((c, False))
    return sizes


# ---------- assisted cleanup ----------

CLEANUP_STRATEGIES = ("shortest", "oldest", "newest", "alphabetical")
CLEANUP_ACTIONS = ("delete", "quarantine")


def generate_cleanup_script(db_path: Path, min_size: int,
                            strategy: str = "shortest",
                            action: str = "quarantine") -> str:
    """Produce a shell script proposing actions on duplicate files. The user
    is expected to review and edit before running it.

    Strategies pick which copy to KEEP:
      shortest      — shortest rel_path (likely the "main" location)
      oldest        — smallest mtime (the original)
      newest        — largest mtime
      alphabetical  — sorted ascending
    """
    if strategy not in CLEANUP_STRATEGIES:
        raise ValueError(f"unknown strategy: {strategy}")
    if action not in CLEANUP_ACTIONS:
        raise ValueError(f"unknown action: {action}")

    conn = open_db(db_path)
    drv = conn.execute(
        "SELECT label, root_path FROM drive LIMIT 1"
    ).fetchone()
    if not drv:
        sys.exit("db has no drive record")
    label, root_path = drv
    sid = latest_snapshot_id(conn)
    if sid is None:
        sys.exit("no snapshots")

    groups = conn.execute(
        "SELECT full_hash, COUNT(*) c FROM entries"
        " WHERE snapshot_id=? AND is_dir=0 AND full_hash IS NOT NULL"
        "   AND size >= ?"
        " GROUP BY full_hash HAVING c > 1",
        (sid, min_size),
    ).fetchall()

    stamp = time.strftime("%Y%m%dT%H%M%S")
    out: list[str] = [
        "#!/usr/bin/env bash",
        f"# drive-xray cleanup plan",
        f"# Drive label : {label}",
        f"# Drive root  : {root_path}",
        f"# Generated   : {time.strftime('%Y-%m-%dT%H:%M:%S')}",
        f"# Strategy    : keep '{strategy}'   (action: {action})",
        f"# Min size    : {human(min_size)} ({min_size} bytes)",
        "#",
        "# Review every line. Comment out (#) anything you want to keep.",
        "# Hardlinks share storage with another path; lines pointing at",
        "# hardlinked-but-already-kept inodes are commented out.",
        "#",
        "# After review, run with:  bash <this-file>",
        "",
        "set -euo pipefail",
        "",
    ]
    if action == "quarantine":
        out += [
            f'QUARANTINE="$HOME/.drive-xray-quarantine/{label}-{stamp}"',
            'mkdir -p "$QUARANTINE"',
            'echo "moving copies to: $QUARANTINE"',
            "",
        ]

    total_freeable = 0
    n_actions = 0
    n_hardlink_notes = 0

    for gnum, (fh, _count) in enumerate(groups, 1):
        members = conn.execute(
            "SELECT rel_path, size, mtime, inode, device FROM entries"
            " WHERE snapshot_id=? AND full_hash=? AND is_dir=0",
            (sid, fh),
        ).fetchall()
        # group by (inode, device) so each physical file is represented once
        by_inode: dict[tuple, list[tuple]] = defaultdict(list)
        for rp, sz, mt, ino, dev in members:
            key = (ino, dev) if ino is not None else (rp, dev)
            by_inode[key].append((rp, sz, mt))

        if len(by_inode) < 2:
            # all paths point to a single physical file (just hardlinks) →
            # no real duplication, skip
            continue

        # one "representative" path per inode (first by rel_path) +
        # mtime + size + the list of hardlink siblings
        reps: list[tuple] = []
        for key, paths in by_inode.items():
            paths_sorted = sorted(paths, key=lambda x: x[0])
            rp0, sz0, mt0 = paths_sorted[0]
            siblings = [p for p in paths_sorted[1:]]
            reps.append((key, rp0, sz0, mt0, siblings))

        # choose keeper
        if strategy == "shortest":
            keeper_idx = min(range(len(reps)), key=lambda i: len(reps[i][1]))
        elif strategy == "oldest":
            keeper_idx = min(range(len(reps)),
                             key=lambda i: reps[i][3] or float("inf"))
        elif strategy == "newest":
            keeper_idx = max(range(len(reps)),
                             key=lambda i: reps[i][3] or 0)
        else:  # alphabetical
            keeper_idx = min(range(len(reps)), key=lambda i: reps[i][1])

        size = reps[0][2]
        out.append(f"# === Group {gnum}: {len(by_inode)} distinct copies"
                   f" of {human(size)}  ·  hash={fh.hex()[:12]} ===")
        keeper_rp = reps[keeper_idx][1]
        out.append(f'#   KEEP   : {keeper_rp}')
        # note hardlinks to the keeper (no action — already retained)
        for sib_rp, _, _ in reps[keeper_idx][4]:
            out.append(f'#   keep↳hl: {sib_rp}  (hardlink to KEEP)')
            n_hardlink_notes += 1

        for i, (key, rp, sz, mt, siblings) in enumerate(reps):
            if i == keeper_idx:
                continue
            full = _shell_quote(f"{root_path}/{rp}")
            if action == "delete":
                out.append(f"rm   {full}  # {human(sz)}")
            else:
                safe = rp.replace("/", "__").replace(" ", "_")
                dst_name = _shell_quote(f"g{gnum:04d}_i{i}__{safe}")
                out.append(f'mv   {full}  "$QUARANTINE"/{dst_name}  # {human(sz)}')
            n_actions += 1
            total_freeable += sz
            for sib_rp, _, _ in siblings:
                sib_full = _shell_quote(f"{root_path}/{sib_rp}")
                out.append(f"#    ↳ hardlink (same inode, no extra space): "
                           f"{sib_full}")
                n_hardlink_notes += 1
        out.append("")

    out += [
        f"# ── Summary ──────────────────────────────────────",
        f"# Actions     : {n_actions}",
        f"# Hardlink notes: {n_hardlink_notes}",
        f"# Reclaimable : ~{human(total_freeable)} ({total_freeable} bytes)",
    ]
    conn.close()
    return "\n".join(out) + "\n"


def _shell_quote(s: str) -> str:
    """Conservative single-quote shell escape."""
    return "'" + s.replace("'", "'\\''") + "'"


# ---------- compact ----------

def compact_db(db_path: Path) -> None:
    """VACUUM + WAL checkpoint to physically shrink the .db file. Also runs
    the v2→v3 migration via open_db if it hasn't happened yet."""
    before = db_path.stat().st_size if db_path.exists() else 0
    wal = Path(str(db_path) + "-wal")
    before_wal = wal.stat().st_size if wal.exists() else 0
    # Step 1: open & migrate. Commit and close cleanly so no locks remain.
    conn = open_db(db_path)
    conn.commit()
    conn.close()
    # Step 2: fresh autocommit connection for checkpoint + VACUUM + checkpoint.
    # VACUUM in WAL mode writes a new copy through the WAL, so we need a
    # second checkpoint after VACUUM to fold it back in and truncate the WAL.
    print("  checkpoint + vacuum + checkpoint...", file=sys.stderr)
    conn = sqlite3.connect(db_path, isolation_level=None)
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    conn.execute("VACUUM")
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")  # fold VACUUM output back in
    conn.close()
    after = db_path.stat().st_size
    after_wal = wal.stat().st_size if wal.exists() else 0
    print(f"  {human(before)} + {human(before_wal)} (wal)"
          f" → {human(after)} + {human(after_wal)} (wal)"
          f"  [-{human(before + before_wal - after - after_wal)}]",
          file=sys.stderr)


# ---------- compare ----------

def compare(db_a: Path, db_b: Path, min_size: int) -> None:
    # use open_db so older .db files migrate to v4 on the fly
    ca = open_db(db_a)
    cb = open_db(db_b)
    da = ca.execute("SELECT label, root_path FROM drive LIMIT 1").fetchone()
    db = cb.execute("SELECT label, root_path FROM drive LIMIT 1").fetchone()
    va = get_hash_version(ca)
    vb = get_hash_version(cb)
    sid_a = latest_snapshot_id(ca)
    sid_b = latest_snapshot_id(cb)
    print(f"A = {da[0]} ({da[1]})  [hash v{va}, snapshot {sid_a}]")
    print(f"B = {db[0]} ({db[1]})  [hash v{vb}, snapshot {sid_b}]")
    if va != vb:
        print(f"\nWARNING: partial-hash versions differ. Matches will be unreliable.\n"
              f"Re-index both drives with the current version (v{HASH_VERSION}).\n",
              file=sys.stderr)
    print()

    # index B by (size, partial)
    b_index: dict[tuple[int, str], list[tuple[str, str | None]]] = defaultdict(list)
    for size, partial, rel, fh in cb.execute(
        "SELECT size, partial_hash, rel_path, full_hash FROM entries"
        " WHERE snapshot_id=? AND is_dir=0 AND size >= ? AND partial_hash IS NOT NULL",
        (sid_b, min_size),
    ):
        b_index[(size, partial)].append((rel, fh))

    matches = 0
    matched_size = 0
    confirmed = 0
    only_a = 0
    for size, partial, rel_a, fh_a in ca.execute(
        "SELECT size, partial_hash, rel_path, full_hash FROM entries"
        " WHERE snapshot_id=? AND is_dir=0 AND size >= ? AND partial_hash IS NOT NULL",
        (sid_a, min_size),
    ):
        hits = b_index.get((size, partial))
        if not hits:
            only_a += 1
            continue
        for rel_b, fh_b in hits:
            tag = "≈"  # likely match
            if fh_a and fh_b:
                if fh_a == fh_b:
                    tag = "="
                    confirmed += 1
                else:
                    continue  # full hashes differ → not a match
            matches += 1
            matched_size += size
            print(f"  {tag} {human(size):>8}  A:{rel_a}\n           B:{rel_b}")
    print(f"\nmatches: {matches} ({human(matched_size)})  confirmed-by-full-hash: {confirmed}")
    print(f"only in A: {only_a} files")
    ca.close(); cb.close()


# ---------- cross-drive dedupe ----------

def cross_dedupe(
    db_labels: list[tuple[Path, str]],
    min_size: int = 1024 * 1024,
) -> list[dict]:
    """Find files duplicated across multiple drives.

    db_labels — list of (db_path, drive_label).
    Returns groups sorted by wasted_bytes desc. Each group:
      {"size": int, "confirmed": bool, "wasted_bytes": int,
       "copies": [{"drive": str, "path": str}, ...]}
    Drives that are not mounted / have no snapshot are silently skipped.
    """
    # (size, partial_hash) → [(rel_path, full_hash_or_None, drive_label)]
    index: dict[tuple[int, bytes], list] = defaultdict(list)

    for db_path, label in db_labels:
        try:
            conn = open_db(db_path)
        except Exception:
            continue
        sid = latest_snapshot_id(conn)
        if sid is None:
            conn.close()
            continue
        # deduplicate by (inode, device) within the same db so APFS firmlinks
        # (e.g. /Users/… == /System/Volumes/Data/Users/…, same inode+dev)
        # don't create false cross-drive duplicates.
        _seen_inodes: set[tuple[int, int]] = set()
        for size, partial, rel, fh, ino, dev in conn.execute(
            "SELECT size, partial_hash, rel_path, full_hash, inode, device"
            " FROM entries"
            " WHERE snapshot_id=? AND is_dir=0 AND size>=? AND partial_hash IS NOT NULL",
            (sid, min_size),
        ):
            if ino is not None and dev is not None:
                key = (int(ino), int(dev))
                if key in _seen_inodes:
                    continue
                _seen_inodes.add(key)
            index[(size, partial)].append((rel, fh, label))
        conn.close()

    groups: list[dict] = []
    for (size, _partial), copies in index.items():
        # skip groups entirely within one drive
        if len({c[2] for c in copies}) < 2:
            continue

        all_have_fh = all(c[1] is not None for c in copies)
        if all_have_fh:
            # group further by exact full_hash
            by_fh: dict[bytes, list] = defaultdict(list)
            for rel, fh, lbl in copies:
                by_fh[fh].append({"drive": lbl, "path": rel})
            for sub in by_fh.values():
                if len({c["drive"] for c in sub}) < 2:
                    continue
                _drive_counts: dict[str, int] = {}
                for _c in sub:
                    _drive_counts[_c["drive"]] = _drive_counts.get(_c["drive"], 0) + 1
                groups.append({
                    "size": size,
                    "confirmed": True,
                    "wasted_bytes": size * (len(sub) - 1),
                    "copies": sub,
                    # drives that appear >1× in same group → likely firmlink
                    "intra_drives": [d for d, n in _drive_counts.items() if n > 1],
                })
        else:
            _drive_counts2: dict[str, int] = {}
            for _c in copies:
                _drive_counts2[_c[2]] = _drive_counts2.get(_c[2], 0) + 1
            groups.append({
                "size": size,
                "confirmed": False,
                "wasted_bytes": size * (len(copies) - 1),
                "copies": [{"drive": c[2], "path": c[0]} for c in copies],
                "intra_drives": [d for d, n in _drive_counts2.items() if n > 1],
            })

    groups.sort(key=lambda g: -g["wasted_bytes"])
    return groups


def read_drive_index_opts(db_labels: list[tuple[Path, str]]) -> dict[str, dict]:
    """Return {label: {"one_fs": bool, "skip_cloud": bool}} from each db's
    drive table. Drives that can't be read are omitted."""
    result: dict[str, dict] = {}
    for db_path, label in db_labels:
        try:
            conn = open_db(db_path)
            row = conn.execute(
                "SELECT opt_one_fs, opt_skip_cloud FROM drive LIMIT 1"
            ).fetchone()
            conn.close()
            if row:
                result[label] = {
                    "one_fs": bool(row[0]),
                    "skip_cloud": bool(row[1]),
                }
        except Exception:
            pass
    return result


# ---------- cli ----------

def main():
    p = argparse.ArgumentParser(prog="drive-xray")
    p.add_argument(
        "--version", action="version",
        version=f"drive-xray {DX_VERSION} "
                f"(schema v{SCHEMA_VERSION} · hash v{HASH_VERSION})",
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    pi = sub.add_parser("index", help="x-ray a drive into a sqlite db")
    pi.add_argument("root", type=Path)
    pi.add_argument("--db", type=Path, help="output .db path (default: ~/tools/drive-xray/<label>.db)")
    pi.add_argument("--label", help="human label for this drive")
    pi.add_argument("--full", action="store_true", help="also compute full BLAKE2b of every file (slow)")
    pi.add_argument("-x", "--one-filesystem", action="store_true",
                    help="don't cross mount points (skips /Volumes/* and APFS firmlinks)")
    pi.add_argument("--skip-cloud", action="store_true",
                    help="skip cloud-sync folders (iCloud, OneDrive, Google Drive, Dropbox, Box, MEGA, ...)")

    pd = sub.add_parser("dedupe", help="find duplicates within an indexed drive")
    pd.add_argument("db", type=Path)
    pd.add_argument("--min-size", type=int, default=1024, help="ignore files smaller than this (bytes)")
    pd.add_argument("--workers", type=int, default=None, metavar="N",
                    help="I/O threads for full hashing (default: min(4, cpu_count))")
    grp = pd.add_mutually_exclusive_group()
    grp.add_argument("--files-only", dest="mode", action="store_const", const="files-only")
    grp.add_argument("--dirs-only", dest="mode", action="store_const", const="dirs-only")
    pd.set_defaults(mode="both")

    pc = sub.add_parser("compare", help="compare two indexed drives")
    pc.add_argument("db_a", type=Path)
    pc.add_argument("db_b", type=Path)
    pc.add_argument("--min-size", type=int, default=1024)

    pr = sub.add_parser("refresh",
                        help="re-index, overwriting the latest snapshot in"
                             " place (reuses hashes of unchanged files)")
    pr.add_argument("db", type=Path)
    pr.add_argument("--full", action="store_true",
                    help="also compute full BLAKE2b for newly hashed files")

    ps = sub.add_parser("snapshot",
                        help="create a new snapshot, preserving previous ones")
    ps_sub = ps.add_subparsers(dest="snap_cmd")
    ps_take = ps_sub.add_parser("take",
                                 help="take a new snapshot (default action)")
    ps_take.add_argument("db", type=Path)
    ps_take.add_argument("--full", action="store_true")
    ps_take.add_argument("--no-prune", action="store_true",
                         help="skip auto-prune after the snapshot")
    ps_take.add_argument("--keep-last", type=int, default=10)
    ps_take.add_argument("--keep-monthly", type=int, default=12)
    ps_list = ps_sub.add_parser("list", help="list snapshots in the db")
    ps_list.add_argument("db", type=Path)
    # default if no subcommand: take
    ps.add_argument("db_default", nargs="?", type=Path,
                    help=argparse.SUPPRESS)

    ppr = sub.add_parser("prune",
                         help="apply retention policy to old snapshots")
    ppr.add_argument("db", type=Path)
    ppr.add_argument("--keep-last", type=int, default=10)
    ppr.add_argument("--keep-monthly", type=int, default=12)

    pdf = sub.add_parser("diff",
                         help="diff two snapshots (default: previous → latest)")
    pdf.add_argument("db", type=Path)
    pdf.add_argument("--from", dest="from_id", type=int)
    pdf.add_argument("--to", dest="to_id", type=int)
    pdf.add_argument("--top", type=int, default=10,
                     help="show top N folders by growth (default 10)")

    pk = sub.add_parser("compact",
                        help="VACUUM + WAL checkpoint to shrink the .db file"
                             " (also runs schema migration if needed)")
    pk.add_argument("db", type=Path)

    pcl = sub.add_parser("cleanup",
                         help="generate a shell script proposing deletes/moves"
                              " for duplicate files (review before running)")
    pcl.add_argument("db", type=Path)
    pcl.add_argument("--strategy", default="shortest",
                     choices=list(CLEANUP_STRATEGIES),
                     help="which copy to KEEP (default: shortest path)")
    pcl.add_argument("--action", default="quarantine",
                     choices=list(CLEANUP_ACTIONS),
                     help="what to do with copies (default: quarantine)")
    pcl.add_argument("--min-size", type=int, default=1048576,
                     help="ignore groups smaller than this (bytes, default 1 MB)")
    pcl.add_argument("-o", "--out", type=Path,
                     help="write to this file (default: stdout)")

    pe = sub.add_parser("export", help="export duplicate groups as CSV or XLSX")
    pe.add_argument("db", type=Path)
    pe.add_argument("out", type=Path, help="output file (.csv or .xlsx)")
    pe.add_argument("--min-size", type=int, default=1024)
    pe.add_argument("--format", choices=("csv", "xlsx"),
                    help="override format (default: inferred from extension)")
    pe.add_argument("--workers", type=int, default=None, metavar="N",
                    help="I/O threads for full hashing (default: min(4, cpu_count))")

    sub.add_parser("drives", help="list all drives registered in the central index")

    pif = sub.add_parser(
        "import-folder",
        help="scan a folder for .db files and register them in the central index",
    )
    pif.add_argument("folder", type=Path, help="folder to scan")

    pcfg = sub.add_parser("config", help="manage drive-xray configuration")
    pcfg_sub = pcfg.add_subparsers(dest="cfg_cmd")
    pcfg_sub.add_parser("show", help="show current configuration")
    pcfg_set = pcfg_sub.add_parser("set-db-dir",
                                   help="set the folder where .db files are saved")
    pcfg_set.add_argument("path", type=Path, help="new db folder path")

    pxd = sub.add_parser(
        "cross-dedupe",
        help="find duplicate files across multiple drives (works offline)",
    )
    pxd.add_argument(
        "dbs", nargs="*", type=Path,
        help="db files to compare (default: all registered drives)",
    )
    pxd.add_argument("--all", dest="use_all", action="store_true",
                     help="use all drives from the central registry")
    pxd.add_argument("--min-size", type=int, default=1024 * 1024,
                     metavar="BYTES", help="ignore files smaller than this (default 1 MB)")

    args = p.parse_args()

    if args.cmd == "index":
        label = args.label or args.root.name
        db = args.db or get_db_dir() / f"{label}.db"
        db.parent.mkdir(parents=True, exist_ok=True)
        opts = []
        if args.one_filesystem: opts.append("one-filesystem")
        if args.skip_cloud: opts.append("skip-cloud")
        suffix = f" ({', '.join(opts)})" if opts else ""
        print(f"indexing {args.root} → {db}{suffix}", file=sys.stderr)
        index_drive(args.root, db, label, args.full,
                    one_fs=args.one_filesystem, skip_cloud=args.skip_cloud)
        # Only auto-register when --db was NOT specified explicitly (i.e. the
        # app chose the path inside DB_DIR). An explicit --db path is managed
        # by the caller (tests, scripts) and must not pollute the registry.
        if args.db is None:
            registry_register(db, label, args.root)
    elif args.cmd == "dedupe":
        dedupe(args.db, args.min_size, args.mode, workers=args.workers)
    elif args.cmd == "compare":
        compare(args.db_a, args.db_b, args.min_size)
    elif args.cmd == "refresh":
        print(f"refreshing {args.db}", file=sys.stderr)
        refresh_drive(args.db, do_full=args.full)
        # update last_indexed in registry (read label/root from db)
        _drv = open_db(args.db).execute(
            "SELECT label, root_path FROM drive LIMIT 1"
        ).fetchone()
        if _drv:
            registry_register(args.db, _drv[0], Path(_drv[1]))
    elif args.cmd == "snapshot":
        # support both "dx snapshot <db>" and "dx snapshot take <db>"
        snap_cmd = getattr(args, "snap_cmd", None) or "take"
        if snap_cmd == "list":
            conn = open_db(args.db)
            snaps = list_snapshots(conn)
            conn.close()
            print(f"  {len(snaps)} snapshot(s) in {args.db.name}:")
            for s in snaps:
                size = human(s["total_size"] or 0)
                print(f"  #{s['id']:>3}  {s['taken_at']}   "
                      f"{(s['total_files'] or 0):>10,} files   {size:>10}"
                      f"   {s.get('label') or ''}")
        else:  # take
            db = args.db if args.db else args.db_default
            if db is None:
                sys.exit("usage: dx snapshot [take|list] <db>")
            print(f"taking snapshot of {db}", file=sys.stderr)
            sid = snapshot_drive(db, do_full=args.full,
                                  auto_prune=not args.no_prune,
                                  keep_last=args.keep_last,
                                  keep_monthly=args.keep_monthly)
            print(f"  new snapshot id: {sid}", file=sys.stderr)
    elif args.cmd == "prune":
        conn = open_db(args.db)
        pruned = prune_snapshots(conn, keep_last=args.keep_last,
                                 keep_monthly=args.keep_monthly)
        conn.commit()
        conn.close()
        if pruned:
            print(f"  pruned {len(pruned)} snapshot(s): {pruned}")
        else:
            print("  no snapshots to prune")
    elif args.cmd == "diff":
        d = diff_snapshots(args.db, args.from_id, args.to_id, top_n=args.top)
        print_diff(d)
    elif args.cmd == "compact":
        print(f"compacting {args.db}", file=sys.stderr)
        compact_db(args.db)
    elif args.cmd == "cleanup":
        script = generate_cleanup_script(
            args.db, args.min_size, args.strategy, args.action,
        )
        if args.out:
            args.out.write_text(script, encoding="utf-8")
            print(f"  wrote cleanup script to {args.out}", file=sys.stderr)
        else:
            sys.stdout.write(script)
    elif args.cmd == "export":
        fmt = args.format or args.out.suffix.lstrip(".").lower()
        if fmt not in ("csv", "xlsx"):
            sys.exit(f"cannot infer format from extension {args.out.suffix!r}; use --format")
        export_duplicates(args.db, args.out, fmt, args.min_size,
                          workers=args.workers)
    elif args.cmd == "import-folder":
        results = import_folder(args.folder)
        new = [r for r in results if not r["already_registered"]]
        already = [r for r in results if r["already_registered"]]
        if not results:
            print(f"  no valid .db files found in {args.folder}")
        else:
            if new:
                print(f"  registered {len(new)} new drive(s):")
                for r in new:
                    print(f"    {r['label']:<20}  {r['db']}")
            if already:
                print(f"  {len(already)} already registered (skipped):")
                for r in already:
                    print(f"    {r['label']:<20}  {r['db']}")
    elif args.cmd == "config":
        cfg_cmd = getattr(args, "cfg_cmd", None) or "show"
        cfg = read_config()
        if cfg_cmd == "show":
            print(f"  db_dir: {cfg.get('db_dir', f'{_DEFAULT_DB_DIR}  (default)')}")
        elif cfg_cmd == "set-db-dir":
            p = args.path.expanduser().resolve()
            p.mkdir(parents=True, exist_ok=True)
            cfg["db_dir"] = str(p)
            write_config(cfg)
            print(f"  db_dir set to: {p}")
    elif args.cmd == "drives":
        entries = registry_list()
        if not entries:
            print("  no drives registered yet — run `dx index` to add one")
        else:
            print(f"  {'LABEL':<20}  {'LAST INDEXED':<20}  {'ROOT':<40}  DB")
            print(f"  {'-'*20}  {'-'*20}  {'-'*40}  {'-'*30}")
            for e in entries:
                status = "" if e["exists"] else " [MISSING]"
                print(
                    f"  {e['label']:<20}  {e['last_indexed']:<20}"
                    f"  {e['root']:<40}  {e['db']}{status}"
                )
    elif args.cmd == "cross-dedupe":
        if args.use_all or not args.dbs:
            reg = registry_list()
            db_labels = [(e["db"], e["label"]) for e in reg if e["exists"]]
        else:
            db_labels = []
            for db in args.dbs:
                try:
                    conn = open_db(db)
                    row = conn.execute(
                        "SELECT label FROM drive LIMIT 1"
                    ).fetchone()
                    conn.close()
                    db_labels.append((db, row[0] if row else db.stem))
                except Exception as exc:
                    print(f"  skip {db}: {exc}", file=sys.stderr)
        if len(db_labels) < 2:
            sys.exit("cross-dedupe needs at least 2 drives. Use --all or pass db files.")
        labels_str = ", ".join(f"{lbl}({db.name})" for db, lbl in db_labels)
        print(f"cross-dedupe across {len(db_labels)} drives: {labels_str}", file=sys.stderr)
        groups = cross_dedupe(db_labels, min_size=args.min_size)
        if not groups:
            print("  no cross-drive duplicates found.")
        else:
            total_wasted = sum(g["wasted_bytes"] for g in groups)
            print(f"  {len(groups)} groups · {human(total_wasted)} wasted\n")
            for i, g in enumerate(groups, 1):
                tag = "=" if g["confirmed"] else "≈"
                print(f"  [{i}] {tag} {human(g['size'])}")
                for c in g["copies"]:
                    print(f"       {c['drive']:<20}  {c['path']}")
                if i >= 200:
                    print(f"  … {len(groups)-200} more groups not shown")
                    break


if __name__ == "__main__":
    main()
